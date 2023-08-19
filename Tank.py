from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QFileDialog, QApplication, QWidget, QTableWidget, \
    QTableWidgetItem, QProxyStyle, QLabel, QScrollArea, QVBoxLayout, QMainWindow, QProgressBar, QFrame, QHeaderView, \
    QTableView, QToolTip, QComboBox, QLineEdit, QSpinBox, QDateEdit, QPushButton, QTabWidget, QGraphicsScene,  QGraphicsProxyWidget
from PyQt5.QtGui import *
from PyQt5.QtGui import QGuiApplication, QPixmap
import numpy as np
import pandas as pd
import math
from scipy.stats import poisson
import numba
import matplotlib.pyplot as plt
import os
import os.path as path
import importlib
import openpyxl
import importlib.util
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from datetime import datetime
from matplotlib.ticker import FormatStrFormatter
from assimulo.problem import Implicit_Problem
from assimulo.solvers import IDA

pd.set_option('display.max_columns', 100)

# get a list of all files in Content/Modules
filelist = os.listdir("neuralNetwork")
my_module = []
nN_name = ['']
current_dir = os.getcwd()
full_file_directory = [current_dir + "\\neuralNetwork\\" + x for x in filelist]
file_value = 0

# import every file ending in .py
for fname in filelist:
    if path.splitext(fname)[1] == '.py':
        test_spec = importlib.util.spec_from_file_location(path.splitext(fname)[0], full_file_directory[file_value])  # load the module
        test_module = importlib.util.module_from_spec(test_spec)
        test_spec.loader.exec_module(test_module)
        my_module.append(test_module)
        nN_name.append(path.splitext(fname)[0])
        file_value += 1

componentList = ['Ethane', 'Methane', 'Propane', 'Nitrogen', 'Hydrogen']

enthalpyLiqList = ['']
enthalpyVapList = ['']
specificheatLiqList = ['']
specificheatVapList = ['']
vapourPressureList = ['']
liquidDensityList = ['']

enthalpyLiqList.extend([x for x in nN_name if x[:2] == 'EL'])
enthalpyVapList.extend([x for x in nN_name if x[:2] == 'EV'])
specificheatLiqList.extend([x for x in nN_name if x[:2] == 'CL'])
specificheatVapList.extend([x for x in nN_name if x[:2] == 'CV'])
vapourPressureList.extend([x for x in nN_name if x[:2] == 'VP'])
liquidDensityList.extend([x for x in nN_name if len(x) > 0 if x[0] == 'D'])

class VerticalLabel(QLabel):

    def __init__(self, *args):
        QLabel.__init__(self, *args)

    def paintEvent(self, event):
        QLabel.paintEvent(self, event)
        painter = QPainter(self)
        painter.rotate(270)
        painter.translate(-1*self.height(), 0)
        QLabel.render(self, painter)


class VerticalLineEdit(QLineEdit):
    def __init__(self, *args):
        QLineEdit.__init__(self, *args)

    def paintEvent(self, event):
        QLineEdit.paintEvent(self, event)
        painter = QPainter(self)
        painter.rotate(270)
        painter.translate(-1*self.height(), 0)
        self.setGeometry(self.x(),self.y(),self.height(),self.width())
        QLineEdit.render(self, painter)
        painter.end()

    def minimumSizeHint(self):
        size = QLabel.minimumSizeHint(self)
        return QtCore.QSize(size.height(), size.width())

    def sizeHint(self):
        size = QLabel.sizeHint(self)
        return QtCore.QSize(size.height(), size.width())



class Ui_MainWindow(QMainWindow):
    global componentList, nN_name, MW_Dict, Ap_dict, Bp_dict, Cp_dict, Dp_dict, Ad_dict, Bd_dict, Cd_dict, \
        enthalpyLiqList, enthalpyVapList, specificheatLiqList, specificheatVapList, vapourPressureList, liquidDensityList

    def __init__(self):
        super(Ui_MainWindow, self).__init__()
        self.setupUi(MainWindow)

    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.setFixedSize(1000, 1000)

        MainWindow.setStyleSheet("background-color: rgb(235, 236, 240);")
        MainWindow.setWindowTitle("Tank Dashboard")

        self.designButton = QPushButton(MainWindow)
        self.designButton.setGeometry(0, 0, 250, 50)
        self.designButton.setObjectName("designButton")
        self.designButton.setText('Design')
        self.designButton.clicked.connect(self.toDesign)
        self.designButton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.designButton.setStyleSheet(
            "QPushButton#designButton{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")

        self.heatLossButton = QPushButton(MainWindow)
        self.heatLossButton.setGeometry(250, 0, 250, 50)
        self.heatLossButton.setObjectName("heatLossButton")
        self.heatLossButton.setText('Heat Leak')
        self.heatLossButton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.heatLossButton.clicked.connect(self.toHeatLoss)
        self.heatLossButton.setStyleSheet(
            "QPushButton#heatLossButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#heatLossButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")

        self.initialButton = QPushButton(MainWindow)
        self.initialButton.setGeometry(500, 0, 250, 50)
        self.initialButton.setObjectName("initialButton")
        self.initialButton.setText('Solver Settings')
        self.initialButton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.initialButton.clicked.connect(self.toInitialConditions)
        self.initialButton.setToolTip("Save Table button needs to be clicked before accessing other tabs.")
        self.initialButton.setStyleSheet(
            "QPushButton#initialButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#initialButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")

        self.resultsButton = QPushButton(MainWindow)
        self.resultsButton.setGeometry(750, 0, 250, 50)
        self.resultsButton.setObjectName("resultsButton")
        self.resultsButton.setText('Results')
        self.resultsButton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.resultsButton.setEnabled(False)
        self.resultsButton.clicked.connect(self.resultsfun)
        self.resultsButton.setStyleSheet(
            "QPushButton#resultsButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#resultsButton::hover{background-color: rgb(255, 0, 0); color: rgb(255, 255, 255); border: 2px solid black}")

        self.progressBar = QLabel(MainWindow)
        self.progressBar.setGeometry(0, 970, 1000, 30)
        self.progressBar.setObjectName("progressBar")
        self.progressBar.setText('Incomplete')
        self.progressBar.setAlignment(QtCore.Qt.AlignCenter)
        self.progressBar.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.progressBar.setStyleSheet(
            "QLabel#progressBar{background-color: rgb(255, 0, 0); color: rgb(0, 0, 0); border: 2px solid black;}")
        self.progressBar.hide()

        self.designlabel = QFrame(MainWindow)
        self.designlabel.setGeometry(0, 50, 1000, 950)
        self.designlabel.show()

        self.complabel = QFrame(MainWindow)
        self.complabel.setGeometry(0, 50, 1000, 950)
        self.complabel.hide()

        self.heatlosslabel = QFrame(MainWindow)
        self.heatlosslabel.setGeometry(0, 50, 1000, 950)
        self.heatlosslabel.hide()

        self.initiallabel = QLabel(MainWindow)
        self.initiallabel.setGeometry(0, 50, 1000, 950)
        self.initiallabel.hide()

        self.resultslabel = QLabel(MainWindow)
        self.resultslabel.setGeometry(0, 50, 1000, 950)
        self.resultslabel.hide()

        self.defaultbutton = QtWidgets.QPushButton(self.designlabel)
        self.defaultbutton.setGeometry(250, 815, 150, 60)
        self.defaultbutton.setObjectName("defaultbutton")
        self.defaultbutton.setText('Load\nBase Case')
        self.defaultbutton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.defaultbutton.setStyleSheet(
            "QPushButton#defaultbutton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#defaultbutton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.defaultbutton.hide()

        self.resetbutton = QtWidgets.QPushButton(self.designlabel)
        self.resetbutton.setGeometry(425, 815, 150, 60)
        self.resetbutton.setObjectName("resetbutton")
        self.resetbutton.setText('Reset')
        self.resetbutton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.resetbutton.setStyleSheet(
            "QPushButton#resetbutton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#resetbutton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        #self.resetbutton.clicked.connect(self.resetfun)
        self.resetbutton.hide()

        self.tanklabel = QLabel(self.designlabel)
        self.tanklabel.setGeometry(250, 50, 500, 790)
        self.tankimage = QPixmap('tank.png')
        self.tanklabel.setPixmap(self.tankimage)
        self.tanklabel.show()

        self.noComponentsLabel = QLabel(self.designlabel)
        self.noComponentsLabel.setGeometry(325, 235, 290, 25)
        self.noComponentsLabel.setObjectName("noComponentsLabel")
        self.noComponentsLabel.setText("Number of Components:")
        self.noComponentsLabel.setFont(QFont('Arial', 15, weight=QtGui.QFont.Bold))

        self.noComponentstext = QComboBox(self.designlabel)
        self.noComponentstext.setGeometry(618, 230, 60, 40)
        self.noComponentstext.setObjectName("noComponentstext")
        self.noComponentstext.addItems(['', '1', '2', '3', '4', '5'])
        self.noComponentstext.setEditable(False)
        self.noComponentstext.currentTextChanged.connect(self.noComponentschanged)
        self.noComponentstext.setStyleSheet("QComboBox#noComponentstext{ background-color: rgb(255, 255, 255);}"
                                            "QComboBox#noComponentstext{ border: 2px solid black;}"
                                            "QComboBox#noComponentstext{ border-radius: 5px;}"
                                            "QComboBox#noComponentstext{ font-size: 15pt; font-family: Arial;}")

        self.initialPressureLabel = QLabel(self.designlabel)
        self.initialPressureLabel.setGeometry(325, 192, 320, 25)
        self.initialPressureLabel.setObjectName("initialPressureLabel")
        self.initialPressureLabel.setText("Initial Tank Pressure (kPa):")
        self.initialPressureLabel.setFont(QFont('Arial', 15, weight=QtGui.QFont.Bold))
        self.initialPressureLabel.show()

        self.initialPressuretext = QLineEdit(self.designlabel)
        self.initialPressuretext.setGeometry(655, 185, 80, 40)
        self.initialPressuretext.setObjectName("initialPressuretext")
        self.initialPressuretext.setFont(QFont('Arial', 15))
        self.initialPressuretext.setStyleSheet("QLineEdit#initialPressuretext{ background-color: rgb(255, 255, 255);}"
                                               "QLineEdit#initialPressuretext{ border: 2px solid black;}"
                                               "QLineEdit#initialPressuretext{ border-radius: 5px;}")
        self.initialPressuretext.show()

        self.MWButton = QtWidgets.QPushButton(self.designlabel)
        self.MWButton.setGeometry(325, 270, 200, 50)
        self.MWButton.setObjectName("MWButton")
        self.MWButton.setText('Molecular Weights')
        self.MWButton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.MWButton.setStyleSheet(
            "QPushButton#MWButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#MWButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.MWButton.clicked.connect(self.MWfun)

        # self.jacketStartLabel = QLabel(self.designlabel)
        # self.jacketStartLabel.setGeometry(325, 802, 320, 25)
        # self.jacketStartLabel.setObjectName("initialPressureLabel")
        # self.jacketStartLabel.setText("Jacket Start Point (%):")
        # self.jacketStartLabel.setFont(QFont('Arial', 15, weight=QtGui.QFont.Bold))
        # self.jacketStartLabel.show()

        self.jacketEndLabel = QLabel(self.designlabel)
        self.jacketEndLabel.setGeometry(325, 330, 320, 25)
        self.jacketEndLabel.setObjectName("initialPressureLabel")
        self.jacketEndLabel.setText("Jacket End Point (%):")
        self.jacketEndLabel.setFont(QFont('Arial', 15, weight=QtGui.QFont.Bold))
        self.jacketEndLabel.show()

        self.jacketStartValue = QLineEdit(self.designlabel)
        self.jacketStartValue.setGeometry(590, 780, 80, 40)
        self.jacketStartValue.textChanged.connect(self.initialchanged)
        self.jacketStartValue.setFont(QFont('Arial', 15))
        self.jacketStartValue.setStyleSheet("QLineEdit{ background-color: rgb(255, 255, 255);}"
                                            "QLineEdit{ border: 2px solid black;}"
                                            "QLineEdit{ border-radius: 5px;}")

        self.jacketEndValue = QLineEdit(self.designlabel)
        self.jacketEndValue.setGeometry(600, 321, 80, 40)
        self.jacketEndValue.textChanged.connect(self.initialchanged)
        self.jacketEndValue.setFont(QFont('Arial', 15))
        self.jacketEndValue.setStyleSheet("QLineEdit{ background-color: rgb(255, 255, 255);}"
                                          "QLineEdit{ border: 2px solid black;}"
                                          "QLineEdit{ border-radius: 5px;}")

        self.tankHeighttext = VerticalLineEdit(self.designlabel)
        self.tankHeighttext.setGeometry(263, 415, 40, 60)
        self.tankHeighttext.setObjectName("tankHeighttext")
        self.tankHeighttext.setFont(QFont('Arial', 14))
        self.tankHeighttext.setStyleSheet("QLineEdit#tankHeighttext{ background-color: rgb(255, 255, 255);}"
                                          "QLineEdit#tankHeighttext{ border: 2px solid black;}"
                                          "QLineEdit#tankHeighttext{ border-radius: 5px;}")
        self.tankHeighttext.setAlignment(QtCore.Qt.AlignVCenter)
        self.tankHeighttext.show()

        self.initialHeighttext = VerticalLineEdit(self.designlabel)
        self.initialHeighttext.setGeometry(670, 410, 40, 60)
        self.initialHeighttext.setObjectName("initialHeighttext")
        self.initialHeighttext.setFont(QFont('Arial', 14))
        self.initialHeighttext.setStyleSheet("QLineEdit#initialHeighttext{ background-color: rgb(255, 255, 255);}"
                                             "QLineEdit#initialHeighttext{ border: 2px solid black;}"
                                             "QLineEdit#initialHeighttext{ border-radius: 5px;}")
        self.initialHeighttext.show()


        self.tankDiameterLabel = QLabel(self.designlabel)
        self.tankDiameterLabel.setGeometry(325, 136, 220, 40)
        self.tankDiameterLabel.setObjectName("tankDiameterlabel")
        self.tankDiameterLabel.setText("Tank Diameter (m):")
        self.tankDiameterLabel.setFont(QFont('Arial', 15, weight=QtGui.QFont.Bold))
        self.tankDiameterLabel.show()

        self.tankDiametertext = QLineEdit(self.designlabel)
        self.tankDiametertext.setGeometry(560, 134, 60, 40)
        self.tankDiametertext.setObjectName("tankDiametertext")
        self.tankDiametertext.setFont(QFont('Arial', 14))
        self.tankDiametertext.setStyleSheet("QLineEdit#tankDiametertext{ background-color: rgb(255, 255, 255);}"
                                            "QLineEdit#tankDiametertext{ border: 2px solid black;}"
                                            "QLineEdit#tankDiametertext{ border-radius: 5px;}")
        self.tankDiametertext.show()

        self.feedarrowfun()
        self.productarrowfun()

        self.noStreamsLabel = QLabel(self.designlabel)
        self.noStreamsLabel.setGeometry(10, 60, 250, 60)
        self.noStreamsLabel.setObjectName("noStreamsLabel")
        self.noStreamsLabel.setText("Number of Feed Streams:")
        self.noStreamsLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.noStreamsLabel.hide()

        self.noStreamstext = QComboBox(self.designlabel)
        self.noStreamstext.setGeometry(270, 75, 50, 28)
        self.noStreamstext.addItems(['', '0', '1', '2', '3'])
        self.noStreamstext.setObjectName("noStreamstext")
        self.noStreamstext.setFont(QFont('Arial', 12))
        self.noStreamstext.setEditable(False)
        self.noStreamstext.setStyleSheet("QComboBox#noStreamstext{ background-color: rgb(255, 255, 255);}"
                                         "QComboBox#noStreamstext{ border: 2px solid black;}"
                                         "QComboBox#noStreamstext{ border-radius: 5px;}")
        # self.noStreamstext.currentTextChanged.connect(self.feedarrowfun)
        self.noStreamstext.hide()

        self.noProductsLabel = QLabel(self.designlabel)
        self.noProductsLabel.setGeometry(660, 60, 280, 60)
        self.noProductsLabel.setObjectName("noProductsLabel")
        self.noProductsLabel.setText("Number of Product Streams:")
        self.noProductsLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.noProductsLabel.hide()

        self.noProductstext = QComboBox(self.designlabel)
        self.noProductstext.setGeometry(940, 75, 50, 28)
        self.noProductstext.addItems(['', '0', '1', '2', '3'])
        self.noProductstext.setObjectName("noProductstext")
        self.noProductstext.setFont(QFont('Arial', 12))
        self.noProductstext.setEditable(False)
        self.noProductstext.setStyleSheet("QComboBox#noProductstext{ background-color: rgb(255, 255, 255);}"
                                          "QComboBox#noProductstext{ border: 2px solid black;}"
                                          "QComboBox#noProductstext{ border-radius: 5px;}")
        # self.noProductstext.currentTextChanged.connect(self.productarrowfun)
        self.noProductstext.hide()


        ## Heat Loss Tab
        self.sigmaLabel = QLabel(self.heatlosslabel)
        self.sigmaLabel.setGeometry(5, 15, 600, 25)
        self.sigmaLabel.setText("Evaporation/Condensation Coefficient ((kmol/kg)\u2070\u0387\u2075.s.m\u207b\u00b9):")
        self.sigmaLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.sigmaText = QLineEdit(self.heatlosslabel)
        self.sigmaText.setGeometry(580, 12, 200, 35)
        self.sigmaText.setFont(QFont('Arial', 12))
        self.sigmaText.setObjectName("sigmaText")
        self.sigmaText.textChanged.connect(self.heatleakchanged)
        self.sigmaText.setStyleSheet("QLineEdit#sigmaText{ background-color: rgb(255, 255, 255);}"
                                     "QLineEdit#sigmaText{ border: 2px solid black;}"
                                     "QLineEdit#sigmaText{ border-radius: 5px;}")

        self.UlLabel = QLabel(self.heatlosslabel)
        self.UlLabel.setGeometry(5, 75, 520, 25)
        self.UlLabel.setObjectName("UlLabel")
        self.UlLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.UlLabel.setText("Liquid Phase Film Heat Transfer Coefficient (W/(m\u00b2.K)):")

        self.Ultext = QLineEdit(self.heatlosslabel)
        self.Ultext.setGeometry(540, 72, 80, 35)
        self.Ultext.setObjectName("Ultext")
        self.Ultext.setFont(QFont('Arial', 12))
        self.Ultext.textChanged.connect(self.heatleakchanged)
        self.Ultext.setStyleSheet("QLineEdit#Ultext{ background-color: rgb(255, 255, 255);}"
                                  "QLineEdit#Ultext{ border: 2px solid black;}"
                                  "QLineEdit#Ultext{ border-radius: 5px;}")

        self.UvLabel = QLabel(self.heatlosslabel)
        self.UvLabel.setGeometry(5, 135, 530, 25)
        self.UvLabel.setObjectName("UvLabel")
        self.UvLabel.setText("Vapour Phase Film Heat Transfer Coefficient (W/(m\u00b2.K)):")
        self.UvLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.Uvtext = QLineEdit(self.heatlosslabel)
        self.Uvtext.setGeometry(550, 132, 80, 35)
        self.Uvtext.setObjectName("Uvtext")
        self.Uvtext.setFont(QFont('Arial', 12))
        self.Uvtext.textChanged.connect(self.heatleakchanged)
        self.Uvtext.setStyleSheet("QLineEdit#Uvtext{ background-color: rgb(255, 255, 255);}"
                                  "QLineEdit#Uvtext{ border: 2px solid black;}"
                                  "QLineEdit#Uvtext{ border-radius: 5px;}")

        self.UvwLabel = QLabel(self.heatlosslabel)
        self.UvwLabel.setGeometry(5, 195, 500, 25)
        self.UvwLabel.setObjectName("UvwLabel")
        self.UvwLabel.setText("Wall-Vapour Heat Transfer Coefficient (W/(m\u00b2.K)):")
        self.UvwLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.Uvwtext = QLineEdit(self.heatlosslabel)
        self.Uvwtext.setGeometry(480, 192, 80, 35)
        self.Uvwtext.setObjectName("Uvwtext")
        self.Uvwtext.setFont(QFont('Arial', 12))
        self.Uvwtext.textChanged.connect(self.heatleakchanged)
        self.Uvwtext.setStyleSheet("QLineEdit#Uvwtext{ background-color: rgb(255, 255, 255);}"
                                   "QLineEdit#Uvwtext{ border: 2px solid black;}"
                                   "QLineEdit#Uvwtext{ border-radius: 5px;}")

        self.UlwLabel = QLabel(self.heatlosslabel)
        self.UlwLabel.setGeometry(5, 255, 500, 25)
        self.UlwLabel.setObjectName("UvwLabel")
        self.UlwLabel.setText("Wall-Liquid Heat Transfer Coefficient (W/(m\u00b2.K)):")
        self.UlwLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.Ulwtext = QLineEdit(self.heatlosslabel)
        self.Ulwtext.setGeometry(480, 252, 80, 35)
        self.Ulwtext.setObjectName("Ulwtext")
        self.Ulwtext.setFont(QFont('Arial', 12))
        self.Ulwtext.textChanged.connect(self.heatleakchanged)
        self.Ulwtext.setStyleSheet("QLineEdit#Ulwtext{ background-color: rgb(255, 255, 255);}"
                                   "QLineEdit#Ulwtext{ border: 2px solid black;}"
                                   "QLineEdit#Ulwtext{ border-radius: 5px;}")

        self.UrLabel = QLabel(self.heatlosslabel)
        self.UrLabel.setGeometry(5, 315, 550, 25)
        self.UrLabel.setObjectName("UrLabel")
        self.UrLabel.setText("Tank Roof-Vapour Heat Transfer Coefficient (W/(m\u00b2.K)):")
        self.UrLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.Urtext = QLineEdit(self.heatlosslabel)
        self.Urtext.setGeometry(550, 312, 80, 35)
        self.Urtext.setObjectName("Urtext")
        self.Urtext.setFont(QFont('Arial', 12))
        self.Urtext.textChanged.connect(self.heatleakchanged)
        self.Urtext.setStyleSheet("QLineEdit#Urtext{ background-color: rgb(255, 255, 255);}"
                                  "QLineEdit#Urtext{ border: 2px solid black;}"
                                  "QLineEdit#Urtext{ border-radius: 5px;}")

        self.UbLabel = QLabel(self.heatlosslabel)
        self.UbLabel.setGeometry(5, 375, 550, 25)
        self.UbLabel.setObjectName("UbLabel")
        self.UbLabel.setText("Tank Bottom-Liquid Heat Transfer Coefficient (W/(m\u00b2.K)):")
        self.UbLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.Ubtext = QLineEdit(self.heatlosslabel)
        self.Ubtext.setGeometry(550, 372, 80, 35)
        self.Ubtext.setObjectName("Ubtext")
        self.Ubtext.setFont(QFont('Arial', 12))
        self.Ubtext.textChanged.connect(self.heatleakchanged)
        self.Ubtext.setStyleSheet("QLineEdit#Ubtext{ background-color: rgb(255, 255, 255);}"
                                  "QLineEdit#Ubtext{ border: 2px solid black;}"
                                  "QLineEdit#Ubtext{ border-radius: 5px;}")

        self.UvrLabel = QLabel(self.heatlosslabel)
        self.UvrLabel.setGeometry(5, 435, 510, 25)
        self.UvrLabel.setObjectName("UvrLabel")
        self.UvrLabel.setText("Jacket-Vapour Heat Transfer Coefficient (W/(m\u00b2.K)):")
        self.UvrLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.Uvrtext = QLineEdit(self.heatlosslabel)
        self.Uvrtext.setGeometry(500, 432, 80, 35)
        self.Uvrtext.setObjectName("Uvrtext")
        self.Uvrtext.setFont(QFont('Arial', 12))
        self.Uvrtext.textChanged.connect(self.heatleakchanged)
        self.Uvrtext.setStyleSheet("QLineEdit#Uvrtext{ background-color: rgb(255, 255, 255);}"
                                   "QLineEdit#Uvrtext{ border: 2px solid black;}"
                                   "QLineEdit#Uvrtext{ border-radius: 5px;}")

        self.UlrLabel = QLabel(self.heatlosslabel)
        self.UlrLabel.setGeometry(5, 495, 510, 25)
        self.UlrLabel.setObjectName("UvrLabel")
        self.UlrLabel.setText("Jacket-Liquid Heat Transfer Coefficient (W/(m\u00b2.K)):")
        self.UlrLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.Ulrtext = QLineEdit(self.heatlosslabel)
        self.Ulrtext.setGeometry(500, 492, 80, 35)
        self.Ulrtext.setObjectName("Ulrtext")
        self.Ulrtext.setFont(QFont('Arial', 12))
        self.Ulrtext.textChanged.connect(self.heatleakchanged)
        self.Ulrtext.setStyleSheet("QLineEdit#Ulrtext{ background-color: rgb(255, 255, 255);}"
                                   "QLineEdit#Ulrtext{ border: 2px solid black;}"
                                   "QLineEdit#Ulrtext{ border-radius: 5px;}")

        self.groundTempLabel = QLabel(self.heatlosslabel)
        self.groundTempLabel.setGeometry(5, 555, 300, 25)
        self.groundTempLabel.setObjectName("groundTempLabel")
        self.groundTempLabel.setText("Ground Temperature (K):")
        self.groundTempLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.groundTemptext = QLineEdit(self.heatlosslabel)
        self.groundTemptext.setGeometry(270, 552, 80, 35)
        self.groundTemptext.setObjectName("groundTemptext")
        self.groundTemptext.setFont(QFont('Arial', 12))
        self.groundTemptext.textChanged.connect(self.heatleakchanged)
        self.groundTemptext.setStyleSheet("QLineEdit#groundTemptext{ background-color: rgb(255, 255, 255);}"
                                          "QLineEdit#groundTemptext{ border: 2px solid black;}"
                                          "QLineEdit#groundTemptext{ border-radius: 5px;}")

        self.ambTempLabel = QLabel(self.heatlosslabel)
        self.ambTempLabel.setGeometry(5, 615, 300, 25)
        self.ambTempLabel.setObjectName("ambTempLabel")
        self.ambTempLabel.setText("Ambient Temperature (K):")
        self.ambTempLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.ambTemptext = QLineEdit(self.heatlosslabel)
        self.ambTemptext.setGeometry(270, 612, 80, 35)
        self.ambTemptext.setObjectName("ambTemptext")
        self.ambTemptext.setFont(QFont('Arial', 12))
        self.ambTemptext.textChanged.connect(self.heatleakchanged)
        self.ambTemptext.setStyleSheet("QLineEdit#ambTemptext{ background-color: rgb(255, 255, 255);}"
                                       "QLineEdit#ambTemptext{ border: 2px solid black;}"
                                       "QLineEdit#ambTemptext{ border-radius: 5px;}")

        self.roofTempLabel = QLabel(self.heatlosslabel)
        self.roofTempLabel.setGeometry(5, 675, 300, 25)
        self.roofTempLabel.setObjectName("roofTempLabel")
        self.roofTempLabel.setText("Roof Temperature (K):")
        self.roofTempLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.roofTemptext = QLineEdit(self.heatlosslabel)
        self.roofTemptext.setGeometry(241, 672, 80, 35)
        self.roofTemptext.setObjectName("roofTemptext")
        self.roofTemptext.setFont(QFont('Arial', 12))
        self.roofTemptext.textChanged.connect(self.heatleakchanged)
        self.roofTemptext.setStyleSheet("QLineEdit#roofTemptext{ background-color: rgb(255, 255, 255);}"
                                        "QLineEdit#roofTemptext{ border: 2px solid black;}"
                                        "QLineEdit#roofTemptext{ border-radius: 5px;}")

        self.refridgeTempLabel = QLabel(self.heatlosslabel)
        self.refridgeTempLabel.setGeometry(5, 735, 300, 25)
        self.refridgeTempLabel.setObjectName("refridgeTempLabel")
        self.refridgeTempLabel.setText("Refrigerant Temperature (K):")
        self.refridgeTempLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.refridgeTemptext = QLineEdit(self.heatlosslabel)
        self.refridgeTemptext.setGeometry(300, 732, 80, 35)
        self.refridgeTemptext.setObjectName("refridgeTemptext")
        self.refridgeTemptext.textChanged.connect(self.heatleakchanged)
        self.refridgeTemptext.setFont(QFont('Arial', 12))
        self.refridgeTemptext.setStyleSheet("QLineEdit#refridgeTemptext{ background-color: rgb(255, 255, 255);}"
                                            "QLineEdit#refridgeTemptext{ border: 2px solid black;}"
                                            "QLineEdit#refridgeTemptext{ border-radius: 5px;}")

        #
        # self.jacketButton = QtWidgets.QPushButton(self.heatlosslabel)
        # self.jacketButton.setGeometry(400, 615, 200, 50)
        # self.jacketButton.setObjectName("jacketButton")
        # self.jacketButton.setText('Optional Inputs')
        # self.jacketButton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        # self.jacketButton.setStyleSheet(
        #     "QPushButton#jacketButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
        #     "QPushButton#jacketButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        # self.jacketButton.clicked.connect(self.jacketfun)

        self.progressWindow = QLabel()
        self.progressWindow.setFixedSize(600, 300)
        self.progressWindow.setWindowTitle("Analysing Data")
        self.progressWindow.hide()

        self.progressWindowtext = QLabel(self.progressWindow)
        self.progressWindowtext.setGeometry(125, 80, 350, 80)
        self.progressWindowtext.setFont(QFont('Arial', 12))
        self.progressWindowtext.setAlignment(QtCore.Qt.AlignCenter)
        self.progressWindowtext.hide()

        self.noLDisksLabel = QLabel(self.initiallabel)
        self.noLDisksLabel.setGeometry(0, 25, 300, 25)
        self.noLDisksLabel.setObjectName("noLDisksLabel")
        self.noLDisksLabel.setText("Number of Liquid Disks:")
        self.noLDisksLabel.setFont(QFont('Arial', 13, weight=QtGui.QFont.Bold))
        self.noLDisksLabel.show()

        self.noLDiskstext = QLineEdit(self.initiallabel)
        self.noLDiskstext.setGeometry(260, 20, 60, 40)
        self.noLDiskstext.setObjectName("noLDiskstext")
        self.noLDiskstext.setFont(QFont('Arial', 13))
        self.noLDiskstext.textChanged.connect(self.diskTablefun)
        self.noLDiskstext.setStyleSheet("QLineEdit#noLDiskstext{ background-color: rgb(255, 255, 255);}"
                                        "QLineEdit#noLDiskstext{ border: 2px solid black;}"
                                        "QLineEdit#noLDiskstext{ border-radius: 5px;}")
        self.noLDiskstext.show()

        self.noVDisksLabel = QLabel(self.initiallabel)
        self.noVDisksLabel.setGeometry(340, 25, 300, 25)
        self.noVDisksLabel.setObjectName("noLDisksLabel")
        self.noVDisksLabel.setText("Number of Vapour Disks:")
        self.noVDisksLabel.setFont(QFont('Arial', 13, weight=QtGui.QFont.Bold))
        self.noVDisksLabel.show()

        self.noVDiskstext = QLineEdit(self.initiallabel)
        self.noVDiskstext.setGeometry(605, 20, 60, 40)
        self.noVDiskstext.setObjectName("noLDiskstext")
        self.noVDiskstext.setFont(QFont('Arial', 13))
        self.noVDiskstext.textChanged.connect(self.diskTablefun)
        self.noVDiskstext.setStyleSheet("QLineEdit#noLDiskstext{ background-color: rgb(255, 255, 255);}"
                                        "QLineEdit#noLDiskstext{ border: 2px solid black;}"
                                        "QLineEdit#noLDiskstext{ border-radius: 5px;}")
        self.noVDiskstext.show()

        self.abstolLabel = QLabel(self.initiallabel)
        self.abstolLabel.setGeometry(0, 85, 380, 25)
        self.abstolLabel.setObjectName("abstolLabel")
        self.abstolLabel.setText("Absolute DAE Solver Tolerance:")
        self.abstolLabel.setFont(QFont('Arial', 13, weight=QtGui.QFont.Bold))
        self.abstolLabel.show()

        self.abstoltext = QLineEdit(self.initiallabel)
        self.abstoltext.setGeometry(340, 80, 100, 40)
        self.abstoltext.setObjectName("abstoltext")
        self.abstoltext.setFont(QFont('Arial', 13))
        self.abstoltext.setText('0.01')
        self.abstoltext.setStyleSheet(
            "QLineEdit#abstoltext{ background-color: rgb(255, 255, 255);}"
            "QLineEdit#abstoltext{ border: 2px solid black;}"
            "QLineEdit#abstoltext{ border-radius: 5px;}")
        self.abstoltext.show()

        self.reltolLabel = QLabel(self.initiallabel)
        self.reltolLabel.setGeometry(460, 85, 380, 25)
        self.reltolLabel.setObjectName("reltolLabel")
        self.reltolLabel.setText("Relative DAE Solver Tolerance:")
        self.reltolLabel.setFont(QFont('Arial', 13, weight=QtGui.QFont.Bold))
        self.reltolLabel.show()

        self.reltoltext = QLineEdit(self.initiallabel)
        self.reltoltext.setGeometry(800, 80, 100, 40)
        self.reltoltext.setObjectName("reltoltext")
        self.reltoltext.setFont(QFont('Arial', 13))
        self.reltoltext.setText('0.0001')
        self.reltoltext.setStyleSheet(
            "QLineEdit#reltoltext{ background-color: rgb(255, 255, 255);}"
            "QLineEdit#reltoltext{ border: 2px solid black;}"
            "QLineEdit#reltoltext{ border-radius: 5px;}")
        self.reltoltext.show()

        self.numberofIterationsLabel = QLabel(self.initiallabel)
        self.numberofIterationsLabel.setGeometry(0, 145, 255, 30)
        self.numberofIterationsLabel.setObjectName("numberofIterationsLabel")
        self.numberofIterationsLabel.setText("Running Time (min):")
        self.numberofIterationsLabel.setFont(QFont('Arial', 13, weight=QtGui.QFont.Bold))
        self.numberofIterationsLabel.show()

        self.numberofIterationstext = QLineEdit(self.initiallabel)
        self.numberofIterationstext.setGeometry(225, 140, 100, 40)
        self.numberofIterationstext.setObjectName("numberofIterationstext")
        self.numberofIterationstext.setFont(QFont('Arial', 13))
        self.numberofIterationstext.setStyleSheet(
            "QLineEdit#numberofIterationstext{ background-color: rgb(255, 255, 255);}"
            "QLineEdit#numberofIterationstext{ border: 2px solid black;}"
            "QLineEdit#numberofIterationstext{ border-radius: 5px;}")
        self.numberofIterationstext.show()

        self.diskTableLabel = QLabel(self.initiallabel)
        self.diskTableLabel.setGeometry(360, 200, 280, 30)
        self.diskTableLabel.setObjectName("diskTableLabel")
        self.diskTableLabel.setText("Disk Initial Conditions Table")
        self.diskTableLabel.setFont(QFont('Arial', 13, weight=QtGui.QFont.Bold))
        self.diskTableLabel.hide()

        self.diskTable = QTableWidget(self.initiallabel)

        # self.diskbutton = QPushButton(self.initiallabel)
        # self.diskbutton.setObjectName("diskbutton")
        # self.diskbutton.setGeometry(338, 420, 150, 60)
        # self.diskbutton.setText("Disk Initial\nConditions")
        # self.diskbutton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        # self.diskbutton.setStyleSheet(
        #     "QPushButton#diskbutton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
        #     "QPushButton#diskbutton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        # self.diskbutton.clicked.connect(self.diskTablefun)

        self.analysebutton = QtWidgets.QPushButton(self.initiallabel)
        self.analysebutton.setGeometry(510, 880, 150, 60)
        self.analysebutton.setObjectName("analysebutton")
        self.analysebutton.setText('Run')
        self.analysebutton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.analysebutton.setStyleSheet(
            "QPushButton#analysebutton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#analysebutton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.analysebutton.clicked.connect(self.analysefun)
        self.analysebutton.setEnabled(True)
        self.analysebutton.show()

        self.saveInputsbutton = QtWidgets.QPushButton(self.initiallabel)
        self.saveInputsbutton.setGeometry(340, 880, 150, 60)
        self.saveInputsbutton.setObjectName("saveInputsbutton")
        self.saveInputsbutton.setText('Save Inputs')
        self.saveInputsbutton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.saveInputsbutton.setToolTip('Save Inputs to relevant excel files')
        self.saveInputsbutton.setStyleSheet(
            "QPushButton#saveInputsbutton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#saveInputsbutton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.saveInputsbutton.clicked.connect(self.saveinputfun)
        self.saveInputsbutton.setEnabled(True)
        self.saveInputsbutton.show()

        self.plot1button = QPushButton(self.resultslabel)
        self.plot1button.setObjectName("plot1button")
        self.plot1button.setGeometry(10, 10, 100, 60)
        self.plot1button.setText("Pressure")
        self.plot1button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot1button.clicked.connect(self.plot1fun)

        self.plot2button = QPushButton(self.resultslabel)
        self.plot2button.setObjectName("plot2button")
        self.plot2button.setGeometry(120, 10, 100, 60)
        self.plot2button.setText("Liquid\nLevel")
        self.plot2button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.clicked.connect(self.plot2fun)

        self.plot3button = QPushButton(self.resultslabel)
        self.plot3button.setObjectName("plot3button")
        self.plot3button.setGeometry(230, 10, 110, 60)
        self.plot3button.setText("Temperature")
        self.plot3button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.clicked.connect(self.plot3fun)

        self.plot4button = QPushButton(self.resultslabel)
        self.plot4button.setObjectName("plot4button")
        self.plot4button.setGeometry(350, 10, 100, 60)
        self.plot4button.setText("Vapour\nFlows")
        self.plot4button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.clicked.connect(self.plot4fun)

        self.plot5button = QPushButton(self.resultslabel)
        self.plot5button.setObjectName("plot5button")
        self.plot5button.setGeometry(460, 10, 130, 60)
        self.plot5button.setText("Liquid\nComposition")
        self.plot5button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.clicked.connect(self.plot5fun)

        self.plot6button = QPushButton(self.resultslabel)
        self.plot6button.setObjectName("plot6button")
        self.plot6button.setGeometry(600, 10, 130, 60)
        self.plot6button.setText("Vapour\nComposition")
        self.plot6button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.clicked.connect(self.plot6fun)

        self.plot7button = QPushButton(self.resultslabel)
        self.plot7button.setObjectName("plot7button")
        self.plot7button.setGeometry(740, 10, 120, 60)
        self.plot7button.setText("Product\nTemperature")
        self.plot7button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.clicked.connect(self.plot7fun)

        self.plot8button = QPushButton(self.resultslabel)
        self.plot8button.setObjectName("plot8button")
        self.plot8button.setGeometry(870, 10, 120, 60)
        self.plot8button.setText("Product\nPressure")
        self.plot8button.setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.clicked.connect(self.plot8fun)

        self.plot_widget = QWidget(self.resultslabel)
        self.plot_widget.setGeometry(QtCore.QRect(0, 80, 1000, 740))
        self.fig, self.ax = plt.subplots(1, 1, dpi=100)
        plt.subplots_adjust(left=0.10, right=0.98, top=0.94, bottom=0.10, wspace=0.20, hspace=0.35)
        self.plotting = FigureCanvas(self.fig)
        plot_box = QVBoxLayout()
        plot_box.addWidget(self.plotting)
        self.plot_widget.setLayout(plot_box)
        self.plot_widget.show()

        self.plot_widget2 = QWidget()
        self.fig2, self.ax2 = plt.subplots(1, 1, dpi=100)
        plt.subplots_adjust(left=0.15, right=0.98, top=0.94, bottom=0.10, wspace=0.20, hspace=0.35)
        self.plotting2 = FigureCanvas(self.fig2)
        plot_box = QVBoxLayout()
        plot_box.addWidget(self.plotting2)
        self.plot_widget2.setLayout(plot_box)
        self.plot_widget2.hide()

        self.exportResultsbutton = QPushButton(self.resultslabel)
        self.exportResultsbutton.setGeometry(340 ,825, 150, 75)
        self.exportResultsbutton.setObjectName("exportResultsbutton")
        self.exportResultsbutton.setText("Export Results")
        self.exportResultsbutton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.exportResultsbutton.setStyleSheet(
            "QPushButton#exportResultsbutton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#exportResultsbutton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.exportResultsbutton.clicked.connect(self.resultsExcel)

        self.exportGraphsbutton = QPushButton(self.resultslabel)
        self.exportGraphsbutton.setGeometry(510, 825, 150, 75)
        self.exportGraphsbutton.setObjectName("exportGraphsbutton")
        self.exportGraphsbutton.setText("Export Graphs")
        self.exportGraphsbutton.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
        self.exportGraphsbutton.setStyleSheet(
            "QPushButton#exportGraphsbutton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#exportGraphsbutton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.exportGraphsbutton.clicked.connect(self.resultsGraph)

        # self.defaultfun()

    def defaultfun(self):
        self.design_df = pd.read_excel('Base Case.xlsx', sheet_name='Design', header=None)
        self.tankDiametertext.setText(str(int(self.design_df.iloc[0,1])))
        self.tankHeighttext.setText(str(int(self.design_df.iloc[1,1])))
        self.initialPressuretext.setText(str(self.design_df.iloc[2,1]))
        self.initialHeighttext.setText(str(int(self.design_df.iloc[3,1])))
        self.noComponentstext.setCurrentText(str(int(self.design_df.iloc[4,1])))
        self.jacketStartValue.setText(str(self.design_df.iloc[5,1]))
        self.jacketEndValue.setText(str(self.design_df.iloc[6,1]))

        self.jacketLocationList = [float(self.jacketStartValue.text()), float(self.jacketEndValue.text())]


        self.feedLocationList = [str(int(x)) if x != "" else "N. A." for x in self.design_df.iloc[12:17, 1].fillna('')]
        value=0
        for key in self.feedLocationValueDict:
            self.feedLocationValueDict[key].setText(self.feedLocationList[value])
            value+=1

        self.productLocationList = [str(int(x)) if x != "" else "N. A." for x in self.design_df.iloc[19:24, 1].fillna('')]
        value=0
        for key in self.productLocationValueDict:
            self.productLocationValueDict[key].setText(self.productLocationList[value])
            value+=1

        self.heatLoss_df = pd.read_excel("Base Case.xlsx", sheet_name='Heat Loss', header=None)
        print(self.heatLoss_df.iloc[0,1])
        self.sigmaText.setText(str(self.heatLoss_df.iloc[0, 1]))
        self.Ultext.setText(str(self.heatLoss_df.iloc[1, 1]))
        self.Uvtext.setText(str(self.heatLoss_df.iloc[2, 1]))
        self.Uvwtext.setText(str(self.heatLoss_df.iloc[3, 1]))
        self.Ulwtext.setText(str(self.heatLoss_df.iloc[4, 1]))
        self.Urtext.setText(str(self.heatLoss_df.iloc[5, 1]))
        self.Ubtext.setText(str(self.heatLoss_df.iloc[6, 1]))
        self.groundTemptext.setText(str(self.heatLoss_df.iloc[7, 1]))
        self.ambTemptext.setText(str(self.heatLoss_df.iloc[8, 1]))
        self.roofTemptext.setText(str(self.heatLoss_df.iloc[9, 1]))
        self.Uvrtext.setText(str(self.heatLoss_df.iloc[10,1]))
        self.Ulrtext.setText(str(self.heatLoss_df.iloc[11,1]))
        self.refridgeTemptext.setText(str(self.heatLoss_df.iloc[12,1]))

        self.initial_df = pd.read_excel("Base Case.xlsx", sheet_name='Initial', header=None)
        self.noLDiskstext.setText(str(int(self.initial_df.iloc[0,1])))
        self.noVDiskstext.setText(str(int(self.initial_df.iloc[1,1])))
        self.abstoltext.setText(str(self.initial_df.iloc[2,1]))
        self.reltoltext.setText(str(self.initial_df.iloc[3,1]))
        self.numberofIterationstext.setText(str(int(self.initial_df.iloc[4,1])))

        # self.optional_df = pd.read_excel("Base Case.xlsx", sheet_name='Optional', header=None)
        # self.jacketLocationList = self.optional_df.iloc[:, 1].values.tolist()

        I = int(self.noComponentstext.currentText())
        if int(self.noLDiskstext.text()) > 0:
            arrangedName = ['Component {}'.format(str(i+1)) for i in range(I)]
            T0 = pd.read_excel("Initial condition.xlsx", "Data", usecols="A").values.tolist()
            T0 = np.array(T0)
            init_DF = pd.read_excel("Initial condition.xlsx", "Data")
            x0 = np.zeros((len(T0), I))
            for i in range(I):
                x0[:, i] = init_DF[arrangedName[i]].values.tolist()
            self.diskinitCombined = np.concatenate((T0, x0), axis=1)

        self.MW_list = [16, 30, 44.1, 28]

    def noComponentschanged(self, text):
        if text != "":
            self.MWButton.setEnabled(True)
            self.MWButton.setStyleSheet(
                "QPushButton#MWButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#MWButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
        try:
            self.MW_list = self.MW_list[:int(text)]
            print(self.MW_list)
            self.compTable.setColumnCount(0)
            self.compTable.setRowCount(0)
        except:
            pass

    def noDiskschanged(self, text):
        if text != "":
            self.initialButton.setEnabled(True)
            self.initialButton.setStyleSheet(
                "QPushButton#initialButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
                "QPushButton#initialButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}"
            )

    def MWfun(self):
        self.MWWindow = QFrame()
        length = 300
        heightTable = 60 * int(self.noComponentstext.currentText()) + 10
        self.MWWindow.setFixedSize(length, heightTable + 50)
        self.MWWindow.setWindowTitle('MW')
        self.MWWindow.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                    "QTabBar::tab { height: 50px;}")

        self.MWWindowLabel = QLabel(self.MWWindow)
        self.MWWindowLabel.setGeometry(5, 0, 500, 50)
        self.MWWindowLabel.setObjectName("MWWindowLabel")
        self.MWWindowLabel.setText("Molecular Weight (kg/kmol)")
        self.MWWindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

        self.compTablehorzList = ['Component {}'.format(str(i + 1)) for i in
                                  range(int(self.noComponentstext.currentText()))]
        self.MWLabeldict = dict.fromkeys(self.compTablehorzList, 0)
        self.MWdict = dict.fromkeys(self.compTablehorzList, 0)

        value = 1

        for key in self.MWLabeldict:
            self.MWLabeldict[key] = QLabel(self.MWWindow)
            self.MWLabeldict[key].setGeometry(5, 60 * value, 200, 50)
            self.MWLabeldict[key].setText(key + ":")
            self.MWLabeldict[key].setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            value += 1

        value = 1
        for key in self.MWdict:
            self.MWdict[key] = QLineEdit(self.MWWindow)
            self.MWdict[key].setGeometry(180, 60 * value + 12, 60, 28)
            objectName = "LineEdit{}".format(str(value))
            self.MWdict[key].setObjectName(objectName)
            try:
                self.MWdict[key].setText(str(self.MW_list[value-1]))
            except:
                self.MWdict[key].setText("")
            self.MWdict[key].textChanged.connect(self.MWchanged)
            self.MWdict[key].setFont(QFont('Arial', 12))
            stylesheetline1 = "QLineEdit#{}{{ background-color: rgb(255, 255, 255);}}".format(objectName)
            stylesheetline2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
            stylesheetline3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)
            self.MWdict[key].setStyleSheet(stylesheetline1+stylesheetline2+stylesheetline3)
            value += 1

        self.MWWindow.show()

    def MWchanged(self):
        try:
            self.MW_list2 = []
            for key in self.MWdict:
                self.MW_list2.append(str(self.MWdict[key].text()))
            self.MW_list = self.MW_list2
            print(self.MW_list)
        except:
            pass

    def toDesign(self):
        self.designButton.setStyleSheet(
            "QPushButton#designButton{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.heatLossButton.setStyleSheet(
            "QPushButton#heatLossButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#heatLossButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.initialButton.setStyleSheet(
            "QPushButton#initialButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#initialButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.resultsButton.setStyleSheet(
            "QPushButton#resultsButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#resultsButton::hover{background-color: rgb(255, 0, 0); color: rgb(255, 255, 255); border: 2px solid black}")

        self.designlabel.show()
        self.heatlosslabel.hide()
        self.initiallabel.hide()
        self.complabel.hide()
        self.resultslabel.hide()

    def feedarrowfun(self):
        noOfFeedStreams = 5
        feedArrowKeys = ['Feed Arrow {}'.format(str(i+1)) for i in range(noOfFeedStreams)]
        feedKeys = ['Feed {}'.format(str(i+1)) for i in range(noOfFeedStreams)]
        self.feedArrowDict = dict.fromkeys(feedArrowKeys, 0)
        self.feedButtonDict = dict.fromkeys(feedKeys, 0)
        self.feedLocationLabelDict = dict.fromkeys(feedKeys, 0)
        self.feedLocationValueDict = dict.fromkeys(feedKeys, 0)
        self.feedButtonfunList = [self.feedButton1fun, self.feedButton2fun, self.feedButton3fun,
                                  self.feedButton4fun, self.feedButton5fun]
        value = 0

        for key in self.feedArrowDict:
            self.feedArrowDict[key] = QLabel(self.designlabel)
            self.feedArrowDict[key].setGeometry(50, 150 + 150*value, 200, 25)
            self.arrowimage = QPixmap('arrow.png')
            self.feedArrowDict[key].setPixmap(self.arrowimage)
            self.feedArrowDict[key].show()
            value += 1

        value = 0
        for key in self.feedButtonDict:
            self.feedButtonDict[key] = QPushButton(self.designlabel)
            self.feedButtonDict[key].setGeometry(90, 110 + 150*value, 100, 40)
            self.feedButtonDict[key].setFont(QFont('Arial', 14, weight=QtGui.QFont.Bold))
            self.feedButtonDict[key].setText('Feed {}'.format(str(value+1)))
            objectName = "FeedButton{}".format(str(value+1))
            self.feedButtonDict[key].setObjectName(objectName)
            self.feedButtonDict[key].clicked.connect(self.feedButtonfunList[value])

            string1 = "QPushButton#{}{{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}}".format(objectName)
            string2 = "QPushButton#{}::hover{{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}}".format(objectName)

            self.feedButtonDict[key].setStyleSheet(string1+string2)
            self.feedButtonDict[key].show()
            value += 1

        value = 0
        for key in self.feedLocationLabelDict:
            self.feedLocationLabelDict[key] = QLabel(self.designlabel)
            self.feedLocationLabelDict[key].setGeometry(40, 170 + 150*value, 120, 40)
            self.feedLocationLabelDict[key].setText("Height (%):")
            self.feedLocationLabelDict[key].setFont(QFont('Arial', 14, weight=QtGui.QFont.Bold))
            self.feedLocationLabelDict[key].show()
            value += 1

        value = 0
        for key in self.feedLocationValueDict:
            self.feedLocationValueDict[key] = QLineEdit(self.designlabel)
            objectName = "FeedLocation{}".format(str(value+1))
            self.feedLocationValueDict[key].textChanged.connect(self.feedValid)
            self.feedLocationValueDict[key].setObjectName(objectName)
            self.feedLocationValueDict[key].setGeometry(170, 170 + 150*value, 60, 40)
            self.feedLocationValueDict[key].setFont(QFont('Arial', 14))

            string1 = "QLineEdit#{}{{ background-color: rgb(255, 255, 255);}}".format(objectName)
            string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
            string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)

            self.feedLocationValueDict[key].setStyleSheet(string1+string2+string3)
            self.feedLocationValueDict[key].show()
            value += 1

    def feedValid(self):

        def isfloat(value):
            try:
                float(value)
                return True
            except ValueError:
                return False

        value = 0
        for key in self.feedLocationValueDict:
            objectName = "FeedLocation{}".format(str(value + 1))
            if isfloat(self.feedLocationValueDict[key].text()) and float(self.feedLocationValueDict[key].text()) <= 100 and float(self.feedLocationValueDict[key].text()) >= 0:
                string1 = "QLineEdit#{}{{ background-color: rgb(0, 255, 0);}}".format(objectName)
                string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
                string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)
                self.feedLocationValueDict[key].setStyleSheet(string1 + string2 + string3)

            elif self.feedLocationValueDict[key].text() == "":
                self.feedLocationValueDict[key].setText("N. A.")
                string1 = "QLineEdit#{}{{ background-color: rgb(255, 255, 255);}}".format(objectName)
                string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
                string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)
                self.feedLocationValueDict[key].setStyleSheet(string1 + string2 + string3)

            else:
                string1 = "QLineEdit#{}{{ background-color: rgb(255, 255, 255);}}".format(objectName)
                string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
                string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)
                self.feedLocationValueDict[key].setStyleSheet(string1 + string2 + string3)
            value += 1

    def productarrowfun(self):
        noOfProductStreams = 5
        productArrowKeys = ['Product Arrow {}'.format(str(i + 1)) for i in range(noOfProductStreams)]
        productKeys = ['Product {}'.format(str(i + 1)) for i in range(noOfProductStreams)]
        self.productArrowDict = dict.fromkeys(productArrowKeys, 0)
        self.productButtonDict = dict.fromkeys(productKeys, 0)
        self.productLocationLabelDict = dict.fromkeys(productKeys, 0)
        self.productLocationValueDict = dict.fromkeys(productKeys, 0)
        self.productButtonfunList = [self.productButton1fun, self.productButton2fun, self.productButton3fun,
                                     self.productButton4fun, self.productButton5fun]
        value = 0

        for key in self.productArrowDict:
            self.productArrowDict[key] = QLabel(self.designlabel)
            self.productArrowDict[key].setGeometry(750, 150 + 150*value, 200, 25)
            self.arrowimage = QPixmap('arrow.png')
            self.productArrowDict[key].setPixmap(self.arrowimage)
            self.productArrowDict[key].show()
            value += 1

        value = 0
        for key in self.productButtonDict:
            self.productButtonDict[key] = QPushButton(self.designlabel)
            self.productButtonDict[key].setGeometry(790, 110 + 150 * value, 120, 40)
            self.productButtonDict[key].setFont(QFont('Arial', 14, weight=QtGui.QFont.Bold))
            self.productButtonDict[key].setText('Product {}'.format(str(value + 1)))
            objectName = "FeedButton{}".format(str(value+1))
            self.productButtonDict[key].setObjectName(objectName)

            string1 = "QPushButton#{}{{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}}".format(
                objectName)
            string2 = "QPushButton#{}::hover{{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}}".format(
                objectName)

            self.productButtonDict[key].setStyleSheet(string1+string2)
            self.productButtonDict[key].clicked.connect(self.productButtonfunList[value])
            self.productButtonDict[key].show()
            value += 1

        value = 0
        for key in self.productLocationLabelDict:
            self.productLocationLabelDict[key] = QLabel(self.designlabel)
            self.productLocationLabelDict[key].setGeometry(760, 170 + 150 * value, 120, 40)
            self.productLocationLabelDict[key].setText("Height (%):")
            self.productLocationLabelDict[key].setFont(QFont('Arial', 14, weight=QtGui.QFont.Bold))
            self.productLocationLabelDict[key].show()
            value += 1

        value = 0
        for key in self.productLocationValueDict:
            self.productLocationValueDict[key] = QLineEdit(self.designlabel)
            objectName = "FeedLocation{}".format(str(value + 1))
            self.productLocationValueDict[key].setObjectName(objectName)
            self.productLocationValueDict[key].setGeometry(890, 170 + 150 * value, 60, 40)
            self.productLocationValueDict[key].textChanged.connect(self.productValid)
            self.productLocationValueDict[key].setFont(QFont('Arial', 14))

            string1 = "QLineEdit#{}{{ background-color: rgb(255, 255, 255);}}".format(objectName)
            string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
            string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)

            self.productLocationValueDict[key].setStyleSheet(string1 + string2 + string3)
            self.productLocationValueDict[key].show()
            value += 1

    def productValid(self):
        def isfloat(value):
            try:
                float(value)
                return True
            except ValueError:
                return False

        value = 0
        for key in self.productLocationValueDict:
            objectName = "FeedLocation{}".format(str(value + 1))
            if isfloat(self.productLocationValueDict[key].text()) and float(
                    self.productLocationValueDict[key].text()) <= 100 and \
                    float(self.productLocationValueDict[key].text()) >= 0:
                string1 = "QLineEdit#{}{{ background-color: rgb(0, 255, 0);}}".format(objectName)
                string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
                string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)
                self.productLocationValueDict[key].setStyleSheet(string1 + string2 + string3)

            elif self.productLocationValueDict[key].text() == "":
                self.productLocationValueDict[key].setText("N. A.")
                string1 = "QLineEdit#{}{{ background-color: rgb(255, 255, 255);}}".format(objectName)
                string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
                string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)
                self.productLocationValueDict[key].setStyleSheet(string1 + string2 + string3)

            else:
                string1 = "QLineEdit#{}{{ background-color: rgb(255, 255, 255);}}".format(objectName)
                string2 = "QLineEdit#{}{{ border: 2px solid black;}}".format(objectName)
                string3 = "QLineEdit#{}{{ border-radius: 5px;}}".format(objectName)
                self.productLocationValueDict[key].setStyleSheet(string1 + string2 + string3)
            value += 1

    def feedButton1fun(self):
        try:
            noOfComponents = int(self.noComponentstext.currentText())
            noOfColumns = noOfComponents + 4
            self.feed_stream_1DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Af(1)')
            self.feed_stream_1DF = self.feed_stream_1DF.iloc[:, :noOfColumns]

            self.feed1TableWindow = QFrame()
            length = 100 * noOfColumns + 30
            self.feed1TableWindow.setFixedSize(length, 550)
            self.feed1TableWindow.setWindowTitle('Feed 1')
            self.feed1TableWindow.show()

            self.feed1Table = QTableWidget(self.feed1TableWindow)
            self.feed1Table.setGeometry(0, 0, length, 500)
            self.feed1Table.horizontalHeader().setDefaultSectionSize(100)
            self.feed1horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)', 'Temperature\n(K)', 'Pressure\n(kPa)']
            self.feed1compositionList = ['MassFrac {}'.format(str(i+1)) for i in range(int(self.noComponentstext.currentText()))]
            self.feed1horizontalList = self.feed1horizontalList + self.feed1compositionList
            self.feed1Table.setColumnCount(len(self.feed1horizontalList))
            self.feed1Table.setHorizontalHeaderLabels(self.feed1horizontalList)
            self.feed1Table.horizontalHeader().setStretchLastSection(True)
            self.feed1Table.verticalHeader().setVisible(False)
            self.feed1Table.setRowCount(self.feed_stream_1DF.shape[0])
            self.feed1Table.cellClicked.connect(self.getRowFeed1)
            self.feed1Table.show()

            for i in range(self.feed1Table.rowCount()):
                for j in range(self.feed1Table.columnCount()):
                    value = str(self.feed_stream_1DF.iloc[i,j])
                    self.feed1Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.savefeed1button = QPushButton(self.feed1TableWindow)
            self.savefeed1button.setObjectName("savefeed1button")
            self.savefeed1button.setGeometry(length-100, 500, 100, 50)
            self.savefeed1button.setText('Save')
            self.savefeed1button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.savefeed1button.clicked.connect(self.savefeed1fun)
            self.savefeed1button.setStyleSheet(
                "QPushButton#savefeed1button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#savefeed1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.savefeed1button.show()

            self.addfeed1button = QPushButton(self.feed1TableWindow)
            self.addfeed1button.setObjectName("addfeed1button")
            self.addfeed1button.setGeometry(length-440, 500, 150, 50)
            self.addfeed1button.setText('Add Row')
            self.addfeed1button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addfeed1button.clicked.connect(self.addfeed1fun)
            self.addfeed1button.setStyleSheet(
                "QPushButton#addfeed1button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addfeed1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addfeed1button.show()

            self.delfeed1button = QPushButton(self.feed1TableWindow)
            self.delfeed1button.setObjectName("delfeed1button")
            self.delfeed1button.setGeometry(length-270, 500, 150, 50)
            self.delfeed1button.setText('Delete Row')
            self.delfeed1button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delfeed1button.clicked.connect(self.delfeed1fun)
            self.delfeed1button.setStyleSheet(
                "QPushButton#delfeed1button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delfeed1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delfeed1button.show()

        except:
            self.noFeed1Window = QFrame()
            length = 700
            heightTable = 60
            self.noFeed1Window.setFixedSize(length, heightTable + 50)
            self.noFeed1Window.setWindowTitle('Excel Column Not Found')
            self.noFeed1Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                             "QTabBar::tab { height: 50px;}")
            self.noFeed1WindowLabel = QLabel(self.noFeed1Window)
            self.noFeed1WindowLabel.setGeometry(5, 20, 700, 70)
            self.noFeed1WindowLabel.setObjectName("noFeed1WindowLabel")
            self.noFeed1WindowLabel.setText(
                "Error:\n\nSheet name Af(1) is not found in Feed_Product flows.xlsx Excel file.")
            self.noFeed1WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noFeed1Window.show()

    def feedButton2fun(self):
        try:
            noOfComponents = int(self.noComponentstext.currentText())
            noOfColumns = noOfComponents + 4
            self.feed_stream_2DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Af(2)')
            self.feed_stream_2DF = self.feed_stream_2DF.iloc[:, :noOfColumns]

            self.feed2TableWindow = QFrame()
            length = 100 * noOfColumns + 30
            self.feed2TableWindow.setFixedSize(length, 550)
            self.feed2TableWindow.setWindowTitle('Feed 2')
            self.feed2TableWindow.show()

            self.feed2Table = QTableWidget(self.feed2TableWindow)
            self.feed2Table.setGeometry(0, 0, length, 500)
            self.feed2Table.horizontalHeader().setDefaultSectionSize(100)
            self.feed2horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)', 'Temperature\n(K)', 'Pressure\n(kPa)']
            self.feed2compositionList = ['MassFrac {}'.format(str(i + 1)) for i in range(int(self.noComponentstext.currentText()))]
            self.feed2horizontalList = self.feed2horizontalList + self.feed2compositionList
            self.feed2Table.setColumnCount(len(self.feed2horizontalList))
            self.feed2Table.setHorizontalHeaderLabels(self.feed2horizontalList)
            self.feed2Table.horizontalHeader().setStretchLastSection(True)
            self.feed2Table.verticalHeader().setVisible(False)
            self.feed2Table.setRowCount(self.feed_stream_2DF.shape[0])
            self.feed2Table.cellClicked.connect(self.getRowFeed2)
            self.feed2Table.show()

            for i in range(self.feed2Table.rowCount()):
                for j in range(self.feed2Table.columnCount()):
                    value = str(self.feed_stream_2DF.iloc[i, j])
                    self.feed2Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.savefeed2button = QPushButton(self.feed2TableWindow)
            self.savefeed2button.setObjectName("savefeed2button")
            self.savefeed2button.setGeometry(length - 100, 500, 100, 50)
            self.savefeed2button.setText('Save')
            self.savefeed2button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.savefeed2button.clicked.connect(self.savefeed2fun)
            self.savefeed2button.setStyleSheet(
                "QPushButton#savefeed2button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#savefeed2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.savefeed2button.show()

            self.addfeed2button = QPushButton(self.feed2TableWindow)
            self.addfeed2button.setObjectName("addfeed2button")
            self.addfeed2button.setGeometry(length - 440, 500, 150, 50)
            self.addfeed2button.setText('Add Row')
            self.addfeed2button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addfeed2button.clicked.connect(self.addfeed2fun)
            self.addfeed2button.setStyleSheet(
                "QPushButton#addfeed2button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addfeed2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addfeed2button.show()

            self.delfeed2button = QPushButton(self.feed2TableWindow)
            self.delfeed2button.setObjectName("delfeed2button")
            self.delfeed2button.setGeometry(length - 270, 500, 150, 50)
            self.delfeed2button.setText('Delete Row')
            self.delfeed2button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delfeed2button.clicked.connect(self.delfeed2fun)
            self.delfeed2button.setStyleSheet(
                "QPushButton#delfeed2button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delfeed2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delfeed2button.show()

        except:
            self.noFeed2Window = QFrame()
            length = 700
            heightTable = 60
            self.noFeed2Window.setFixedSize(length, heightTable + 50)
            self.noFeed2Window.setWindowTitle('Excel Column Not Found')
            self.noFeed2Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                             "QTabBar::tab { height: 50px;}")
            self.noFeed2WindowLabel = QLabel(self.noFeed2Window)
            self.noFeed2WindowLabel.setGeometry(5, 20, 700, 70)
            self.noFeed2WindowLabel.setObjectName("noFeed2WindowLabel")
            self.noFeed2WindowLabel.setText(
                "Error:\n\nSheet name Af(2) is not found in Feed_Product flows.xlsx Excel file.")
            self.noFeed2WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noFeed2Window.show()

    def feedButton3fun(self):
        try:
            noOfComponents = int(self.noComponentstext.currentText())
            noOfColumns = noOfComponents + 4
            self.feed_stream_3DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Af(3)')
            self.feed_stream_3DF = self.feed_stream_3DF.iloc[:, :noOfColumns]

            self.feed3TableWindow = QFrame()
            length = 100 * noOfColumns + 30
            self.feed3TableWindow.setFixedSize(length, 550)
            self.feed3TableWindow.setWindowTitle('Feed 3')
            self.feed3TableWindow.show()

            self.feed3Table = QTableWidget(self.feed3TableWindow)
            self.feed3Table.setGeometry(0, 0, length, 500)
            self.feed3Table.horizontalHeader().setDefaultSectionSize(100)
            self.feed3horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)', 'Temperature\n(K)', 'Pressure\n(kPa)']
            self.feed3compositionList = ['MassFrac {}'.format(str(i + 1)) for i in range(int(self.noComponentstext.currentText()))]
            self.feed3horizontalList = self.feed3horizontalList + self.feed3compositionList
            self.feed3Table.setColumnCount(len(self.feed3horizontalList))
            self.feed3Table.setHorizontalHeaderLabels(self.feed3horizontalList)
            self.feed3Table.horizontalHeader().setStretchLastSection(True)
            self.feed3Table.verticalHeader().setVisible(False)
            self.feed3Table.setRowCount(self.feed_stream_3DF.shape[0])
            self.feed3Table.cellClicked.connect(self.getRowFeed3)
            self.feed3Table.show()

            for i in range(self.feed3Table.rowCount()):
                for j in range(self.feed3Table.columnCount()):
                    value = str(self.feed_stream_3DF.iloc[i, j])
                    self.feed3Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.savefeed3button = QPushButton(self.feed3TableWindow)
            self.savefeed3button.setObjectName("savefeed3button")
            self.savefeed3button.setGeometry(length - 100, 500, 100, 50)
            self.savefeed3button.setText('Save')
            self.savefeed3button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.savefeed3button.clicked.connect(self.savefeed3fun)
            self.savefeed3button.setStyleSheet(
                "QPushButton#savefeed3button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#savefeed3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.savefeed3button.show()

            self.addfeed3button = QPushButton(self.feed3TableWindow)
            self.addfeed3button.setObjectName("addfeed3button")
            self.addfeed3button.setGeometry(length - 440, 500, 150, 50)
            self.addfeed3button.setText('Add Row')
            self.addfeed3button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addfeed3button.clicked.connect(self.addfeed3fun)
            self.addfeed3button.setStyleSheet(
                "QPushButton#addfeed3button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addfeed3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addfeed3button.show()

            self.delfeed3button = QPushButton(self.feed3TableWindow)
            self.delfeed3button.setObjectName("delfeed3button")
            self.delfeed3button.setGeometry(length - 270, 500, 150, 50)
            self.delfeed3button.setText('Delete Row')
            self.delfeed3button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delfeed3button.clicked.connect(self.delfeed3fun)
            self.delfeed3button.setStyleSheet(
                "QPushButton#delfeed3button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delfeed3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delfeed3button.show()

        except:
            self.noFeed3Window = QFrame()
            length = 700
            heightTable = 60
            self.noFeed3Window.setFixedSize(length, heightTable + 50)
            self.noFeed3Window.setWindowTitle('Excel Column Not Found')
            self.noFeed3Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                             "QTabBar::tab { height: 50px;}")
            self.noFeed3WindowLabel = QLabel(self.noFeed3Window)
            self.noFeed3WindowLabel.setGeometry(5, 20, 700, 70)
            self.noFeed3WindowLabel.setObjectName("noFeed3WindowLabel")
            self.noFeed3WindowLabel.setText(
                "Error:\n\nSheet name Af(3) is not found in Feed_Product flows.xlsx Excel file.")
            self.noFeed3WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noFeed3Window.show()

    def feedButton4fun(self):
        try:
            noOfComponents = int(self.noComponentstext.currentText())
            noOfColumns = noOfComponents + 4
            self.feed_stream_4DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Af(4)')
            self.feed_stream_4DF = self.feed_stream_4DF.iloc[:, :noOfColumns]

            self.feed4TableWindow = QFrame()
            length = 100 * noOfColumns + 30
            self.feed4TableWindow.setFixedSize(length, 550)
            self.feed4TableWindow.setWindowTitle('Feed 4')
            self.feed4TableWindow.show()

            self.feed4Table = QTableWidget(self.feed4TableWindow)
            self.feed4Table.setGeometry(0, 0, length, 500)
            self.feed4Table.horizontalHeader().setDefaultSectionSize(100)
            self.feed4horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)', 'Temperature\n(K)', 'Pressure\n(kPa)']
            self.feed4compositionList = ['MassFrac {}'.format(str(i + 1)) for i in
                                         range(int(self.noComponentstext.currentText()))]
            self.feed4horizontalList = self.feed4horizontalList + self.feed4compositionList
            self.feed4Table.setColumnCount(len(self.feed4horizontalList))
            self.feed4Table.setHorizontalHeaderLabels(self.feed4horizontalList)
            self.feed4Table.horizontalHeader().setStretchLastSection(True)
            self.feed4Table.verticalHeader().setVisible(False)
            self.feed4Table.setRowCount(self.feed_stream_4DF.shape[0])
            self.feed4Table.cellClicked.connect(self.getRowFeed4)
            self.feed4Table.show()

            for i in range(self.feed4Table.rowCount()):
                for j in range(self.feed4Table.columnCount()):
                    value = str(self.feed_stream_4DF.iloc[i, j])
                    self.feed4Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.savefeed4button = QPushButton(self.feed4TableWindow)
            self.savefeed4button.setObjectName("savefeed4button")
            self.savefeed4button.setGeometry(length - 100, 500, 100, 50)
            self.savefeed4button.setText('Save')
            self.savefeed4button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.savefeed4button.clicked.connect(self.savefeed4fun)
            self.savefeed4button.setStyleSheet(
                "QPushButton#savefeed4button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#savefeed4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.savefeed4button.show()

            self.addfeed4button = QPushButton(self.feed4TableWindow)
            self.addfeed4button.setObjectName("addfeed4button")
            self.addfeed4button.setGeometry(length - 440, 500, 150, 50)
            self.addfeed4button.setText('Add Row')
            self.addfeed4button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addfeed4button.clicked.connect(self.addfeed4fun)
            self.addfeed4button.setStyleSheet(
                "QPushButton#addfeed4button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addfeed4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addfeed4button.show()

            self.delfeed4button = QPushButton(self.feed4TableWindow)
            self.delfeed4button.setObjectName("delfeed4button")
            self.delfeed4button.setGeometry(length - 270, 500, 150, 50)
            self.delfeed4button.setText('Delete Row')
            self.delfeed4button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delfeed4button.clicked.connect(self.delfeed4fun)
            self.delfeed4button.setStyleSheet(
                "QPushButton#delfeed4button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delfeed4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delfeed4button.show()


        except:
            self.noFeed4Window = QFrame()
            length = 700
            heightTable = 60
            self.noFeed4Window.setFixedSize(length, heightTable + 50)
            self.noFeed4Window.setWindowTitle('Excel Column Not Found')
            self.noFeed4Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                             "QTabBar::tab { height: 50px;}")
            self.noFeed4WindowLabel = QLabel(self.noFeed4Window)
            self.noFeed4WindowLabel.setGeometry(5, 20, 700, 70)
            self.noFeed4WindowLabel.setObjectName("noFeed4WindowLabel")
            self.noFeed4WindowLabel.setText("Error:\n\nSheet name Af(4) is not found in Feed_Product flows.xlsx Excel file.")
            self.noFeed4WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noFeed4Window.show()

    def feedButton5fun(self):
        try:
            noOfComponents = int(self.noComponentstext.currentText())
            noOfColumns = noOfComponents + 4
            self.feed_stream_5DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Af(5)')
            self.feed_stream_5DF = self.feed_stream_5DF.iloc[:, :noOfColumns]

            self.feed5TableWindow = QFrame()
            length = 100 * noOfColumns + 30
            self.feed5TableWindow.setFixedSize(length, 550)
            self.feed5TableWindow.setWindowTitle('Feed 5')
            self.feed5TableWindow.show()

            self.feed5Table = QTableWidget(self.feed5TableWindow)
            self.feed5Table.setGeometry(0, 0, length, 500)
            self.feed5Table.horizontalHeader().setDefaultSectionSize(100)
            self.feed5horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)', 'Temperature\n(K)', 'Pressure\n(kPa)']
            self.feed5compositionList = ['MassFrac {}'.format(str(i + 1)) for i in
                                         range(int(self.noComponentstext.currentText()))]
            self.feed5horizontalList = self.feed5horizontalList + self.feed5compositionList
            self.feed5Table.setColumnCount(len(self.feed5horizontalList))
            self.feed5Table.setHorizontalHeaderLabels(self.feed5horizontalList)
            self.feed5Table.horizontalHeader().setStretchLastSection(True)
            self.feed5Table.verticalHeader().setVisible(False)
            self.feed5Table.setRowCount(self.feed_stream_5DF.shape[0])
            self.feed5Table.cellClicked.connect(self.getRowFeed5)
            self.feed5Table.show()

            for i in range(self.feed5Table.rowCount()):
                for j in range(self.feed5Table.columnCount()):
                    value = str(self.feed_stream_5DF.iloc[i, j])
                    self.feed5Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.savefeed5button = QPushButton(self.feed5TableWindow)
            self.savefeed5button.setObjectName("savefeed5button")
            self.savefeed5button.setGeometry(length - 100, 500, 100, 50)
            self.savefeed5button.setText('Save')
            self.savefeed5button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.savefeed5button.clicked.connect(self.savefeed5fun)
            self.savefeed5button.setStyleSheet(
                "QPushButton#savefeed5button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#savefeed5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.savefeed5button.show()

            self.addfeed5button = QPushButton(self.feed5TableWindow)
            self.addfeed5button.setObjectName("addfeed5button")
            self.addfeed5button.setGeometry(length - 440, 500, 150, 50)
            self.addfeed5button.setText('Add Row')
            self.addfeed5button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addfeed5button.clicked.connect(self.addfeed5fun)
            self.addfeed5button.setStyleSheet(
                "QPushButton#addfeed5button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addfeed5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addfeed5button.show()

            self.delfeed5button = QPushButton(self.feed5TableWindow)
            self.delfeed5button.setObjectName("delfeed5button")
            self.delfeed5button.setGeometry(length - 270, 500, 150, 50)
            self.delfeed5button.setText('Delete Row')
            self.delfeed5button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delfeed5button.clicked.connect(self.delfeed5fun)
            self.delfeed5button.setStyleSheet(
                "QPushButton#delfeed5button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delfeed5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delfeed5button.show()

        except:
            self.noFeed5Window = QFrame()
            length = 700
            heightTable = 60
            self.noFeed5Window.setFixedSize(length, heightTable + 50)
            self.noFeed5Window.setWindowTitle('Excel Column Not Found')
            self.noFeed5Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                             "QTabBar::tab { height: 50px;}")
            self.noFeed5WindowLabel = QLabel(self.noFeed5Window)
            self.noFeed5WindowLabel.setGeometry(5, 20, 700, 70)
            self.noFeed5WindowLabel.setObjectName("noFeed5WindowLabel")
            self.noFeed5WindowLabel.setText(
                "Error:\n\nSheet name Af(5) is not found in Feed_Product flows.xlsx Excel file.")
            self.noFeed5WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noFeed5Window.show()

    def productButton1fun(self):
        try:
            self.product_stream_1DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Ag(1)')
            self.product_stream_1DF = self.product_stream_1DF.iloc[:, :2]

            self.product1TableWindow = QFrame()
            length = 440
            self.product1TableWindow.setWindowTitle("Product 1")
            self.product1TableWindow.setFixedSize(length, 600)
            self.product1TableWindow.show()

            self.product1nameLabel = QLabel(self.product1TableWindow)
            self.product1nameLabel.setGeometry(165, 0, 230, 50)
            self.product1nameLabel.setText("Product 1")
            self.product1nameLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.product1nameLabel.show()

            self.product1Table = QTableWidget(self.product1TableWindow)
            self.product1Table.setGeometry(105, 50, 230, 500)
            self.product1Table.horizontalHeader().setDefaultSectionSize(100)
            self.product1horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)']
            self.product1Table.setColumnCount(len(self.product1horizontalList))
            self.product1Table.setHorizontalHeaderLabels(self.product1horizontalList)
            self.product1Table.horizontalHeader().setStretchLastSection(True)
            self.product1Table.verticalHeader().setVisible(False)
            self.product1Table.setRowCount(self.product_stream_1DF.shape[0])
            self.product1Table.cellClicked.connect(self.getRowProd1)
            self.product1Table.show()

            for i in range(self.product1Table.rowCount()):
                for j in range(self.product1Table.columnCount()):
                    value = str(self.product_stream_1DF.iloc[i, j])
                    self.product1Table.setItem(i, j, QTableWidgetItem(value))
                    self.product1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.saveproduct1button = QPushButton(self.product1TableWindow)
            self.saveproduct1button.setObjectName("saveproduct1button")
            self.saveproduct1button.setGeometry(length - 100, 550, 100, 50)
            self.saveproduct1button.setText('Save')
            self.saveproduct1button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.saveproduct1button.clicked.connect(self.saveproduct1fun)
            self.saveproduct1button.setStyleSheet(
                "QPushButton#saveproduct1button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#saveproduct1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.saveproduct1button.show()

            self.addprod1button = QPushButton(self.product1TableWindow)
            self.addprod1button.setObjectName("addprod1button")
            self.addprod1button.setGeometry(length - 440, 550, 150, 50)
            self.addprod1button.setText('Add Row')
            self.addprod1button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addprod1button.clicked.connect(self.addprod1fun)
            self.addprod1button.setStyleSheet(
                "QPushButton#addprod1button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addprod1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addprod1button.show()

            self.delprod1button = QPushButton(self.product1TableWindow)
            self.delprod1button.setObjectName("delprod1button")
            self.delprod1button.setGeometry(length - 270, 550, 150, 50)
            self.delprod1button.setText('Delete Row')
            self.delprod1button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delprod1button.clicked.connect(self.delprod1fun)
            self.delprod1button.setStyleSheet(
                "QPushButton#delprod1button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delprod1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delprod1button.show()

        except:
            self.noProduct1Window = QFrame()
            length = 700
            heightTable = 60
            self.noProduct1Window.setFixedSize(length, heightTable + 50)
            self.noProduct1Window.setWindowTitle('Excel Column Not Found')
            self.noProduct1Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                             "QTabBar::tab { height: 50px;}")
            self.noProduct1WindowLabel = QLabel(self.noProduct1Window)
            self.noProduct1WindowLabel.setGeometry(5, 20, 700, 70)
            self.noProduct1WindowLabel.setObjectName("noProduct1WindowLabel")
            self.noProduct1WindowLabel.setText(
                "Error:\n\nSheet name Ag(1) is not found in Feed_Product flows.xlsx Excel file.")
            self.noProduct1WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noProduct1Window.show()

    def productButton2fun(self):
        try:
            self.product_stream_2DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Ag(2)')
            self.product_stream_2DF = self.product_stream_2DF.iloc[:, :2]

            self.product2TableWindow = QFrame()
            length = 440
            self.product2TableWindow.setWindowTitle("Product 2")
            self.product2TableWindow.setFixedSize(length, 600)
            self.product2TableWindow.show()

            self.product2nameLabel = QLabel(self.product2TableWindow)
            self.product2nameLabel.setGeometry(165, 0, 230, 50)
            self.product2nameLabel.setText("Product 2")
            self.product2nameLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.product2nameLabel.show()

            self.product2Table = QTableWidget(self.product2TableWindow)
            self.product2Table.setGeometry(105, 50, 230, 500)
            self.product2Table.horizontalHeader().setDefaultSectionSize(100)
            self.product2horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)']
            self.product2Table.setColumnCount(len(self.product2horizontalList))
            self.product2Table.setHorizontalHeaderLabels(self.product2horizontalList)
            self.product2Table.horizontalHeader().setStretchLastSection(True)
            self.product2Table.verticalHeader().setVisible(False)
            self.product2Table.setRowCount(self.product_stream_2DF.shape[0])
            self.product2Table.cellClicked.connect(self.getRowProd2)
            self.product2Table.show()

            for i in range(self.product2Table.rowCount()):
                for j in range(self.product2Table.columnCount()):
                    value = str(self.product_stream_2DF.iloc[i, j])
                    self.product2Table.setItem(i, j, QTableWidgetItem(value))
                    self.product2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.saveproduct2button = QPushButton(self.product2TableWindow)
            self.saveproduct2button.setObjectName("saveproduct2button")
            self.saveproduct2button.setGeometry(length - 100, 550, 100, 50)
            self.saveproduct2button.setText('Save')
            self.saveproduct2button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.saveproduct2button.clicked.connect(self.saveproduct2fun)
            self.saveproduct2button.setStyleSheet(
                "QPushButton#saveproduct2button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#saveproduct2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.saveproduct2button.show()

            self.addprod2button = QPushButton(self.product2TableWindow)
            self.addprod2button.setObjectName("addprod2button")
            self.addprod2button.setGeometry(length - 440, 550, 150, 50)
            self.addprod2button.setText('Add Row')
            self.addprod2button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addprod2button.clicked.connect(self.addprod2fun)
            self.addprod2button.setStyleSheet(
                "QPushButton#addprod2button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addprod2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addprod2button.show()

            self.delprod2button = QPushButton(self.product2TableWindow)
            self.delprod2button.setObjectName("delprod2button")
            self.delprod2button.setGeometry(length - 270, 550, 150, 50)
            self.delprod2button.setText('Delete Row')
            self.delprod2button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delprod2button.clicked.connect(self.delprod2fun)
            self.delprod2button.setStyleSheet(
                "QPushButton#delprod2button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delprod2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delprod2button.show()

        except:
            self.noProduct2Window = QFrame()
            length = 700
            heightTable = 60
            self.noProduct2Window.setFixedSize(length, heightTable + 50)
            self.noProduct2Window.setWindowTitle('Excel Column Not Found')
            self.noProduct2Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                                "QTabBar::tab { height: 50px;}")
            self.noProduct2WindowLabel = QLabel(self.noProduct2Window)
            self.noProduct2WindowLabel.setGeometry(5, 20, 700, 70)
            self.noProduct2WindowLabel.setObjectName("noProduct2WindowLabel")
            self.noProduct2WindowLabel.setText(
                "Error:\n\nSheet name Ag(2) is not found in Feed_Product flows.xlsx Excel file.")
            self.noProduct2WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noProduct2Window.show()

    def productButton3fun(self):
        try:
            self.product_stream_3DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Ag(3)')
            self.product_stream_3DF = self.product_stream_3DF.iloc[:, :2]

            self.product3TableWindow = QFrame()
            length = 440
            self.product3TableWindow.setWindowTitle("Product 3")
            self.product3TableWindow.setFixedSize(length, 600)
            self.product3TableWindow.show()

            self.product3nameLabel = QLabel(self.product3TableWindow)
            self.product3nameLabel.setGeometry(165, 0, 230, 50)
            self.product3nameLabel.setText("Product 3")
            self.product3nameLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.product3nameLabel.show()

            self.product3Table = QTableWidget(self.product3TableWindow)
            self.product3Table.setGeometry(105, 50, 230, 500)
            self.product3Table.horizontalHeader().setDefaultSectionSize(100)
            self.product3horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)']
            self.product3Table.setColumnCount(len(self.product3horizontalList))
            self.product3Table.setHorizontalHeaderLabels(self.product3horizontalList)
            self.product3Table.horizontalHeader().setStretchLastSection(True)
            self.product3Table.verticalHeader().setVisible(False)
            self.product3Table.setRowCount(self.product_stream_3DF.shape[0])
            self.product3Table.cellClicked.connect(self.getRowProd3)
            self.product3Table.show()

            for i in range(self.product3Table.rowCount()):
                for j in range(self.product3Table.columnCount()):
                    value = str(self.product_stream_3DF.iloc[i, j])
                    self.product3Table.setItem(i, j, QTableWidgetItem(value))
                    self.product3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.saveproduct3button = QPushButton(self.product3TableWindow)
            self.saveproduct3button.setObjectName("saveproduct3button")
            self.saveproduct3button.setGeometry(length - 100, 550, 100, 50)
            self.saveproduct3button.setText('Save')
            self.saveproduct3button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.saveproduct3button.clicked.connect(self.saveproduct3fun)
            self.saveproduct3button.setStyleSheet(
                "QPushButton#saveproduct3button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#saveproduct3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.saveproduct3button.show()

            self.addprod3button = QPushButton(self.product3TableWindow)
            self.addprod3button.setObjectName("addprod3button")
            self.addprod3button.setGeometry(length - 440, 550, 150, 50)
            self.addprod3button.setText('Add Row')
            self.addprod3button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addprod3button.clicked.connect(self.addprod3fun)
            self.addprod3button.setStyleSheet(
                "QPushButton#addprod3button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addprod3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addprod3button.show()

            self.delprod3button = QPushButton(self.product3TableWindow)
            self.delprod3button.setObjectName("delprod3button")
            self.delprod3button.setGeometry(length - 270, 550, 150, 50)
            self.delprod3button.setText('Delete Row')
            self.delprod3button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delprod3button.clicked.connect(self.delprod3fun)
            self.delprod3button.setStyleSheet(
                "QPushButton#delprod3button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delprod3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delprod3button.show()
        except:
            self.noProduct3Window = QFrame()
            length = 700
            heightTable = 60
            self.noProduct3Window.setFixedSize(length, heightTable + 50)
            self.noProduct3Window.setWindowTitle('Excel Column Not Found')
            self.noProduct3Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                                "QTabBar::tab { height: 50px;}")
            self.noProduct3WindowLabel = QLabel(self.noProduct3Window)
            self.noProduct3WindowLabel.setGeometry(5, 20, 700, 70)
            self.noProduct3WindowLabel.setObjectName("noProduct3WindowLabel")
            self.noProduct3WindowLabel.setText(
                "Error:\n\nSheet name Ag(3) is not found in Feed_Product flows.xlsx Excel file.")
            self.noProduct3WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noProduct3Window.show()

    def productButton4fun(self):
        try:
            self.product_stream_4DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Ag(4)')
            self.product_stream_4DF = self.product_stream_4DF.iloc[:, :2]

            self.product4TableWindow = QFrame()
            length = 440
            self.product4TableWindow.setWindowTitle("Product 4")
            self.product4TableWindow.setFixedSize(length, 600)
            self.product4TableWindow.show()

            self.product4nameLabel = QLabel(self.product4TableWindow)
            self.product4nameLabel.setGeometry(165, 0, 230, 50)
            self.product4nameLabel.setText("Product 4")
            self.product4nameLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.product4nameLabel.show()

            self.product4Table = QTableWidget(self.product4TableWindow)
            self.product4Table.setGeometry(105, 50, 230, 500)
            self.product4Table.horizontalHeader().setDefaultSectionSize(100)
            self.product4horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)']
            self.product4Table.setColumnCount(len(self.product4horizontalList))
            self.product4Table.setHorizontalHeaderLabels(self.product4horizontalList)
            self.product4Table.horizontalHeader().setStretchLastSection(True)
            self.product4Table.verticalHeader().setVisible(False)
            self.product4Table.setRowCount(self.product_stream_4DF.shape[0])
            self.product4Table.cellClicked.connect(self.getRowProd4)
            self.product4Table.show()

            for i in range(self.product4Table.rowCount()):
                for j in range(self.product4Table.columnCount()):
                    value = str(self.product_stream_4DF.iloc[i, j])
                    self.product4Table.setItem(i, j, QTableWidgetItem(value))
                    self.product4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.saveproduct4button = QPushButton(self.product4TableWindow)
            self.saveproduct4button.setObjectName("saveproduct4button")
            self.saveproduct4button.setGeometry(length - 100, 550, 100, 50)
            self.saveproduct4button.setText('Save')
            self.saveproduct4button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.saveproduct4button.clicked.connect(self.saveproduct4fun)
            self.saveproduct4button.setStyleSheet(
                "QPushButton#saveproduct4button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#saveproduct4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.saveproduct4button.show()

            self.addprod4button = QPushButton(self.product4TableWindow)
            self.addprod4button.setObjectName("addprod4button")
            self.addprod4button.setGeometry(length - 440, 550, 150, 50)
            self.addprod4button.setText('Add Row')
            self.addprod4button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addprod4button.clicked.connect(self.addprod4fun)
            self.addprod4button.setStyleSheet(
                "QPushButton#addprod4button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addprod4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addprod4button.show()

            self.delprod4button = QPushButton(self.product4TableWindow)
            self.delprod4button.setObjectName("delprod4button")
            self.delprod4button.setGeometry(length - 270, 550, 150, 50)
            self.delprod4button.setText('Delete Row')
            self.delprod4button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delprod4button.clicked.connect(self.delprod4fun)
            self.delprod4button.setStyleSheet(
                "QPushButton#delprod4button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delprod4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delprod4button.show()
        except:
            self.noProduct4Window = QFrame()
            length = 700
            heightTable = 60
            self.noProduct4Window.setFixedSize(length, heightTable + 50)
            self.noProduct4Window.setWindowTitle('Excel Column Not Found')
            self.noProduct4Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                                "QTabBar::tab { height: 50px;}")
            self.noProduct4WindowLabel = QLabel(self.noProduct4Window)
            self.noProduct4WindowLabel.setGeometry(5, 20, 700, 70)
            self.noProduct4WindowLabel.setObjectName("noProduct4WindowLabel")
            self.noProduct4WindowLabel.setText(
                "Error:\n\nSheet name Ag(4) is not found in Feed_Product flows.xlsx Excel file.")
            self.noProduct4WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noProduct4Window.show()

    def productButton5fun(self):
        try:
            self.product_stream_5DF = pd.read_excel("Feed_Product flows.xlsx", sheet_name='Ag(5)')
            self.product_stream_5DF = self.product_stream_5DF.iloc[:, :2]

            self.product5TableWindow = QFrame()
            length = 440
            self.product5TableWindow.setWindowTitle("Product 5")
            self.product5TableWindow.setFixedSize(length, 600)
            self.product5TableWindow.show()

            self.product5nameLabel = QLabel(self.product5TableWindow)
            self.product5nameLabel.setGeometry(165, 0, 230, 50)
            self.product5nameLabel.setText("Product 5")
            self.product5nameLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.product5nameLabel.show()

            self.product5Table = QTableWidget(self.product5TableWindow)
            self.product5Table.setGeometry(105, 50, 230, 500)
            self.product5Table.horizontalHeader().setDefaultSectionSize(100)
            self.product5horizontalList = ['Time\n(min)', 'Flow Rate\n(kg/s)']
            self.product5Table.setColumnCount(len(self.product5horizontalList))
            self.product5Table.setHorizontalHeaderLabels(self.product5horizontalList)
            self.product5Table.horizontalHeader().setStretchLastSection(True)
            self.product5Table.verticalHeader().setVisible(False)
            self.product5Table.setRowCount(self.product_stream_5DF.shape[0])
            self.product5Table.cellClicked.connect(self.getRowProd5)
            self.product5Table.show()

            for i in range(self.product5Table.rowCount()):
                for j in range(self.product5Table.columnCount()):
                    value = str(self.product_stream_5DF.iloc[i, j])
                    self.product5Table.setItem(i, j, QTableWidgetItem(value))
                    self.product5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

            self.saveproduct5button = QPushButton(self.product5TableWindow)
            self.saveproduct5button.setObjectName("saveproduct5button")
            self.saveproduct5button.setGeometry(length - 100, 550, 100, 50)
            self.saveproduct5button.setText('Save')
            self.saveproduct5button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.saveproduct5button.clicked.connect(self.saveproduct5fun)
            self.saveproduct5button.setStyleSheet(
                "QPushButton#saveproduct5button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#saveproduct5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.saveproduct5button.show()

            self.addprod5button = QPushButton(self.product5TableWindow)
            self.addprod5button.setObjectName("addprod5button")
            self.addprod5button.setGeometry(length - 440, 550, 150, 50)
            self.addprod5button.setText('Add Row')
            self.addprod5button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.addprod5button.clicked.connect(self.addprod5fun)
            self.addprod5button.setStyleSheet(
                "QPushButton#addprod5button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#addprod5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.addprod5button.show()

            self.delprod5button = QPushButton(self.product5TableWindow)
            self.delprod5button.setObjectName("delprod5button")
            self.delprod5button.setGeometry(length - 270, 550, 150, 50)
            self.delprod5button.setText('Delete Row')
            self.delprod5button.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            self.delprod5button.clicked.connect(self.delprod5fun)
            self.delprod5button.setStyleSheet(
                "QPushButton#delprod5button{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
                "QPushButton#delprod5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}"
            )
            self.delprod5button.show()
        except:
            self.noProduct5Window = QFrame()
            length = 700
            heightTable = 60
            self.noProduct5Window.setFixedSize(length, heightTable + 50)
            self.noProduct5Window.setWindowTitle('Excel Column Not Found')
            self.noProduct5Window.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                                "QTabBar::tab { height: 50px;}")
            self.noProduct5WindowLabel = QLabel(self.noProduct5Window)
            self.noProduct5WindowLabel.setGeometry(5, 20, 700, 70)
            self.noProduct5WindowLabel.setObjectName("noProduct5WindowLabel")
            self.noProduct5WindowLabel.setText(
                "Error:\n\nSheet name Ag(5) is not found in Feed_Product flows.xlsx Excel file.")
            self.noProduct5WindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            self.noProduct5Window.show()

    def getRowFeed1(self):
        self.table_row = self.feed1Table.currentItem().row()

    def getRowFeed2(self):
        self.table_row = self.feed2Table.currentItem().row()

    def getRowFeed3(self):
        self.table_row = self.feed3Table.currentItem().row()

    def getRowFeed4(self):
        self.table_row = self.feed4Table.currentItem().row()

    def getRowFeed5(self):
        self.table_row = self.feed5Table.currentItem().row()

    def getRowProd1(self):
        self.table_row = self.product1Table.currentItem().row()

    def getRowProd2(self):
        self.table_row = self.product2Table.currentItem().row()

    def getRowProd3(self):
        self.table_row = self.product3Table.currentItem().row()

    def getRowProd4(self):
        self.table_row = self.product4Table.currentItem().row()

    def getRowProd5(self):
        self.table_row = self.product5Table.currentItem().row()

    def addfeed1fun(self):
        df_top = self.feed_stream_1DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_1DF.iloc[self.table_row:, :]

        self.feed1Table.setRowCount(self.feed_stream_1DF.shape[0]+1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.feed1Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.feed1Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.feed1Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed1Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.feed1Table.setItem(k, j, QTableWidgetItem(value))
                self.feed1Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.feed1Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_1DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delfeed1fun(self):
        df_top = self.feed_stream_1DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_1DF.iloc[self.table_row+1:, :]
        self.feed1Table.setRowCount(self.feed_stream_1DF.shape[0]-1)

        for i in range(df_top.shape[0]):
            for j in range(self.feed1Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.feed1Table.setItem(i, j, QTableWidgetItem(value))
                self.feed1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed1Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.feed1Table.setItem(k, j, QTableWidgetItem(value))
                self.feed1Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_1DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)


    def savefeed1fun(self):
        for i in range(self.feed1Table.rowCount()):
            for j in range(self.feed1Table.columnCount()):
                self.feed_stream_1DF.iloc[i, j] = self.feed1Table.item(i,j).text()

        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Af(1)"]
        for r in range(self.feed_stream_1DF.shape[0]):
            for s in range(self.feed_stream_1DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.feed_stream_1DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addfeed2fun(self):
        df_top = self.feed_stream_2DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_2DF.iloc[self.table_row:, :]

        self.feed2Table.setRowCount(self.feed_stream_2DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.feed2Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.feed2Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.feed2Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed2Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.feed2Table.setItem(k, j, QTableWidgetItem(value))
                self.feed2Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.feed2Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_2DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delfeed2fun(self):
        df_top = self.feed_stream_2DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_2DF.iloc[self.table_row+1:, :]
        self.feed2Table.setRowCount(self.feed_stream_2DF.shape[0]-1)

        for i in range(df_top.shape[0]):
            for j in range(self.feed2Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.feed2Table.setItem(i, j, QTableWidgetItem(value))
                self.feed2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed2Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.feed2Table.setItem(k, j, QTableWidgetItem(value))
                self.feed2Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_2DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def savefeed2fun(self):
        for i in range(self.feed2Table.rowCount()):
            for j in range(self.feed2Table.columnCount()):
                self.feed_stream_2DF.iloc[i, j] = self.feed2Table.item(i,j).text()

        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Af(2)"]
        for r in range(self.feed_stream_2DF.shape[0]):
            for s in range(self.feed_stream_2DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.feed_stream_2DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addfeed3fun(self):
        df_top = self.feed_stream_3DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_3DF.iloc[self.table_row:, :]

        self.feed3Table.setRowCount(self.feed_stream_3DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.feed3Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.feed3Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.feed3Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed3Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.feed3Table.setItem(k, j, QTableWidgetItem(value))
                self.feed3Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.feed3Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_3DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delfeed3fun(self):
        df_top = self.feed_stream_3DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_3DF.iloc[self.table_row+1:, :]
        self.feed3Table.setRowCount(self.feed_stream_3DF.shape[0]-1)

        for i in range(df_top.shape[0]):
            for j in range(self.feed3Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.feed3Table.setItem(i, j, QTableWidgetItem(value))
                self.feed3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed3Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.feed3Table.setItem(k, j, QTableWidgetItem(value))
                self.feed3Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_3DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def savefeed3fun(self):
        for i in range(self.feed3Table.rowCount()):
            for j in range(self.feed3Table.columnCount()):
                self.feed_stream_3DF.iloc[i, j] = self.feed3Table.item(i, j).text()

        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Af(3)"]
        for r in range(self.feed_stream_3DF.shape[0]):
            for s in range(self.feed_stream_3DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.feed_stream_3DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addfeed4fun(self):
        df_top = self.feed_stream_4DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_4DF.iloc[self.table_row:, :]

        self.feed4Table.setRowCount(self.feed_stream_4DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.feed4Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.feed4Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.feed4Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed4Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.feed4Table.setItem(k, j, QTableWidgetItem(value))
                self.feed4Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.feed4Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_4DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delfeed4fun(self):
        df_top = self.feed_stream_4DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_4DF.iloc[self.table_row+1:, :]
        self.feed4Table.setRowCount(self.feed_stream_4DF.shape[0]-1)

        for i in range(df_top.shape[0]):
            for j in range(self.feed4Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.feed4Table.setItem(i, j, QTableWidgetItem(value))
                self.feed4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed4Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.feed4Table.setItem(k, j, QTableWidgetItem(value))
                self.feed4Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_4DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def savefeed4fun(self):
        for i in range(self.feed4Table.rowCount()):
            for j in range(self.feed4Table.columnCount()):
                self.feed_stream_4DF.iloc[i, j] = self.feed4Table.item(i, j).text()

        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Af(4)"]
        for r in range(self.feed_stream_4DF.shape[0]):
            for s in range(self.feed_stream_4DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.feed_stream_4DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addfeed5fun(self):
        df_top = self.feed_stream_5DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_5DF.iloc[self.table_row:, :]

        self.feed5Table.setRowCount(self.feed_stream_5DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.feed5Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.feed5Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.feed5Table.setItem(i, j, QTableWidgetItem(value))
                    self.feed5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed5Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.feed5Table.setItem(k, j, QTableWidgetItem(value))
                self.feed5Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.feed5Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_5DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delfeed5fun(self):
        df_top = self.feed_stream_5DF.iloc[:self.table_row, :]
        df_bottom = self.feed_stream_5DF.iloc[self.table_row+1:, :]
        self.feed5Table.setRowCount(self.feed_stream_5DF.shape[0]-1)

        for i in range(df_top.shape[0]):
            for j in range(self.feed5Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.feed5Table.setItem(i, j, QTableWidgetItem(value))
                self.feed5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.feed5Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.feed5Table.setItem(k, j, QTableWidgetItem(value))
                self.feed5Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.feed_stream_5DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def savefeed5fun(self):
        for i in range(self.feed5Table.rowCount()):
            for j in range(self.feed5Table.columnCount()):
                self.feed_stream_5DF.iloc[i, j] = self.feed5Table.item(i, j).text()

        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Af(5)"]
        for r in range(self.feed_stream_5DF.shape[0]):
            for s in range(self.feed_stream_5DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.feed_stream_5DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addprod1fun(self):
        df_top = self.product_stream_1DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_1DF.iloc[self.table_row:, :]

        self.product1Table.setRowCount(self.product_stream_1DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.product1Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.product1Table.setItem(i, j, QTableWidgetItem(value))
                    self.product1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.product1Table.setItem(i, j, QTableWidgetItem(value))
                    self.product1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product1Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.product1Table.setItem(k, j, QTableWidgetItem(value))
                self.product1Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.product1Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_1DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delprod1fun(self):
        df_top = self.product_stream_1DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_1DF.iloc[self.table_row+1:, :]
        self.product1Table.setRowCount(self.product_stream_1DF.shape[0]-1)

        for i in range(df_top.shape[0]):
            for j in range(self.product1Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.product1Table.setItem(i, j, QTableWidgetItem(value))
                self.product1Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product1Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.product1Table.setItem(k, j, QTableWidgetItem(value))
                self.product1Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_1DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def saveproduct1fun(self):
        for i in range(self.product1Table.rowCount()):
            for j in range(self.product1Table.columnCount()):
                self.product_stream_1DF.iloc[i, j] = self.product1Table.item(i, j).text()

        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Ag(1)"]
        for r in range(self.product_stream_1DF.shape[0]):
            for s in range(self.product_stream_1DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.product_stream_1DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addprod2fun(self):
        df_top = self.product_stream_2DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_2DF.iloc[self.table_row:, :]

        self.product2Table.setRowCount(self.product_stream_2DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.product2Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.product2Table.setItem(i, j, QTableWidgetItem(value))
                    self.product2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.product2Table.setItem(i, j, QTableWidgetItem(value))
                    self.product2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product2Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.product2Table.setItem(k, j, QTableWidgetItem(value))
                self.product2Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.product2Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_2DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delprod2fun(self):
        df_top = self.product_stream_2DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_2DF.iloc[self.table_row+1:, :]
        self.product2Table.setRowCount(self.product_stream_2DF.shape[0]-1)

        for i in range(df_top.shape[0]):
            for j in range(self.product2Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.product2Table.setItem(i, j, QTableWidgetItem(value))
                self.product2Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product2Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.product2Table.setItem(k, j, QTableWidgetItem(value))
                self.product2Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_2DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def saveproduct2fun(self):
        for i in range(self.product2Table.rowCount()):
            for j in range(self.product2Table.columnCount()):
                self.product_stream_2DF.iloc[i, j] = self.product2Table.item(i, j).text()

        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Ag(2)"]
        for r in range(self.product_stream_2DF.shape[0]):
            for s in range(self.product_stream_2DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.product_stream_2DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addprod3fun(self):
        df_top = self.product_stream_3DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_3DF.iloc[self.table_row:, :]

        self.product3Table.setRowCount(self.product_stream_3DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.product3Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.product3Table.setItem(i, j, QTableWidgetItem(value))
                    self.product3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.product3Table.setItem(i, j, QTableWidgetItem(value))
                    self.product3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product3Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.product3Table.setItem(k, j, QTableWidgetItem(value))
                self.product3Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.product3Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_3DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delprod3fun(self):
        df_top = self.product_stream_3DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_3DF.iloc[self.table_row + 1:, :]
        self.product3Table.setRowCount(self.product_stream_3DF.shape[0] - 1)

        for i in range(df_top.shape[0]):
            for j in range(self.product3Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.product3Table.setItem(i, j, QTableWidgetItem(value))
                self.product3Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product3Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.product3Table.setItem(k, j, QTableWidgetItem(value))
                self.product3Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_3DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def saveproduct3fun(self):
        for i in range(self.product3Table.rowCount()):
            for j in range(self.product3Table.columnCount()):
                self.product_stream_3DF.iloc[i, j] = self.product3Table.item(i, j).text()
        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Ag(3)"]
        for r in range(self.product_stream_3DF.shape[0]):
            for s in range(self.product_stream_3DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.product_stream_3DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addprod4fun(self):
        df_top = self.product_stream_4DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_4DF.iloc[self.table_row:, :]

        self.product4Table.setRowCount(self.product_stream_4DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.product4Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.product4Table.setItem(i, j, QTableWidgetItem(value))
                    self.product4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.product4Table.setItem(i, j, QTableWidgetItem(value))
                    self.product4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product4Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.product4Table.setItem(k, j, QTableWidgetItem(value))
                self.product4Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.product4Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_4DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delprod4fun(self):
        df_top = self.product_stream_4DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_4DF.iloc[self.table_row + 1:, :]
        self.product4Table.setRowCount(self.product_stream_4DF.shape[0] - 1)

        for i in range(df_top.shape[0]):
            for j in range(self.product4Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.product4Table.setItem(i, j, QTableWidgetItem(value))
                self.product4Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product4Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.product4Table.setItem(k, j, QTableWidgetItem(value))
                self.product4Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_4DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def saveproduct4fun(self):
        for i in range(self.product4Table.rowCount()):
            for j in range(self.product4Table.columnCount()):
                self.product_stream_4DF.iloc[i, j] = self.product4Table.item(i, j).text()
        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Ag(4)"]
        for r in range(self.product_stream_4DF.shape[0]):
            for s in range(self.product_stream_4DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.product_stream_4DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def addprod5fun(self):
        df_top = self.product_stream_5DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_5DF.iloc[self.table_row:, :]

        self.product5Table.setRowCount(self.product_stream_5DF.shape[0] + 1)

        for i in range(df_top.shape[0] + 1):
            for j in range(self.product5Table.columnCount()):
                if i == self.table_row:
                    value = ""
                    self.product5Table.setItem(i, j, QTableWidgetItem(value))
                    self.product5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

                else:
                    value = str(df_top.iloc[i, j])
                    self.product5Table.setItem(i, j, QTableWidgetItem(value))
                    self.product5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product5Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row + 1
                self.product5Table.setItem(k, j, QTableWidgetItem(value))
                self.product5Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_middle = pd.DataFrame(["" for x in range(self.product5Table.columnCount())]).transpose()
        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_5DF = pd.concat([df_top, df_middle, df_bottom], axis=0, ignore_index=True)

    def delprod5fun(self):
        df_top = self.product_stream_5DF.iloc[:self.table_row, :]
        df_bottom = self.product_stream_5DF.iloc[self.table_row + 1:, :]
        self.product5Table.setRowCount(self.product_stream_5DF.shape[0] - 1)

        for i in range(df_top.shape[0]):
            for j in range(self.product5Table.columnCount()):
                value = str(df_top.iloc[i, j])
                self.product5Table.setItem(i, j, QTableWidgetItem(value))
                self.product5Table.item(i, j).setTextAlignment(QtCore.Qt.AlignCenter)

        for i in range(df_bottom.shape[0]):
            for j in range(self.product5Table.columnCount()):
                value = str(df_bottom.iloc[i, j])
                k = i + self.table_row
                self.product5Table.setItem(k, j, QTableWidgetItem(value))
                self.product5Table.item(k, j).setTextAlignment(QtCore.Qt.AlignCenter)

        df_top = pd.DataFrame(df_top.values.tolist())
        df_bottom = pd.DataFrame(df_bottom.values.tolist())
        self.product_stream_5DF = pd.concat([df_top, df_bottom], axis=0, ignore_index=True)

    def saveproduct5fun(self):
        for i in range(self.product5Table.rowCount()):
            for j in range(self.product5Table.columnCount()):
                self.product_stream_5DF.iloc[i, j] = self.product5Table.item(i, j).text()
        # Save to Feed_Product flows.xlsx
        wb = openpyxl.load_workbook("Feed_Product flows.xlsx")
        ws = wb["Ag(5)"]
        for r in range(self.product_stream_5DF.shape[0]):
            for s in range(self.product_stream_5DF.shape[1]):
                ws.cell(r + 2, s + 1).value = self.product_stream_5DF.iloc[r, s]
        wb.save("Feed_Product flows.xlsx")

    def toHeatLoss(self):
        self.designButton.setStyleSheet(
            "QPushButton#designButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#designButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.heatLossButton.setStyleSheet(
            "QPushButton#heatLossButton{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.initialButton.setStyleSheet(
            "QPushButton#initialButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#initialButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.resultsButton.setStyleSheet(
            "QPushButton#resultsButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#resultsButton::hover{background-color: rgb(255, 0, 0); color: rgb(255, 255, 255); border: 2px solid black}")

        self.designlabel.hide()
        self.heatlosslabel.show()
        self.initiallabel.hide()
        self.complabel.hide()
        self.resultslabel.hide()

        try:
            self.sigmaText.setText(str(self.heatLoss_df.iloc[0,1]))
            self.Ultext.setText(str(self.heatLoss_df.iloc[1,1]))
            self.Uvtext.setText(str(self.heatLoss_df.iloc[2,1]))
            self.Uvwtext.setText(str(self.heatLoss_df.iloc[3,1]))
            self.Ulwtext.setText(str(self.heatLoss_df.iloc[4,1]))
            self.Urtext.setText(str(self.heatLoss_df.iloc[5,1]))
            self.Ubtext.setText(str(self.heatLoss_df.iloc[6,1]))
            self.groundTemptext.setText(str(self.heatLoss_df.iloc[7,1]))
            self.ambTemptext.setText(str(self.heatLoss_df.iloc[8,1]))
            self.roofTemptext.setText(str(self.heatLoss_df.iloc[9,1]))
            self.Uvrtext.setText(str(self.heatLoss_df.iloc[10,1]))
            self.Ulrtext.setText(str(self.heatLoss_df.iloc[11,1]))
            self.refridgeTemptext.setText(str(self.heatLoss_df.iloc[12,1]))

        except:
            pass


    def heatleakchanged(self):
        try:
            self.heatLoss_df.iloc[0,1] = float(self.sigmaText.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[1,1] = float(self.Ultext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[2,1] = float(self.Uvtext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[3,1] = float(self.Uvwtext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[4,1] = float(self.Ulwtext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[5,1] = float(self.Urtext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[6,1] = float(self.Ubtext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[7,1] = float(self.groundTemptext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[8,1] = float(self.roofTemptext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[9,1] = float(self.roofTemptext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[10,1] = float(self.Uvrtext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[11,1] = float(self.Ulrtext.text())
        except:
            pass
        try:
            self.heatLoss_df.iloc[12,1] = float(self.refridgeTemptext.text())
        except:
            pass

    def initialchanged(self):
        try:
            self.jacketLocationList2 = []
            self.jacketLocationList2.append(float(self.jacketStartValue.text()))
            self.jacketLocationList2.append(float(self.jacketEndValue.text()))
            self.jacketLocationList2.append(float(self.Uvrtext.text()))
            self.jacketLocationList2.append(float(self.Ulrtext.text()))
            self.jacketLocationList2.append(float(self.refridgeTemptext.text()))
            self.jacketLocationList = self.jacketLocationList2

        except:
            pass


    def toInitialConditions(self):
        if self.diskTable:
            self.diskTable.setColumnCount(0)
            self.diskTable.setRowCount(0)
            self.diskTable.setGeometry(0,0,0,0)

        self.designButton.setStyleSheet(
            "QPushButton#designButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#designButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.heatLossButton.setStyleSheet(
            "QPushButton#heatLossButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#heatLossButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.initialButton.setStyleSheet(
            "QPushButton#initialButton{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.resultsButton.setStyleSheet(
            "QPushButton#resultsButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#resultsButton::hover{background-color: rgb(255, 0, 0); color: rgb(255, 255, 255); border: 2px solid black}")

        self.diskTable = QTableWidget(self.initiallabel)
        I = int(self.noComponentstext.currentText())
        nodisks = int(self.noLDiskstext.text()) + int(self.noVDiskstext.text())
        length = 270 + 100 * (int(self.noComponentstext.currentText()))
        height = (nodisks + 2) * 40

        if height > 600:
            height = 600

        self.diskTable.setGeometry(np.round((1000 - length) / 2), 250, length, height)
        self.diskTable.setColumnCount(int(self.noComponentstext.currentText()) + 2)
        self.diskTable.setRowCount(nodisks + 2)
        self.diskTable.horizontalHeader().setDefaultSectionSize(100)
        self.diskTable.setColumnWidth(0, 150)
        self.diskTable.setColumnWidth(1, 120)
        self.diskTable.verticalHeader().setDefaultSectionSize(40)
        self.diskTable.verticalHeader().setFixedWidth(100)
        self.diskTable.horizontalHeader().setFixedHeight(60)
        self.diskTableverticalList = ['Disk ' + str(i + 1) for i in range(int(nodisks))]
        self.diskTable.setVerticalHeaderLabels(self.diskTableverticalList)
        self.diskTablehorizontalList = ['Temperature\n(°C)']
        self.compTablehorzList = ['Component {}'.format(str(i + 1)) for i in
                                  range(int(self.noComponentstext.currentText()))]
        self.diskTablehorizontalList = np.concatenate(
            (self.diskTablehorizontalList, self.compTablehorzList)).tolist()
        self.diskTable.setHorizontalHeaderLabels(self.diskTablehorizontalList)
        self.diskTable.horizontalHeader().setStretchLastSection(True)
        self.diskTable.verticalHeader().setStretchLastSection(True)
        self.diskTable.verticalHeader().setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.diskTable.horizontalHeader().setVisible(False)
        self.diskTable.verticalHeader().setVisible(False)
        self.diskTable.horizontalScrollBar().setVisible(False)
        self.diskTable.horizontalScrollBar().setEnabled(False)
        self.diskTable.verticalScrollBar().setVisible(False)
        self.diskTable.verticalScrollBar().setEnabled(False)
        self.diskTable.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.diskTable.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        self.diskTable.setShowGrid(True)
        self.diskTable.setStyleSheet(
            "QTableWidget{border: 2px solid black; gridline-color: black;}"
        )
        self.diskTableLabel.show()
        self.diskTable.show()

        self.diskTable.setSpan(0, 1, 2, 1)
        newItem = QTableWidgetItem("Temperature\n(K)")
        self.diskTable.setItem(0, 1, newItem)
        self.diskTable.item(0, 1).setTextAlignment(QtCore.Qt.AlignCenter)
        self.diskTable.item(0, 1).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.diskTable.item(0, 1).setBackground(QtGui.QColor(230, 255, 255))

        self.diskTable.setSpan(0, 2, 1, I)
        newItem2 = QTableWidgetItem("Component")
        self.diskTable.setItem(0, 2, newItem2)
        self.diskTable.item(0, 2).setTextAlignment(QtCore.Qt.AlignCenter)
        self.diskTable.item(0, 2).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.diskTable.item(0, 2).setBackground(QtGui.QColor(230, 255, 255))

        self.diskTable.setSpan(0, 0, 2, 1)
        newItem3 = QTableWidgetItem("Disk\nNumber")
        self.diskTable.setItem(0, 0, newItem3)
        self.diskTable.item(0, 0).setTextAlignment(QtCore.Qt.AlignCenter)
        self.diskTable.item(0, 0).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
        self.diskTable.item(0, 0).setBackground(QtGui.QColor(230, 255, 255))

        for i in range(I):
            item = QTableWidgetItem(str(i + 1))
            self.diskTable.setItem(1, i + 2, item)
            self.diskTable.item(1, i + 2).setTextAlignment(QtCore.Qt.AlignCenter)
            self.diskTable.item(1, i + 2).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
            self.diskTable.item(1, i + 2).setBackground(QtGui.QColor(230, 255, 255))

        for i in range(int(self.noLDiskstext.text())):
            item = QTableWidgetItem("Liquid Disk {}".format(str(i + 1)))
            self.diskTable.setItem(i + 2, 0, item)
            self.diskTable.item(i + 2, 0).setTextAlignment(QtCore.Qt.AlignCenter)
            self.diskTable.item(i + 2, 0).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
            self.diskTable.item(i + 2, 0).setBackground(QtGui.QColor(230, 255, 255))

        for i in range(int(self.noVDiskstext.text())):
            number = i + int(self.noLDiskstext.text())
            item = QTableWidgetItem("Vapour Disk {}".format(str(number + 1)))
            self.diskTable.setItem(number + 2, 0, item)
            self.diskTable.item(number + 2, 0).setTextAlignment(QtCore.Qt.AlignCenter)
            self.diskTable.item(number + 2, 0).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
            self.diskTable.item(number + 2, 0).setBackground(QtGui.QColor(230, 255, 255))

        if int(self.noComponentstext.currentText()) > 5:
            self.diskTable.horizontalScrollBar().setEnabled(True)
            self.diskTable.horizontalScrollBar().setVisible(True)
            self.diskTable.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        if nodisks > 13:
            self.diskTable.setGeometry(np.round((1000 - length) / 2), 250, length, height)
            self.diskTable.verticalScrollBar().setEnabled(True)
            self.diskTable.verticalScrollBar().setVisible(True)
            self.diskTable.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

        # Populate Disk Table
        for i in range(self.diskTable.rowCount() - 2):
            for j in range(self.diskTable.columnCount() - 1):
                label = QLineEdit()
                label.setStyleSheet("QLineEdit { background-color: white; }"
                                    "QLineEdit { border: none; }")
                label.textChanged.connect(self.saveDiskTable)
                label.setAlignment(QtCore.Qt.AlignCenter)
                self.diskTable.setCellWidget(i + 2, j + 1, label)
                self.diskTable.cellWidget(i + 2, j + 1).setFont(self.compTablefont)
                try:
                    label.setText(str(self.diskinitCombined[i, j]))
                except:
                    pass

        self.diskTable.show()

        self.designlabel.hide()
        self.heatlosslabel.hide()
        self.initiallabel.show()
        self.complabel.hide()
        self.resultslabel.hide()

    def diskTablefun(self):
        self.compTablefont = QFont('Arial', 10)

        def isint(value):
            try:
                int(value)
                return True
            except ValueError:
                return False


        if self.noLDiskstext.text() != "" and self.noVDiskstext.text() != "":
            if isint(self.noLDiskstext.text()) and isint(self.noVDiskstext.text()):
                if int(self.noLDiskstext.text()) >= 0 and int(self.noVDiskstext.text()) >= 0:
                    nodisks = int(self.noLDiskstext.text()) + int(self.noVDiskstext.text())

                    try:
                        self.diskTableLabel.hide()
                        self.diskTable.hide()
                        self.diskTable.setColumnCount(0)
                        self.diskTable.setRowCount(0)
                    except:
                        pass

                    I = int(self.noComponentstext.currentText())
                    if nodisks > 0:
                        arrangedName = ['Component {}'.format(str(i+1)) for i in range(I)]
                        # T0 = pd.read_excel("base_disk_initial_conditions.xlsx", usecols="A").values.tolist()
                        # T0 = np.array(T0)
                        # init_DF = pd.read_excel("base_disk_initial_conditions.xlsx", "Data")
                        # x0 = np.zeros((len(T0), I))
                        # for i in range(I):
                        #     x0[:, i] = init_DF[arrangedName[i]].values.tolist()
                        # self.diskinitCombined = np.concatenate((T0, x0), axis=1)

                        # extra_rows = nodisks - self.diskinitCombined.shape[0]

                        # if extra_rows > 0:
                        #     empty_array = np.zeros((extra_rows, self.diskinitCombined.shape[1]))
                        #     self.diskinitCombined = np.concatenate((self.diskinitCombined, empty_array), axis=0)

                        # length = 270 + 100 * (int(self.noComponentstext.currentText()))
                        # height = (nodisks+2) * 40

                        # if height > 600:
                        #     height = 600

                        # self.diskTable.setGeometry(np.round((1000-length)/2), 250, length, height)
                        self.diskTable.setColumnCount(int(self.noComponentstext.currentText()) + 2)
                        self.diskTable.setRowCount(nodisks+2)
                        self.diskTable.horizontalHeader().setDefaultSectionSize(100)
                        self.diskTable.setColumnWidth(0, 150)
                        self.diskTable.setColumnWidth(1, 120)
                        self.diskTable.verticalHeader().setDefaultSectionSize(40)
                        self.diskTable.verticalHeader().setFixedWidth(100)
                        self.diskTable.horizontalHeader().setFixedHeight(60)
                        self.diskTableverticalList = ['Disk ' + str(i + 1) for i in range(int(nodisks))]
                        self.diskTable.setVerticalHeaderLabels(self.diskTableverticalList)
                        self.diskTablehorizontalList = ['Temperature\n(°C)']
                        self.compTablehorzList = ['Component {}'.format(str(i + 1)) for i in
                                                  range(int(self.noComponentstext.currentText()))]
                        self.diskTablehorizontalList = np.concatenate(
                            (self.diskTablehorizontalList, self.compTablehorzList)).tolist()
                        self.diskTable.setHorizontalHeaderLabels(self.diskTablehorizontalList)
                        self.diskTable.horizontalHeader().setStretchLastSection(True)
                        self.diskTable.verticalHeader().setStretchLastSection(True)
                        self.diskTable.verticalHeader().setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                        self.diskTable.horizontalHeader().setVisible(False)
                        self.diskTable.verticalHeader().setVisible(False)
                        self.diskTable.horizontalScrollBar().setVisible(False)
                        self.diskTable.horizontalScrollBar().setEnabled(False)
                        self.diskTable.verticalScrollBar().setVisible(False)
                        self.diskTable.verticalScrollBar().setEnabled(False)
                        self.diskTable.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
                        self.diskTable.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
                        self.diskTable.setShowGrid(True)
                        self.diskTable.setStyleSheet(
                            "QTableWidget{border: 2px solid black; gridline-color: black;}"
                        )
                        self.diskTableLabel.show()
                        self.diskTable.show()

                        self.diskTable.setSpan(0, 1, 2, 1)
                        newItem = QTableWidgetItem("Temperature\n(K)")
                        self.diskTable.setItem(0, 1, newItem)
                        self.diskTable.item(0, 1).setTextAlignment(QtCore.Qt.AlignCenter)
                        self.diskTable.item(0, 1).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                        self.diskTable.item(0, 1).setBackground(QtGui.QColor(230, 255, 255))

                        self.diskTable.setSpan(0, 2, 1, I)
                        newItem2 = QTableWidgetItem("Component")
                        self.diskTable.setItem(0, 2, newItem2)
                        self.diskTable.item(0, 2).setTextAlignment(QtCore.Qt.AlignCenter)
                        self.diskTable.item(0, 2).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                        self.diskTable.item(0, 2).setBackground(QtGui.QColor(230, 255, 255))

                        self.diskTable.setSpan(0, 0, 2, 1)
                        newItem3 = QTableWidgetItem("Disk\nNumber")
                        self.diskTable.setItem(0, 0, newItem3)
                        self.diskTable.item(0, 0).setTextAlignment(QtCore.Qt.AlignCenter)
                        self.diskTable.item(0, 0).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                        self.diskTable.item(0, 0).setBackground(QtGui.QColor(230, 255, 255))

                        for i in range(I):
                            item = QTableWidgetItem(str(i+1))
                            self.diskTable.setItem(1, i+2, item)
                            self.diskTable.item(1, i+2).setTextAlignment(QtCore.Qt.AlignCenter)
                            self.diskTable.item(1, i+2).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                            self.diskTable.item(1, i+2).setBackground(QtGui.QColor(230, 255, 255))

                        for i in range(int(self.noLDiskstext.text())):
                            item = QTableWidgetItem("Liquid Disk {}".format(str(i + 1)))
                            self.diskTable.setItem(i+2, 0, item)
                            self.diskTable.item(i+2, 0).setTextAlignment(QtCore.Qt.AlignCenter)
                            self.diskTable.item(i+2, 0).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                            self.diskTable.item(i+2, 0).setBackground(QtGui.QColor(230, 255, 255))

                        for i in range(int(self.noVDiskstext.text())):
                            number = i + int(self.noLDiskstext.text())
                            item = QTableWidgetItem("Vapour Disk {}".format(str(number + 1)))
                            self.diskTable.setItem(number+2, 0, item)
                            self.diskTable.item(number+2, 0).setTextAlignment(QtCore.Qt.AlignCenter)
                            self.diskTable.item(number+2, 0).setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                            self.diskTable.item(number+2, 0).setBackground(QtGui.QColor(230, 255, 255))

                        if int(self.noComponentstext.currentText()) > 5:
                            self.diskTable.horizontalScrollBar().setEnabled(True)
                            self.diskTable.horizontalScrollBar().setVisible(True)
                            self.diskTable.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

                        if nodisks > 13:
                            self.diskTable.setGeometry(np.round((1000-length)/2), 250, length, height)
                            self.diskTable.verticalScrollBar().setEnabled(True)
                            self.diskTable.verticalScrollBar().setVisible(True)
                            self.diskTable.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOn)

                        # Populate Disk Table
                        for i in range(self.diskTable.rowCount()-2):
                            for j in range(self.diskTable.columnCount()-1):
                                label = QLineEdit()
                                label.setStyleSheet("QLineEdit { background-color: white; }"
                                                    "QLineEdit { border: none; }")
                                label.textChanged.connect(self.saveDiskTable)
                                label.setAlignment(QtCore.Qt.AlignCenter)
                                self.diskTable.setCellWidget(i+2, j+1, label)
                                self.diskTable.cellWidget(i+2, j+1).setFont(self.compTablefont)
                                try:
                                    label.setText(str(self.diskinitCombined[i, j]))
                                except:
                                    pass

                        self.diskTable.show()


    def saveDiskTable(self):
        try:
            # Save to locally stored variable.
            for i in range(self.diskTable.rowCount()-2):
                for j in range(self.diskTable.columnCount()-1):
                    widget = self.diskTable.cellWidget(i+2, j+1)
                    if isinstance(widget, QLineEdit):
                        text = widget.text()
                        self.diskinitCombined[i, j] = text

            # Save to Initial condition.xlsx
            wb = openpyxl.load_workbook("Initial condition.xlsx")
            ws = wb["Data"]
            for r in range(len(self.diskinitCombined)):
                for s in range(len(self.diskinitCombined[1])):
                    ws.cell(r+2, s+1).value = self.diskinitCombined[r,s]
            wb.save("Initial condition.xlsx")

        except:
            pass
            # self.disksaveerrorWindow = QFrame()
            # length = 720
            # heightTable = 60
            # self.disksaveerrorWindow.setFixedSize(length, heightTable + 50)
            # self.disksaveerrorWindow.setWindowTitle('Disk Table Save Error')
            # self.disksaveerrorWindow.setStyleSheet("QTabBar{ font: 10pt Arial}"
            #                                "QTabBar::tab { height: 50px;}")
            #
            # self.disksaveerrorWindowLabel = QLabel(self.disksaveerrorWindow)
            # self.disksaveerrorWindowLabel.setGeometry(5, 0, 710, 60)
            # self.disksaveerrorWindowLabel.setObjectName("errorWindowLabel")
            # self.disksaveerrorWindowLabel.setText("The table cannot be saved because one or more cells are not of float type.")
            # self.disksaveerrorWindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))
            # self.disksaveerrorWindow.show()

    def resetfun(self):
        if self.designButton.palette().window().color().name() == '#0000ff':
            self.tankDiametertext.setText("")
            self.tankHeighttext.setText("")
            self.initialHeighttext.setText("")
            self.initialPressuretext.setText("")
            for key in self.feedLocationValueDict:
                self.feedLocationValueDict[key].setText("")
            for key in self.productLocationValueDict:
                self.productLocationValueDict[key].setText("")
            self.noComponentstext.setCurrentText("")
            self.noDiskstext.setText("")
            self.noLDiskstext.setText("")
            self.abstoltext.setText("")
            self.reltoltext.setText("")
            self.numberofIterationstext.setText("")

    def data_preprocessing(self):
        def isfloat(value):
            try:
                float(value)
                return True
            except ValueError:
                return False
        ## Preprocessing Ag ['time', 'Ag', 'Bg']
        noProductStream = 0
        for key in self.productLocationValueDict:
            if isfloat(self.productLocationValueDict[key].text()) and float(self.productLocationValueDict[key].text()) <= 100 and \
                    float(self.productLocationValueDict[key].text()) >= 0:
                noProductStream += 1
        Ag_keys = ['Ag({})'.format(str(i + 1)) for i in range(noProductStream)]
        self.Ag_dict = dict.fromkeys(Ag_keys, 0)

        for j in range(noProductStream):
            df = pd.read_excel('Feed_Product flows.xlsx', "Ag({})".format(str(j + 1)))
            time_len = int(df.iloc[-1, 0] / 5) + 1
            time = [5 * i for i in range(time_len)]
            Ag_list = [None] * len(time)
            Bg_list = [0] * len(time)
            for i in range(len(time)):
                if time[i] in df.iloc[:, 0].values:
                    Ag_list[i] = df[df.iloc[:, 0] == time[i]].iloc[:, 1].values.tolist()[0]
                else:
                    k = i
                    while time[k] not in df.iloc[:, 0].values:
                        k = k - 1
                    Ag_list[i] = Ag_list[k]

            for i in range(len(Ag_list) - 1):
                if Ag_list[i] != Ag_list[i + 1]:
                    Bg_list[i] = (Ag_list[i + 1] - Ag_list[i]) / (5 * 60)
                else:
                    pass
            Ag_df = pd.DataFrame([time, Ag_list, Bg_list]).T
            Ag_df.columns = ['Time(min)', 'Ag({})'.format(j + 1), 'Bg({})'.format(j + 1)]
            self.Ag_dict[Ag_keys[j]] = Ag_df

        # Pre-processing Af(1)
        noFeedStream = 0
        for key in self.feedLocationValueDict:
            if isfloat(self.feedLocationValueDict[key].text()) and float(self.feedLocationValueDict[key].text()) <= 100 and \
                    float(self.feedLocationValueDict[key].text()) >= 0:
                noFeedStream += 1

        Af_keys = ['Af({})'.format(str(i + 1)) for i in range(noFeedStream)]
        AZ_keys = ['AZ({})'.format(str(i + 1)) for i in range(noFeedStream)]
        BZ_keys = ['BZ({})'.format(str(i + 1)) for i in range(noFeedStream)]
        self.Af_dict = dict.fromkeys(Af_keys, 0)
        self.AZ_dict = dict.fromkeys(AZ_keys, 0)
        self.BZ_dict = dict.fromkeys(BZ_keys, 0)

        for j in range(noFeedStream):
            print(j)
            df = pd.read_excel('Feed_Product flows.xlsx', "Af({})".format(str(j + 1)))
            time_len = int(df.iloc[-1, 0] / 5) + 1
            time = [5 * i for i in range(time_len)]
            Af_list = [None] * len(time)
            ATf_list = [None] * len(time)
            APf_list = [None] * len(time)
            noOfComponents = int(self.noComponentstext.currentText())
            componentName = ['Z{}'.format(str(i + 1)) for i in range(int(self.noComponentstext.currentText()))]
            AZ_dict = dict.fromkeys(componentName, 0)
            for key in AZ_dict:
                AZ_dict[key] = ['None'] * len(time)

            for i in range(len(time)):
                if time[i] in df.iloc[:, 0].values:
                    Af_list[i] = df[df.iloc[:, 0] == time[i]].iloc[:, 1].values.tolist()[0]
                    ATf_list[i] = df[df.iloc[:, 0] == time[i]].iloc[:, 2].values.tolist()[0]
                    APf_list[i] = df[df.iloc[:, 0] == time[i]].iloc[:, 3].values.tolist()[0]
                    for key in AZ_dict:  # Z1, Z2, Z3,...
                        AZ_dict[key][i] = df[df.iloc[:, 0] == time[i]][key].values.tolist()[0]
                else:
                    k = i
                    while time[k] not in df.iloc[:, 0].values:
                        k = k - 1
                    Af_list[i] = Af_list[k]
                    ATf_list[i] = ATf_list[k]
                    APf_list[i] = APf_list[k]
                    for key in AZ_dict:
                        AZ_dict[key][i] = AZ_dict[key][k]

            Bf_list = [0] * len(time)
            BTf_list = [0] * len(time)
            BPf_list = [0] * len(time)

            BZ_dict = dict.fromkeys(componentName, 0)

            for key in BZ_dict:
                BZ_dict[key] = [0] * len(time)
            for i in range(len(Af_list) - 1):
                if Af_list[i] != Af_list[i + 1]:
                    Bf_list[i] = (Af_list[i + 1] - Af_list[i]) / (5 * 60)
                if ATf_list[i] != ATf_list[i + 1]:
                    BTf_list[i] = (ATf_list[i + 1] - ATf_list[i]) / 5
                if APf_list[i] != APf_list[i + 1]:
                    BPf_list[i] = (APf_list[i + 1] - APf_list[i]) / 5

                for key in BZ_dict:
                    if AZ_dict[key][i] != AZ_dict[key][i + 1]:
                        BZ_dict[key][i] = (AZ_dict[key][i + 1] - AZ_dict[key][i]) / 5

            Af_df = pd.DataFrame([time, Af_list, Bf_list, ATf_list, BTf_list, APf_list, BPf_list])
            Af_df = Af_df.T
            Af_df.columns = ['Time(min)', 'Af({})'.format(str(j + 1)), 'Bf({})'.format(str(j + 1)),
                             'ATf({})'.format(str(j + 1)),
                             'BTf({})'.format(str(j + 1)), 'APf({})'.format(str(j + 1)), 'BPf({})'.format(str(j + 1))]
            self.Af_dict[Af_keys[j]] = Af_df

            AZ_df_list = [AZ_dict[key] for key in AZ_dict]
            AZ_df = pd.DataFrame(AZ_df_list).T
            AZ_df.columns = componentName
            self.AZ_dict[AZ_keys[j]] = AZ_df

            BZ_df_list = [BZ_dict[key] for key in BZ_dict]
            BZ_df = pd.DataFrame(BZ_df_list).T
            BZ_df.columns = componentName
            self.BZ_dict[BZ_keys[j]] = BZ_df

    #  checkfun ensures all user input variables are of the required data type.
    def checkfun(self):
        def isfloat(value):
            try:
                float(value)
                return True
            except ValueError:
                return False

        def isint(value):
            try:
                int(value)
                return True
            except ValueError:
                return False

        self.errorList = []

        value = 0
        non_empty_value = 0
        for key in self.feedLocationValueDict:
            if self.feedLocationValueDict[key].text() != "":
                non_empty_value = value
            value+=1

        empty_value = 0
        value = 0
        for key in self.feedLocationValueDict:
            if self.feedLocationValueDict[key].text() == "":
                empty_value = value
                if empty_value < non_empty_value:
                    string = "Feed streams are not in order."
                    self.errorList.append(string)
                    break
            value += 1

        inner_value = 0
        outer_value = 0
        for key in self.feedLocationValueDict:
            if self.feedLocationValueDict[key].text() == "N. A.":
                for key in self.feedLocationValueDict:
                    if isfloat(self.feedLocationValueDict[key].text()):
                        if float(self.feedLocationValueDict[key].text()) >= 0 and float(self.feedLocationValueDict[key].text()) <= 100 and \
                                inner_value > outer_value:
                            string = "Feed Streams are not in order."
                            self.errorList.append(string)
                            break
                    inner_value += 1
                    print(inner_value)
            outer_value += 1
            break

        value = 0
        for key in self.feedLocationValueDict:
            if self.feedLocationValueDict[key].text() != 'N. A.':
                if isfloat(self.feedLocationValueDict[key].text()):
                    if float(self.feedLocationValueDict[key].text()) >= 0 and float(self.feedLocationValueDict[key].text()) <= 100:
                        print("Feed location {} is a float between 0 and 100. Please proceed.".format(str(value+1)))
                    elif float(self.feedLocationValueDict[key].text()) < 0 or float(self.feedLocationValueDict[key].text()) > 100:
                        string = "Feed location {} is not a float between 0 and 100.".format(str(value+1))
                        self.errorList.append(string)
                else:
                    string = "Feed location {} is not a float between 0 and 100.".format(str(value + 1))
                    self.errorList.append(string)
            value += 1

        value = 0
        non_empty_value = 0
        for key in self.productLocationValueDict:
            if self.productLocationValueDict[key].text() != "":
                non_empty_value = value
            value += 1

        empty_value = 0
        value = 0
        for key in self.productLocationValueDict:
            if self.productLocationValueDict[key].text() == "":
                empty_value = value
                if empty_value < non_empty_value:
                    string = "Product streams are not in order."
                    self.errorList.append(string)
                    break
            value+=1

        inner_value = 0
        outer_value = 0
        for key in self.productLocationValueDict:
            if self.productLocationValueDict[key].text() == "N. A.":
                for key in self.productLocationValueDict:
                    if isfloat(self.productLocationValueDict[key].text()):
                        if float(self.productLocationValueDict[key].text()) >= 0 and float(
                                self.productLocationValueDict[key].text()) <= 100 and \
                                inner_value > outer_value:
                            string = "Product Streams are not in order."
                            self.errorList.append(string)
                            break
                    inner_value += 1
                    print(inner_value)
            outer_value += 1
            break

        value = 0
        for key in self.productLocationValueDict:
            if self.productLocationValueDict[key].text() != 'N. A.':
                if isfloat(self.productLocationValueDict[key].text()):
                    if float(self.productLocationValueDict[key].text()) >= 0 and float(
                            self.productLocationValueDict[key].text()) <= 100:
                        print("Product location {} is a float between 0 and 100. Please proceed.".format(str(value + 1)))
                    elif float(self.productLocationValueDict[key].text()) < 0 or float(
                            self.productLocationValueDict[key].text()) > 100:
                        string = "Product location {} is not a float between 0 and 100.".format(str(value + 1))
                        self.errorList.append(string)
                else:
                    string = "Product location {} is not a float between 0 and 100.".format(str(value + 1))
                    self.errorList.append(string)
            value += 1

        # value = 0
        # for key in self.productLocationValueDict:
        #     if self.productLocationValueDict[key].text() != "":
        #         try:
        #             sheet = 'Ag({})'.format(str(value + 1))
        #             df = pd.read_excel("Feed_Product flows.xlsx", sheet_name=sheet)
        #
        #         except:
        #             string = "Height for product stream {} is given but sheet 'Ag({})' in Feed_Product flows.xlsx is not found.".format(
        #                 str(value + 1), str(value + 1))
        #             self.errorList.append(string)
        #     value += 1


        if isfloat(self.tankDiametertext.text()):
            if float(self.tankDiametertext.text()) > 0:
                print('Tank Diameter is a float. Please proceed.')
            else:
                string = 'Tank Diameter is not a positive float number.'
                self.errorList.append(string)
        else:
            string1 = 'Tank Diameter is not of float number type.'
            self.errorList.append(string1)

        # try:
        #     value = 0
        #     for key in self.feedLocationValueDict:
        #        if isfloat(self.feedLocationValueDict[key].text()) and float(self.feedLocationValueDict[key].text()) <= 100:
        #            print('Feed Location ' + str(value+1) + ' is a float number <= 100. Please proceed.')
        #        else:
        #            string = "Feed Location {} is not a float number <= 100.".format(str(value+1))
        #            self.errorList.append(string)
        #        value += 1
        #
        # except:
        #     string = 'Input for feed stream location is incomplete or not saved.'
        #     self.errorList.append(string)
        #
        # try:
        #     value = 0
        #     for key in self.productLocationValueDict:
        #        if isfloat(self.productLocationValueDict[key].text()) and float(self.productLocationValueDict[key].text()) <= 100:
        #            print('Product Location ' + str(value+1) + ' is a float number <= 100. Please proceed.')
        #        else:
        #            string = "Product Location {} is not a float number <= 100.".format(str(value+1))
        #            self.errorList.append(string)
        #        value += 1
        #
        # except:
        #     string = 'Input for product stream location is incomplete or not saved.'
        #     self.errorList.append(string)

        if isfloat(self.tankHeighttext.text()):
            if float(self.tankHeighttext.text()) > 0:
                print('Tank Height is a float. Please proceed.')
            else:
                string = 'Tank Height is not a positive float number.'
                self.errorList.append(string)
        else:
            string1 = 'Tank Height is not of float type.'
            self.errorList.append(string1)

        if isfloat(self.initialPressuretext.text()):
            if float(self.initialPressuretext.text()) > 0:
                print('Initial Tank Pressure is a float. Please proceed.')
            else:
                string = 'Initial Tank Pressure is not a positive float number.'
                self.errorList.append(string)
        else:
            string1 = 'Initial Tank Pressure is not of float type.'
            self.errorList.append(string1)

        if isfloat(self.initialHeighttext.text()):
            if float(self.initialHeighttext.text()) >= 0 and float(self.initialHeighttext.text()) <= 100:
                print('Initial Liquid Height is a float <= 100. Please proceed.')
            else:
                string = 'Initial Liquid Height is not a positive float number <= 100.'
                self.errorList.append(string)
        else:
            string1 = 'Initial Liquid Height is not of float type.'
            self.errorList.append(string1)

        if isint(self.noLDiskstext.text()):
            if int(self.noLDiskstext.text()) >= 0:
                print('Number of liquid disks is a positive integer. Please proceed.')
            else:
                string = 'Number of liquid disks is not a positive integer.'
                self.errorList.append(string)
        else:
            string3 = 'Number of liquid disks is not an integer type.'
            self.errorList.append(string3)

        if isint(self.noVDiskstext.text()):
            if int(self.noVDiskstext.text()) >= 0:
                print('Number of vapour disks is a positive integer. Please proceed.')
            else:
                string = 'Number of vapour disks is not a positive integer.'
                self.errorList.append(string)
        else:
            string3 = 'Number of vapour disks is not an integer type.'
            self.errorList.append(string3)


        if isfloat(self.abstoltext.text()):
            if float(self.abstoltext.text()) >= 0 and float(self.abstoltext.text()) <= 1:
                print('Absolute tolerance is a positive float type. Please proceed.')
            else:
                string = 'Absolute tolerance is not a positive integer between 0 and 1.'
                self.errorList.append(string)
        else:
            string3 = 'Absolute tolerance is not a positive integer between 0 and 1.'
            self.errorList.append(string3)

        if isfloat(self.reltoltext.text()):
            if float(self.reltoltext.text()) >= 0 and float(self.reltoltext.text()) <= 1:
                print('Relative tolerance is a positive float type. Please proceed.')
            else:
                string = 'Relative tolerance is not a positive integer between 0 and 1.'
                self.errorList.append(string)
        else:
            string3 = 'Relative tolerance is not a positive integer between 0 and 1.'
            self.errorList.append(string3)

        if isint(self.numberofIterationstext.text()):
            if int(self.numberofIterationstext.text()) >= 0 and (int(self.numberofIterationstext.text()) % 5 == 0):
                print('Running time is divisible by 5. Please proceed.')
            else:
                string = 'Running time is not a positive integer divisible by 5.'
                self.errorList.append(string)
        else:
            string3 = 'Running time is not a positive integer divisible by 5.'
            self.errorList.append(string3)

        if isfloat(self.groundTemptext.text()):
            print('Ground Temperature is a float. Please proceed.')
        else:
            string = "Ground Temperature is not a float."
            self.errorList.append(string)

        if isfloat(self.ambTemptext.text()):
            print('Ambient Temperature is a float. Please proceed.')
        else:
            string = "Ambient Temperature is not a float."
            self.errorList.append(string)

        if isfloat(self.roofTemptext.text()):
            print('Roof Temperature is a float. Please proceed.')
        else:
            string = "Roof Temperature is not a float."
            self.errorList.append(string)

        if isfloat(self.sigmaText.text()) and float(self.sigmaText.text()) > 0:
            print('Evaporation Coefficient is a float. Please proceed.')
        else:
            string = "Evaporation Coefficient is not a float."
            self.errorList.append(string)

        if isfloat(self.Ultext.text()) and float(self.Ultext.text()) > 0:
            print('Liquid-Phase Film Heat Transfer Coefficient is a float. Please proceed.')
        else:
            string = "Liquid-Phase Film Heat Transfer Coefficient is not a float."
            self.errorList.append(string)

        if isfloat(self.Uvtext.text()) and float(self.Uvtext.text()) > 0:
            print('Vapour-Phase Film Heat Transfer Coefficient is a float. Please proceed.')
        else:
            string = "Vapour-Phase Film Heat Transfer Coefficient is not a float."
            self.errorList.append(string)

        if isfloat(self.Uvwtext.text()) and float(self.Uvwtext.text()) > 0:
            print('Wall-Vapour Heat Transfer Coefficient is a float. Please proceed.')
        else:
            string = "Wall-Vapour Heat Transfer Coefficient is not a float."
            self.errorList.append(string)

        if isfloat(self.Ulwtext.text()) and float(self.Ulwtext.text()) > 0:
            print('Liquid-Vapour Heat Transfer Coefficient is a float. Please proceed.')
        else:
            string = "Liquid-Vapour Heat Transfer Coefficient is not a float."
            self.errorList.append(string)

        if isfloat(self.Urtext.text()) and float(self.Urtext.text()) > 0:
            print('Tank Roof-Vapour Heat Transfer Coefficient is a float. Please proceed.')
        else:
            string = "Tank Roof-Vapour Heat Transfer Coefficient is not a float."
            self.errorList.append(string)

        if isfloat(self.Ubtext.text()) and float(self.Ubtext.text()) > 0:
            print('Tank Bottom-Liquid Heat Transfer Coefficient is a float. Please proceed.')
        else:
            string = "Tank Bottom-Liquid Heat Transfer Coefficient is not a float."
            self.errorList.append(string)

        try:
            optionalList = ['Jacket Start Location', 'Jacket End Location',
                            'Jacket-Vapour Heat Transfer Coefficient',
                            'Jacket-Liquid Heat Transfer Coefficient', 'Refrigerant Temperature']
            for i in range(len(self.jacketLocationList)):
                if i == 0:
                    if isfloat(self.jacketLocationList[0]):
                        if float(self.jacketLocationList[0]) > float(self.jacketLocationList[1]):
                            string = "Value of jacket start point is greater than the value of jacket end point."
                            self.errorList.append(string)

                if i == 0 or i == 1:
                    if isfloat(self.jacketLocationList[i]):
                        if float(self.jacketLocationList[i]) <= 100 and float(self.jacketLocationList[i]) >= 0:
                            print('Location for jacket {} is valid. Please proceed.'.format(str(i + 1)))
                        else:
                            string = 'Location for jacket {} is a float but it is negative or greater than 1.'.format(
                                str(i + 1))
                            self.errorList.append(string)
                    else:
                        string = 'Location for jacket {} is a not of the required float type.'.format(str(i + 1))
                        self.errorList.append(string)
                elif i == 2 or i == 3:
                    if isfloat(self.jacketLocationList[i]):
                        if float(self.jacketLocationList[i]) > 0:
                            print(optionalList[i] + ' is valid. Please proceed.')
                        else:
                            string = optionalList[i] + ' is not a positive float number.'
                            self.errorList.append(string)
                    else:
                        string = optionalList[i] + ' is not a positive float number.'
                        self.errorList.append(string)
                elif i == 4:
                    if isfloat(self.jacketLocationList[i]):
                        print(optionalList[i] + ' is a float number. Please proceed.')
                    else:
                        string = optionalList[i] + ' is not a positive float number.'
                        self.errorList.append(string)
        except:
            string = 'Jacket location is incomplete or not saved.'
            self.errorList.append(string)

        # checking disk initial conditions
        try:
            noDisks = int(self.noLDiskstext.text()) + int(self.noVDiskstext.text())
            self.diskinitCombined = self.diskinitCombined[:noDisks, :int(self.noComponentstext.currentText())+1]
            diskNumberList = ['Disk {}'.format(i + 1) for i in range(len(self.diskinitCombined))]
            componentList = ['Component {}'.format(j + 1) for j in range(int(self.noComponentstext.currentText()))]
            temperatureVector = self.diskinitCombined[:, 0]
            compVector = self.diskinitCombined[:, 1:]
            for i in range(len(temperatureVector)):
                if isfloat(temperatureVector[i]):
                    if float(temperatureVector[i]) > 0:
                        print('Temperature of ' + diskNumberList[i] + ' is a float number. Please proceed.')
                    else:
                        string = 'Temperature of ' + diskNumberList[i] + ' is not a float number.'
                        self.errorList.append(string)
            for i in range(len(compVector)):
                for j in range(len(compVector[0])):
                    if isfloat(compVector[i, j]):
                        if float(compVector[i, j]) >= 0 and float(compVector[i, j]) <= 1:
                            print(diskNumberList[i] + ' Composition for ' + componentList[
                                j] + ' is between 0 and 1. Please proceed.')
                        else:
                            string = diskNumberList[i] + ' Composition for ' + componentList[
                                j] + ' is between 0 and 1.'
                            self.errorList.append(string)
                    else:
                        string = diskNumberList[i] + ' Composition for ' + componentList[j] + ' is between 0 and 1.'
                        self.errorList.append(string)
            compSum = compVector.sum(axis=1)
            for i in range(len(compSum)):
                if compSum[i] == 1:
                    print('Composition summation for ' + diskNumberList[i] + ' is valid. Please proceed.')
                else:
                    string = 'Composition summation for ' + diskNumberList[i] + ' is not equal to 1.'
                    self.errorList.append(string)
        except:
            string = 'Initial disk conditions is incomplete or not saved.'
            self.errorList.append(string)

        # Checking molecular weight
        noComp = int(self.noComponentstext.currentText())
        MW_list = self.MW_list[:noComp]
        for i in range(len(MW_list)):
            if MW_list[i] == "":
                string = 'Molecular weight for component {} is empty.'.format(str(i+1))
                self.errorList.append(string)
                break
            if isfloat(MW_list[i]):
                if float(MW_list[i]) > 0:
                    print('Molecular weight for component {} is of correct float type.'.format(str(i+1)))
                else:
                    string = 'Molecular weight for component {} is negative or zero.'.format(str(i+1))
                    self.errorList.append(string)
            else:
                string = 'Molecular weight for component {} is not of float type.'.format(str(i+1))
                self.errorList.append(string)


    def analysefun(self):
        # Check whether the respective inputs are of the correct data type.
        self.checkfun()

        if self.errorList != []:
            print(self.errorList)
            self.errorWindow = QFrame()
            length = 800
            heightTable = 50 * len(self.errorList) + 10
            self.errorWindow.setFixedSize(length, heightTable + 50)
            self.errorWindow.setWindowTitle('List of Errors')
            self.errorWindow.setStyleSheet("QTabBar{ font: 10pt Arial}"
                                           "QTabBar::tab { height: 50px;}")
            self.errorWindowLabel = QLabel(self.errorWindow)
            self.errorWindowLabel.setGeometry(5, 0, 800, 50)
            self.errorWindowLabel.setObjectName("errorWindowLabel")
            self.errorWindowLabel.setText("The program cannot be run due to the following errors:")
            self.errorWindowLabel.setFont(QFont('Arial', 12, weight=QtGui.QFont.Bold))

            error_keys = ['Error {}'.format(str(i + 1)) for i in range(len(self.errorList))]
            self.errorLabelDict = dict.fromkeys(error_keys, 0)
            value = 1

            for key in self.errorLabelDict:
                self.errorLabelDict[key] = QLabel(self.errorWindow)
                self.errorLabelDict[key].setGeometry(5, 45 * value + 20, 800, 40)
                self.errorLabelDict[key].setText("• " + self.errorList[value - 1])
                self.errorLabelDict[key].setFont(QFont('Arial', 10, weight=QtGui.QFont.Bold))
                value += 1

            self.errorWindow.show()

        else:
            QGuiApplication.processEvents()
            self.progressWindow.show()
            self.progressWindowtext.setText("Preprocessing Data...")
            self.progressWindowtext.show()
            now = datetime.now()
            current_time = now.strftime("%H:%M:%S")
            print("Started at", current_time)
            QGuiApplication.processEvents()
            # Obtain arranged list from Component Table
            I = int(self.noComponentstext.currentText())

            arrangedVP = vapourPressureList[1:I+1]
            arrangedVP_index = [nN_name.index(x) - 1 for x in arrangedVP]  # Because there is white space in first index

            arrangedLD = liquidDensityList[1:I+1]
            arrangedLD_index = [nN_name.index(x) - 1 for x in arrangedLD]

            arrangedLE = enthalpyLiqList[1:I+1]
            arrangedLE_index = [nN_name.index(x) - 1 for x in arrangedLE]

            arrangedVE = enthalpyVapList[1:I+1]
            arrangedVE_index = [nN_name.index(x) - 1 for x in arrangedVE]

            arrangedLH = specificheatLiqList[1:I+1]
            arrangedLH_index = [nN_name.index(x) - 1 for x in arrangedLH]

            arrangedVH = specificheatVapList[1:I+1]
            arrangedVH_index = [nN_name.index(x) - 1 for x in arrangedVH]
            ## Tank dimensions
            D = float(self.tankDiametertext.text())  ## Tank Diameter (m)
            H = float(self.tankHeighttext.text())  ## Tank Height (m)
            A = 0.25 * math.pi * (D ** 2)

            ## Feed stream parameters
            F = 0 # Number of Feed Streams
            self.feedLocationList = []

            def isfloat(value):
                try:
                    float(value)
                    return True
                except ValueError:
                    return False

            for key in self.feedLocationValueDict:
                if isfloat(self.feedLocationValueDict[key].text()) and float(self.feedLocationValueDict[key].text()) <= 100 and \
                        float(self.feedLocationValueDict[key].text()) >= 0:
                    self.feedLocationList.append(self.feedLocationValueDict[key].text())
                    F += 1

            nf = [float(x) for x in self.feedLocationList if x != ""]
            print(nf)
            nf = [x * H/100 for x in nf]  # feed stream location (Feed, LP, HP)

            ## Product stream parameters
            S = 0 # total number of product streams
            self.productLocationList = []
            for key in self.productLocationValueDict:
                if isfloat(self.productLocationValueDict[key].text()) and float(
                        self.productLocationValueDict[key].text()) <= 100 and \
                        float(self.productLocationValueDict[key].text()) >= 0:
                    self.productLocationList.append(self.productLocationValueDict[key].text())
                    S += 1

            ng = [float(x) for x in self.productLocationList if x != ""]
            ng = [x * H/100 for x in ng]  # Product stream location (Send-out, VRL, BOG)

            if F > 0 or S > 0:
                self.data_preprocessing()
                print('Data Preprocessing Done! Please Proceed.')

            ai = 1
            bi = 1
            Sp = []

            ## Fluid Properties and Constants
            g = 9.81  ## gravitational acceleration (m2/s)
            R = 8.314  ## gas constant
            # Z = float(self.compressFactortext.text())  # Compressibility Factor
            MW = [float(x) for x in self.MW_list]  # component molecular weights (kg/kmol)

            ## Fixed DAE model parameters, scale factors, and reference
            N = int(self.noLDiskstext.text()) + int(self.noVDiskstext.text())
            Tref = 100  # Scale factor for temperature
            LDref = 400  # Reference liquid density (kg/m3)
            VDref = 2  # Reference vapor density (kg/m3)
            Lref = A * H * LDref / N
            Vref = A * H * VDref / N
            NL = int(self.noLDiskstext.text())
            NV = N - NL  # number of vapor disks
            NC = N + N * I
            NE = NC + N - 2
            sigma = float(self.sigmaText.text()) * 1000  # evaporation model coefficient (2E-5,5E-5,1E-4)

            ## Temperatures, heat transfer coefficients, and other parameters
            # global Tamb, Tb, Tj, Tr, Ul, Uv, Ur, Ui, Uvw, Ulw, Ub, Uvr, Ulr, nJ
            Tb = float(self.groundTemptext.text())  # Ground temperature (K)
            Tamb = [float(self.ambTemptext.text()), 0, 0]  # ambient temperature as a function of t (K)
            Tr = float(self.roofTemptext.text())  # roof temperature (K)
            Ul = float(self.Ultext.text())  # liquid phase film heat transfer coefficient (W/m2 K)
            Uv = float(self.Uvtext.text())  # vapor phase film heat transfer coefficient (W/m2 K)
            Ui = (Ul * Uv) / (Ul + Uv)  # interface heat transfer coefficient (W/m2 K)
            Uvw = float(self.Uvwtext.text())  # wall-vapor heat transfer coefficient (W/m2 K)
            Ulw = float(self.Ulwtext.text())  # wall-liquid heat transfer coefficient (W/m2 K)
            Ur = float(self.Urtext.text())  # tank roof-vapor heat transfer coefficient (W/m2 K)
            Ub = float(self.Ubtext.text())  # tank bottom-liquid heat transfer coefficient (W/m2 K)
            ## Optional
            nJ = [float(self.jacketStartValue.text())/100, float(self.jacketEndValue.text())/100]  # jacket location
            Uvr = float(self.Uvrtext.text())  # jacket-vapor heat transfer coefficient (W/m2 K)
            Ulr = float(self.Ulrtext.text())  # jacket-liquid side heat transfer coefficient (W/m2 K)
            Tj = float(self.refridgeTemptext.text())  # refrigerant temperature (K)


            def liq_prop(T):
                E = np.zeros((len(T), I))
                C = np.zeros((len(T), I))
                LD = np.zeros((len(T), I))
                for j in range(len(T)):
                    for i in range(I):
                        if len(T) > 1:
                            E[j, i] = my_module[arrangedLE_index[i]].ANN(np.atleast_2d(T[j]).T)[0]
                            C[j, i] = my_module[arrangedLH_index[i]].ANN(np.atleast_2d(T[j]).T)[0]
                            LD[j, i] = my_module[arrangedLD_index[i]].ans(T[j][0])
                        elif len(T) == 1:
                            E[j, i] = my_module[arrangedLE_index[i]].ANN([T])[0]
                            C[j, i] = my_module[arrangedLH_index[i]].ANN([T])[0]
                            LD[j, i] = my_module[arrangedLD_index[i]].ans(T[j])
                return E, C, LD

            def vap_prop(T):
                Ei = np.zeros((len(T), I))
                Ci = np.zeros((len(T), I))
                for j in range(len(T)):
                    for i in range(I):
                        if len(T) > 1:
                            Ei[j, i] = my_module[arrangedVE_index[i]].ANN(np.atleast_2d(T[j]).T)[0]
                            Ci[j, i] = my_module[arrangedVH_index[i]].ANN(np.atleast_2d(T[j]).T)[0]
                        elif len(T) == 1:
                            Ei[j, i] = my_module[arrangedVE_index[i]].ANN([T])[0]
                            Ci[j, i] = my_module[arrangedVH_index[i]].ANN([T])[0]
                return Ei, Ci

            def Vap_Pressure(T):
                P_s = np.zeros((len(T), I))
                for j in range(len(T)):
                    for i in range(I):
                        P_s[j, i] = my_module[arrangedVP_index[i]].ans(T[j])
                return P_s

            ## Initial conditions in the tank
            h0 = float(self.initialHeighttext.text()) * H/100  # liquid height (m) from a percentage of H
            P0 = float(self.initialPressuretext.text())  # tank pressure (kPa)
            self.noDisks = int(self.noLDiskstext.text()) + int(self.noVDiskstext.text())
            self.diskinitCombined = self.diskinitCombined[:self.noDisks, :int(self.noComponentstext.currentText())+1]
            T0 = self.diskinitCombined[:,0].reshape(-1,1)
            x0 = self.diskinitCombined[:, 1:]
            d0 = np.zeros((N, 1))  # disk density
            Ei, Ci, Di = liq_prop(T0[:NL])
            d0[:NL, 0] = 1 / np.sum((x0[:NL, :] / Di), axis=1)
            Pg = P0
            self.Pg = Pg
            AN1 = [x for x in x0[N-1,:]]
            AN2 = [Pg]
            AN3 = T0[N-1].tolist()
            AN = AN1 + AN2 + AN3
            AN = np.array(AN).reshape(1,-1)
            Z1 = my_module[nN_name.index('ModelZg') - 1].ans(AN)
            Z = Z1[0][0]
            d0[NL:N, 0] = P0 / (Z * R * T0[N - 1]) / np.sum(x0[N - 1, :] / MW)
            W0 = np.zeros((I, N))
            W0[:, :NL] = d0[:NL, 0] * np.transpose(x0[:NL, :]) * (A * h0 / NL / Lref)
            W0[:, NL:N] = d0[NL:N, 0] * np.transpose(x0[NL:N, :]) * (A * (H - h0) / NV / Vref)
            y0 = np.zeros((NE, 1))
            y0[:N] = (T0 / Tref).tolist()
            y0[N:NC] = np.reshape(np.transpose(W0), (len(W0) * len(W0[0]), 1))

            ## Main Program
            t = [0]
            y = y0.reshape(1, -1)[0]  # horizontal list

            if F > 0:
                Fi = [self.Af_dict[key].iloc[0, 1] for key in self.Af_dict]
                FT = [self.Af_dict[key].iloc[0, 3] for key in self.Af_dict]
                FP = [self.Af_dict[key].iloc[0, 5] for key in self.Af_dict]
                Fz = []

            def mass(u):
                # global H, A, R, g, Z, MW, I, N, NL, NV, NE, NC, Tref, Vref, Lref, Ad, Bd, Cd, bi, Sp, Tmin, Tmax
                ### Model Variables
                u = np.array(u)
                u = u.reshape(-1, 1)
                T = u[0:N] * Tref
                Wn = np.transpose(np.reshape(u[N:NC], (int((NC - N) / I), I)))
                Wn[:, :NL] = Wn[:, :NL] * Lref
                Wn[:, NL:N] = Wn[:, NL:N] * Vref
                sumw = Wn.sum(axis=0)
                x = Wn / sumw

                ### Enthalpy calculations
                Ei = np.zeros((N, I))
                Ci = np.zeros((N, I))
                EL, CL, LD = liq_prop(T[:NL])
                EV, CV = vap_prop(T[NL:N])
                Ei[:NL, :] = EL
                Ei[NL:N, :] = EV
                Ei = Ei.T
                Ci[:NL, :] = CL
                Ci[NL:N, :] = CV

                CP = np.zeros((1, N))
                CP[0, 0:NL] = (x[:, :NL] * np.transpose(Ci[:NL, :])).sum(axis=0)
                CP[0, NL:N] = (x[:, NL:N] * np.transpose(Ci[NL:N, :])).sum(axis=0)

                ## Mass Matrix
                Mass = np.zeros((NE, NE))
                # Energy Balance
                for i in range(N):
                    if i <= NL - 1:
                        Mass[i, i] = sumw[i] * CP[0, i] / Lref
                    else:
                        Mass[i, i] = sumw[i] * CP[0, i] / Vref
                    p = i + 1
                    lowerlimit = N + (p - 1) * I
                    upperlimit = N + (p) * I
                    for j in range(lowerlimit, upperlimit):
                        k = j - (N + i * I)
                        Mass[i, j] = Ei[k, i] / Tref

                # Component Balance
                for i in range(N, NC):
                    Mass[i, i] = 1

                return Mass

            def tankfunc(u, t):
                # global H, A, R, g, Z, MW, I, N, NL, NV, NE, NC, Tref, Vref, Lref, Ad, Bd, Cd, bi, Sp, Tmin, Tmax, Pmax, \
                #     Pmin, Tamb, Tb, Tj, Tr, Ul, Uv, Ur, Ui, Uvw, Ulw, Ub, Uvr, Ulr, nJ, Ap, Bp, Cp, Dp, sigma, F, nf, ATf, APf, BTf, \
                #     BPf, CTf, CPf, Cf, CZf, Af, Bf, AZf, BZf, Ag, Bg, Cg, ng, S, D, Tj, ai
                global ai
                u = np.array(u)
                # u = np.atleast_2d(u)
                u = u.reshape(-1, 1)
                dydt = np.zeros((NE, 1))
                ## Ambient Temperature
                t0 = (t / (24 * 3600)) % 24
                Ta = Tamb[0] + Tamb[1] * t0 + Tamb[2] * t0 ** 2
                # Model Variables
                T = u[0:N] * Tref
                Wn = np.transpose(np.reshape(u[N:NC], (int((NC - N) / I), I)))
                Wn[:, :NL] = Wn[:, :NL] * Lref
                Wn[:, NL:N] = Wn[:, NL:N] * Vref
                sumw = Wn.sum(axis=0)
                x = Wn / sumw
                wsq = np.zeros((N - 1, 1))
                wsq[:NL - 1] = u[NC:NC + NL - 1] * 0.01 * Lref
                wsq[NL:N - 1] = u[NC + NL - 1:NE] * 0.01 * Vref

                ### Density & Pressure
                rho = np.zeros((N, 1))
                P = np.zeros((N, 1))
                Mw = np.zeros((N, 1))
                EL, CL, LD = liq_prop(T[:NL])
                rho[:NL, 0] = 1 / (np.transpose(x[:, :NL]) / LD).sum(axis=1)
                h = NL * sumw[0] / A / rho[0, 0]  # liquid level
                Mw[:, 0] = 1 / (np.transpose(x[:, :N]) / MW).sum(axis=1)

                P1 = np.zeros((N-NL, 1))
                P1[:] = self.Pg
                zf = np.transpose(x[:, NL:N])
                AN = np.concatenate((zf, P1, T[NL:N].reshape(-1,1)), axis=1)
                BN = my_module[nN_name.index('ModelZg') - 1].ans(AN)
                BN = np.maximum(BN, 0)
                BN_list = BN.reshape(1, -1).tolist()[0]
                Z = np.zeros((N,1))
                Z[NL:N, 0] = BN_list
                PV = (NV * Z[N-1] * R * T[N - 1, 0] * sumw[N - 1]) / (A * (H - h) * Mw[N - 1, 0])
                self.Pg = PV
                rho1 = PV * Mw[NL:N, 0]
                rho2 = (Z[NL:N, 0] * R * T[NL:N, 0])
                rho[NL:N, 0] = rho1/rho2
                P[NL:N, 0] = PV
                for i in range(NL):
                    P[i] = PV + (g * h / NL) * (0.5 * rho[i] + rho[i + 1:NL].sum()) / 1000
                ## Enthalpy Calculations
                Ei = np.zeros((N, I))
                EL, CL, LD = liq_prop(T[:NL])
                EV, CV = vap_prop(T[NL:N])
                Ei[:NL, :] = EL
                Ei[NL:N, :] = EV
                E = np.zeros((1, N))
                E[0, 0:NL] = (Ei[:NL, :] * np.transpose(x[:, :NL])).sum(axis=1)
                E[0, NL:N] = (Ei[NL:N, :] * np.transpose(x[:, NL:N])).sum(axis=1)
                ## Interface Calculations
                T_I = (Ul * T[NL - 1, 0] + Uv * T[NL, 0]) / (Ul + Uv)
                X = np.transpose(x[:, NL - 1]) / MW / (np.transpose(x[:, NL - 1]) / MW).sum(axis=0)
                Y = np.transpose(x[:, NL]) / MW / (np.transpose(x[:, NL]) / MW).sum(axis=0)
                Ps = Vap_Pressure([T_I])
                K = np.maximum(Ps / PV, 1 * 10 ** (-6))
                ## Evaporation Flux Calculation
                J = sigma * A * PV * (np.minimum(K * X, 1) - Y) * np.sqrt(MW)
                e = np.maximum(J, 0)
                c = np.maximum(J * (-1), 0)
                TI_A = np.array([T_I])
                EL, CL, LD = liq_prop(TI_A)
                EV, CV = vap_prop(TI_A)
                sumc = np.matmul(c, np.transpose(EL))[0][0]
                sume = np.matmul(e, np.transpose(EV))[0][0]
                wsq[NL - 1] = np.sum(J)

                ### Flash stream
                ## Liquid and vapour distribution
                mf = np.zeros((F, 1))
                for r in range(F):
                    if nf[r] > h:
                        mf[r, 0] = np.round((nf[r] - h) / ((H - h) / NV)) + NL
                    else:
                        mf[r, 0] = np.maximum(np.round(nf[r] / (h / NL)), 1)

                ## Flowrate, Temperature and Vapor Fraction
                if F > 0:
                    Tf = np.array([np.array(ATf), np.array(BTf) * t]).sum(axis=0)
                    Pf = np.array([np.array(APf), np.array(BPf) * t]).sum(axis=0)
                    Fi = np.array([np.array(Af), np.array(Bf) * t]).sum(axis=0)
                    zf = np.array([np.array(AZf), np.array(BZf) * t]).sum(axis=0)
                    Ff = np.transpose(np.reshape(Fi, (-1, len(Fi)))) * zf
                    Pf = Pf.reshape(-1, 1)
                    mf_1 = mf.reshape(1, -1)[0].tolist()
                    mf_1 = [int(i)-1 for i in mf_1]
                    P_mf = P[mf_1]
                    Tf = Tf.reshape(-1, 1)
                    AN = np.concatenate((zf, Pf, P_mf, Tf), axis=1)
                    BN = np.maximum(my_module[nN_name.index('ModelF') - 1].ANN(AN), 0)
                    T2 = BN[:, 0]
                    vf = np.minimum(BN[:, 1:], zf)
                    fv = vf * Ff
                    fl = Ff - fv
                    T2_1 = T2.reshape(-1, 1)
                    EL, Ci, Di = liq_prop(T2_1)
                    EV, Ci = vap_prop(T2_1)
                    fel = np.sum((fl * EL), axis=1)
                    fev = np.sum((fv * EV), axis=1)
                    fel = fel.reshape(-1, 1)  # column vector
                    fev = fev.reshape(-1, 1)  # column vector

                ## Liquid and Vapour Distributions
                ff = np.zeros((N, F))
                for r in range(F):
                    if mf[r, 0] > NL:
                        xp = [i for i in range(NL)]
                        ldf = poisson.pmf(xp, mu=1)
                        yp = [i for i in range(N - int(mf[r, 0]) + 1)]
                        vdf = poisson.pmf(yp, mu=1)
                        xp = [i for i in range(int(mf[r, 0]) - NL)]
                        xpf = poisson.pmf(xp, mu=1)
                        for i in range(N):
                            if i <= NL - 1:
                                ff[i, r] = ldf[NL - i - 1]
                            elif i > NL - 1 and i < int(mf[r, 0]) - 1:
                                ff[i, r] = xpf[int(mf[r, 0]) - i - 2]
                            else:
                                # print(i, vdf[i-int(mf.iloc[r, 0])])
                                ff[i, r] = vdf[i - int(mf[r, 0]) + 1]

                    elif mf[r, 0] == NL:
                        xp = [i for i in range(NL)]
                        ldf = poisson.pmf(xp, mu=1)
                        yp = [i for i in range(N - NL)]
                        vdf = poisson.pmf(yp, mu=1)
                        for i in range(N):
                            if i <= NL - 1:
                                ff[i, r] = ldf[NL - i - 1]
                            else:
                                ff[i, r] = vdf[i - NL]

                    else:
                        xp = [i for i in range(int(mf[r, 0]))]
                        ldf = poisson.pmf(xp, mu=1)
                        xp = [i for i in range(NL - int(mf[r, 0]))]
                        vdf = poisson.pmf(xp, mu=1)
                        for i in range(N):
                            if i <= int(mf[r, 0]) - 1:
                                ff[i, r] = ldf[int(mf[r, 0]) - i - 1]
                            elif i > int(mf[r, 0]) - 1 and i <= NL - 1:
                                ff[i, r] = vdf[i - int(mf[r, 0])]
                            else:
                                ff[i, r] = 0

                ff[:NL, :] = ff[:NL, :] / ff[:NL, :].sum(axis=0)
                ff[NL:N, :] = ff[NL:N, :] / ff[NL:N, :].sum(axis=0)
                ff = np.maximum(ff, 0)
                ff = np.nan_to_num(ff)

                # ## Send Out Stream
                if S > 0:
                    Gi = np.array(Ag) + np.array(Bg) * t
                    mg = np.zeros((S, 1))
                    for s in range(S):
                        if ng[s] > h:
                            mg[s, 0] = np.round((ng[s] - h) / ((H - h) / NV)) + NL
                        else:
                            mg[s, 0] = np.maximum(np.round(ng[s] / (h / NL)), 1)

                    pf = np.zeros((N, S))
                    for s in range(S):
                        if mg[s, 0] <= NL:
                            xp = [i for i in range(int(mg[s, 0]))]
                            ldf = poisson.pmf(xp, mu=1)
                            xp = [i for i in range(NL - int(mg[s, 0]))]
                            vdf = poisson.pmf(xp, mu=1)
                            for i in range(N):
                                if i <= mg[s, 0] - 1:
                                    pf[i, s] = ldf[int(mg[s, 0]) - i - 1]
                                elif i > mg[s, 0] - 1 and i <= NL - 1:
                                    pf[i, s] = vdf[i - int(mg[s, 0])]
                                else:
                                    pf[i, s] = 0

                        else:
                            xp = [i for i in range(int(mg[s, 0]) - NL)]
                            ldf = poisson.pmf(xp, mu=1)
                            xp = [i for i in range(N - int(mg[s, 0] + 1))]
                            vdf = poisson.pmf(xp, mu=1)
                            for i in range(N):
                                if i <= NL - 1:
                                    pf[i, s] = 0
                                elif i > NL - 1 and i <= mg[s, 0] - 1:
                                    pf[i, s] = ldf[int(mg[s, 0]) - i - 1]
                                else:
                                    pf[i, s] = vdf[i - int(mg[s, 0])]

                    pf[:NL, :] = pf[:NL, :] / pf[:NL, :].sum(axis=0)
                    pf[NL:N, :] = pf[NL:N, :] / pf[NL:N, :].sum(axis=0)
                    pf = np.maximum(pf, 0)
                    pf = np.nan_to_num(pf)
                    sumg = np.matmul(pf, np.transpose(Gi))

                ### Heat Transfer Calculations
                ## Jacket Location
                mJ = np.zeros((2, 1))
                for l in range(2):
                    if nJ[l] > h - 1:
                        mJ[l, 0] = np.round((nJ[l] - h) / ((H - h) / NV)) + NL
                    else:
                        mJ[l, 0] = np.maximum(np.round(nJ[l] / (h / NL)), 1)

                ## Heat leak from surrounding
                Q = np.zeros((N, 1))
                q = np.zeros((N, 1))
                for i in range(N):
                    if mJ[0, 0] < NL:
                        if mJ[1, 0] + 1 <= NL:
                            if i >= mJ[0, 0]-1 and i <= mJ[1, 0]-1:
                                Q[i, 0] = Ulr * math.pi * D * h * (Tj - T[i, 0]) / NL
                            elif i > NL - 1:
                                Q[i, 0] = Uvw * math.pi * D * (H - h) * (Ta - T[i, 0]) / NV
                            else:
                                Q[i, 0] = Ulw * math.pi * D * h * (Ta - T[i, 0]) / NL

                        else:
                            if i < mJ[0, 0] - 1:
                                Q[i, 0] = Ulw * math.pi * D * h * (Ta - T[i, 0]) / NL
                            elif i >= mJ[0, 0] - 1 and i <= NL - 1:
                                Q[i, 0] = Ulr * math.pi * D * h * (Tj - T[i, 0]) / NL
                            elif i > NL - 1 and i <= mJ[1, 0] - 1:
                                Q[i, 0] = Uvr * math.pi * D * (H - h) * (Tj - T[i, 0]) / NV
                            else:
                                Q[i, 0] = Uvr * math.pi * D * (H - h) * (Ta - T[i, 0]) / NV
                    else:
                        if i >= mJ[0, 0] - 1 and i <= mJ[1, 0] - 1:
                            Q[i, 0] = Uvr * math.pi * D * (H - h) * (Tj - T[i, 0]) / NV
                        elif i <= NL - 1:
                            Q[i, 0] = Ulw * math.pi * D * h * (Ta - T[i, 0]) / NL
                        else:
                            Q[i, 0] = Uvw * math.pi * D * (H - h) * (Ta - T[i, 0]) / NV

                # Heat Transfer across Disks
                q[0, 0] = Ub * A * (Tb - T[0, 0]) - Ul * A * (T[0, 0] - T[1, 0])
                q[1:NL - 1, 0] = Ul * A * (T[:NL - 2, 0] - T[1:NL - 1, 0]) - Ul * A * (T[1:NL - 1, 0] - T[2:NL, 0])
                q[NL - 1, 0] = Ul * A * (T[NL - 2, 0] - T[NL - 1, 0]) - Ui * A * (T[NL - 1, 0] - T[NL, 0])
                q[NL, 0] = Ui * A * (T[NL - 1, 0] - T[NL, 0]) - Uv * A * (T[NL, 0] - T[NL + 1, 0])
                q[NL + 1:N - 1, 0] = Uv * A * (T[NL:N - 2, 0] - T[NL + 1:N - 1, 0]) - Uv * A * (
                        T[NL + 1:N - 1, 0] - T[NL + 2:N, 0])
                q[N - 1, 0] = Uv * A * (T[N - 2, 0] - T[N - 1, 0]) + Ur * A * (Tr - T[N - 1, 0])

                ## Energy Balance
                if F > 0 and S > 0:
                    for i in range(NL):
                        if i == 0:
                            dydt[i] = np.matmul(ff[i, :], fel[:, 0]) - sumg[i] * E[0, i] - np.maximum(wsq[i, 0], 0) * E[
                                0, i] + \
                                      np.maximum(wsq[i, 0] * -1, 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]
                        elif i == NL - 1:
                            dydt[i] = np.matmul(ff[i, :], fel[:, 0]) - sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[
                                0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] + sumc - sume + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = np.matmul(ff[i, :], fel[:, 0]) - sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[
                                0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[
                                          0, i] + np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + \
                                      q[i, 0] + Q[i, 0]

                elif F == 0 and S > 0:
                    for i in range(NL):
                        if i == 0:
                            dydt[i] = -sumg[i] * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * -1, 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]
                        elif i == NL - 1:
                            dydt[i] = -sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] + sumc - sume + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = -sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] \
                                      + np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]

                elif F > 0 and S == 0:
                    for i in range(NL):
                        if i == 0:
                            dydt[i] = np.matmul(ff[i, :], fel[:, 0]) - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * -1, 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]
                        elif i == NL - 1:
                            dydt[i] = np.matmul(ff[i, :], fel[:, 0]) + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] + sumc - sume + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = np.matmul(ff[i, :], fel[:, 0]) + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] \
                                      + np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]

                elif F == 0 and S == 0:
                    for i in range(NL):
                        if i == 0:
                            dydt[i] = - np.maximum(wsq[i, 0], 0) * E[0, i] + np.maximum(wsq[i, 0] * -1, 0) * E[0, i + 1]\
                                      + q[i, 0] + Q[i, 0]
                        elif i == NL - 1:
                            dydt[i] = np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] + sumc - sume + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * -1, 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] \
                                      + np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]

                dydt[:NL] = dydt[:NL] / (Lref * Tref)

                if F > 0 and S > 0:
                    for i in range(NL, N):
                        if i == NL:
                            dydt[i] = np.matmul(ff[i, :], fev[:, 0]) - sumg[i] * E[0, i] - np.maximum(wsq[i, 0], 0) * E[
                                0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + sume + sumc + q[i, 0] + Q[i, 0]
                        elif i == N - 1:
                            dydt[i] = np.matmul(ff[i, :], fev[:, 0]) - sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[
                                0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = np.matmul(ff[i, :], fev[:, 0]) - sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[
                                0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]

                elif F > 0 and S == 0:
                    for i in range(NL, N):
                        if i == NL:
                            dydt[i] = np.matmul(ff[i, :], fev[:, 0]) - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + sume + sumc + q[i, 0] + Q[i, 0]
                        elif i == N - 1:
                            dydt[i] = np.matmul(ff[i, :], fev[:, 0]) + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = np.matmul(ff[i, :], fev[:, 0]) + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]

                elif F == 0 and S > 0:
                    for i in range(NL, N):
                        if i == NL:
                            dydt[i] = -sumg[i] * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + sume + sumc + q[i, 0] + Q[i, 0]
                        elif i == N - 1:
                            dydt[i] = -sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = -sumg[i] * E[0, i] + np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]

                elif F == 0 and S == 0:
                    for i in range(NL, N):
                        if i == NL:
                            dydt[i] = -np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + sume + sumc + q[i, 0] + Q[i, 0]
                        elif i == N - 1:
                            dydt[i] = np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] + q[i, 0] + Q[i, 0]
                        else:
                            dydt[i] = np.maximum(wsq[i - 1, 0], 0) * E[0, i - 1] - \
                                      np.maximum(wsq[i - 1, 0] * (-1), 0) * E[0, i] - np.maximum(wsq[i, 0], 0) * E[0, i] + \
                                      np.maximum(wsq[i, 0] * (-1), 0) * E[0, i + 1] + q[i, 0] + Q[i, 0]

                dydt[NL:N] = dydt[NL:N] / (Vref * Tref)

                # Component Balance
                num1 = int(N + (NL * I))
                if F > 0 and S > 0:
                    for i in range(N, num1):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == 1:
                            dydt[i] = np.matmul(ff[n - 1, :], fl[:, j - 1]) - sumg[n - 1] * x[j - 1, n - 1] - np.maximum(
                                wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n]
                        elif n == NL:
                            dydt[i] = np.matmul(ff[n - 1, :], fl[:, j - 1]) - sumg[n - 1] * x[j - 1, n - 1] - e[0][j - 1] + \
                                      c[0][j - 1] + np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = np.matmul(ff[n - 1, :], fl[:, j - 1]) - sumg[n - 1] * x[j - 1, n - 1] - np.maximum(
                                wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + np.maximum(wsq[n - 2, 0], 0) * x[
                                          j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                elif F > 0 and S == 0:
                    for i in range(N, num1):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == 1:
                            dydt[i] = np.matmul(ff[n - 1, :], fl[:, j - 1]) - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n]
                        elif n == NL:
                            dydt[i] = np.matmul(ff[n - 1, :], fl[:, j - 1]) - e[0][j - 1] + \
                                      c[0][j - 1] + np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = np.matmul(ff[n - 1, :], fl[:, j - 1]) - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + \
                                      np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                elif F == 0 and S > 0:
                    for i in range(N, num1):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == 1:
                            dydt[i] = -sumg[n - 1] * x[j - 1, n - 1] - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n]
                        elif n == NL:
                            dydt[i] = -sumg[n - 1] * x[j - 1, n - 1] - e[0][j - 1] + c[0][j - 1] + \
                                      np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = -sumg[n - 1] * x[j - 1, n - 1] - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + \
                                      np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                elif F == 0 and S == 0:
                    for i in range(N, num1):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == 1:
                            dydt[i] = -np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n]
                        elif n == NL:
                            dydt[i] = -e[0][j - 1] + c[0][j - 1] + np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = -np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + \
                                      np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                dydt[N:num1] = dydt[N:num1] / Lref

                if F > 0 and S > 0:
                    for i in range(num1, NC):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == NL + 1:
                            dydt[i] = np.matmul(ff[n - 1, :], fv[:, j - 1]) - sumg[n - 1] * x[j - 1, n - 1] - np.maximum(
                                wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + e[0][j - 1] - c[0][j - 1]
                        elif n == N:
                            dydt[i] = np.matmul(ff[n - 1, :], fv[:, j - 1]) - sumg[n - 1] * x[j - 1, n - 1] + np.maximum(
                                wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = np.matmul(ff[n - 1, :], fv[:, j - 1]) - sumg[n - 1] * x[j - 1, n - 1] - np.maximum(
                                wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + np.maximum(wsq[n - 2, 0], 0) * x[
                                          j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                elif F > 0 and S == 0:
                    for i in range(num1, NC):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == NL + 1:
                            dydt[i] = np.matmul(ff[n - 1, :], fv[:, j - 1]) - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + e[0][j - 1] - c[0][j - 1]
                        elif n == N:
                            dydt[i] = np.matmul(ff[n - 1, :], fv[:, j - 1]) + np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = np.matmul(ff[n - 1, :], fv[:, j - 1]) - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + \
                                      np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                elif F == 0 and S > 0:
                    for i in range(num1, NC):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == NL + 1:
                            dydt[i] = -sumg[n - 1] * x[j - 1, n - 1] - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + e[0][j - 1] - c[0][j - 1]
                        elif n == N:
                            dydt[i] = -sumg[n - 1] * x[j - 1, n - 1] + np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = -sumg[n - 1] * x[j - 1, n - 1] - np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + \
                                      np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                elif F == 0 and S == 0:
                    for i in range(num1, NC):
                        i2 = i + 1
                        n = int(1 + np.floor((i2 - 1 - N) / I))
                        j = int(i2 - N - (n - 1) * I)
                        if n == NL + 1:
                            dydt[i] = -np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + e[0][j - 1] - c[0][j - 1]
                        elif n == N:
                            dydt[i] = np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]
                        else:
                            dydt[i] = -np.maximum(wsq[n - 1, 0], 0) * x[j - 1, n - 1] + \
                                      np.maximum(wsq[n - 1, 0] * (-1), 0) * x[j - 1, n] + \
                                      np.maximum(wsq[n - 2, 0], 0) * x[j - 1, n - 2] - \
                                      np.maximum(wsq[n - 2, 0] * (-1), 0) * x[j - 1, n - 1]

                dydt[num1:NC] = dydt[num1:NC] / Vref

                ## Interdisk flows
                dydt[NC:NC + NL - 1] = np.transpose(sumw[1:NL][np.newaxis]) / sumw[0] - rho[1:NL] / rho[0, 0]
                dydt[NC + NL - 1:NE] = np.transpose(sumw[NL:N - 1][np.newaxis]) / sumw[N - 1] - rho[NL:N - 1] / rho[
                    N - 1, 0]

                # ai = ai + 1
                TE = np.matmul(sumw, np.transpose(E))[0]

                return dydt

            def f(t, u, du):
                du_array = np.atleast_2d(du).T
                M = mass(u)
                rhs_du = tankfunc(u, t)
                lhs_du = np.matmul(M, du_array)
                resid = rhs_du - lhs_du
                resid = resid.tolist()
                resid = [x[0] for x in resid]
                resid = np.array(resid)
                return resid

            numberofIterations = int(int(self.numberofIterationstext.text())/5)
            for r in range(numberofIterations):
                print(r+1)
                if r == 0:
                    QGuiApplication.processEvents()
                    self.progressWindowtext.setText(
                        "Analysing data in progress...\n\n({} out of {} iterations completed)"
                        .format(str(r), numberofIterations))
                    QGuiApplication.processEvents()
                Af = []
                Bf = []
                AZf = []
                BZf = []
                ATf = []
                BTf = []
                APf = []
                BPf = []
                if F > 0:
                    for i in range(F):
                        Af.append(self.Af_dict['Af({})'.format(str(i + 1))]['Af({})'.format(str(i + 1))][r])
                        Bf.append(self.Af_dict['Af({})'.format(str(i + 1))]['Bf({})'.format(str(i + 1))][r])
                        AZ_df = self.AZ_dict['AZ({})'.format(str(i + 1))]
                        AZf.append(AZ_df.iloc[r, :].values.tolist())
                        BZ_df = self.BZ_dict['BZ({})'.format(str(i + 1))]
                        BZf.append(BZ_df.iloc[r, :].values.tolist())
                        ATf.append(self.Af_dict['Af({})'.format(str(i + 1))]['ATf({})'.format(str(i + 1))][r])
                        BTf.append(self.Af_dict['Af({})'.format(str(i + 1))]['BTf({})'.format(str(i + 1))][r])
                        APf.append(self.Af_dict['Af({})'.format(str(i + 1))]['APf({})'.format(str(i + 1))][r])
                        BPf.append(self.Af_dict['Af({})'.format(str(i + 1))]['BPf({})'.format(str(i + 1))][r])

                Ag = []
                Bg = []
                if S > 0:
                    for i in range(S):
                        Ag.append(self.Ag_dict['Ag({})'.format(str(i + 1))]['Ag({})'.format(str(i + 1))][r])
                        Bg.append(self.Ag_dict['Ag({})'.format(str(i + 1))]['Bg({})'.format(str(i + 1))][r])
                        # Ag.append(self.Ag_dict['Ag({})'.format(str(i + 1))]['Ag({})'.format(str(i + 1))][0])
                        # Bg.append(self.Ag_dict['Ag({})'.format(str(i + 1))]['Bg({})'.format(str(i + 1))][0])
                p = np.ones((1, NE))
                tspan = (0.0, 300.0)

                if r == 0:
                    y0 = [y0[i, 0] for i in range(len(y0))]
                else:
                    y0 = y[-1, :].reshape(1, -1)[0]

                try:
                    dy0 = np.matmul(np.linalg.pinv(mass(y0)), tankfunc(y0, 0))
                    dy0 = [dy0[i, 0] for i in range(len(dy0))]

                    nondiffEq = NE-8
                    abstol = float(self.abstoltext.text())
                    reltol = float(self.reltoltext.text())
                    differential_vars = np.append(np.ones((1, nondiffEq), dtype=bool), np.zeros((1, 8), dtype=bool))
                    differential_vars = [x for x in differential_vars]

                    model = Implicit_Problem(f, y0, dy0, 0)
                    sim = IDA(model)
                    sim.atol = abstol
                    sim.rtol = reltol
                    sim.algvar = differential_vars
                    tfinal = 300.0
                    ncp = 10
                    t2, y2, yd = sim.simulate(tfinal, ncp)

                    t2 = np.atleast_2d(t2).T

                    ## Feed Flash Calculations
                    if F > 0:
                        if r == 0:
                            Fi = np.array(Fi).reshape(1, -1)
                        Af_term = Af + Bf * t2[1:]
                        Fi = np.append(Fi, Af_term, axis=0)

                        if r == 0:
                            FT = np.array(FT).reshape(1, -1)
                        ATf_term = ATf + BTf * t2[1:]
                        FT = np.append(FT, ATf_term, axis=0)

                        if r == 0:
                            FP = np.array(FP).reshape(1, -1)
                        APf_term = APf + BPf * t2[1:]
                        FP = np.append(FP, APf_term, axis=0)

                        if r == 0:
                            dFz = np.zeros((F, len(t2), I))
                            for j in range(F):
                                for p in range(len(t2)):
                                    dFz[j, p, :] = AZf[j] + BZf[j] * t2[p]
                        else:
                            dFz = np.zeros((F, len(t2) - 1, I))
                            for j in range(F):
                                for p in range(len(t2) - 1):
                                    dFz[j, p, :] = AZf[j] + BZf[j] * t2[p + 1]

                        dFz = np.transpose(dFz, (1, 0, 2))
                        if r == 0:
                            Fz = dFz
                        else:
                            Fz = np.append(Fz, dFz, axis=0)

                    if r == 0:
                        t = t2
                        y = y2
                    else:
                        t2 = t2[1:] + t[-1]
                        y2 = y2[1:, :]
                        t = np.append(t, t2, axis=0)
                        y = np.append(y, y2, axis=0)

                    self.progressWindowtext.setText("Analysing data in progress...\n\n({} out of {} iterations completed)"
                                                    .format(str(r + 1), numberofIterations))
                    QGuiApplication.processEvents()
                    if r+1 == numberofIterations:
                        self.progressWindow.close()

                except:
                    self.progressWindowtext.setText('Iterations are not completed.')

            try:
                now2 = datetime.now()
                current_time = now2.strftime("%H:%M:%S")
                print("Done at", current_time)
                ### for generatefun
                self.t = t
                self.y = y
                self.N = N
                self.Tref = Tref
                self.NL = NL
                self.NC = NC
                self.NE = NE
                self.Lref = Lref
                self.Vref = Vref
                self.I = I
                self.A = A
                self.MW = MW
                self.NV = NV
                self.Z = Z
                self.R = R
                self.g = g
                self.Ul = Ul
                self.Uv = Uv
                self.sigma = sigma
                if F > 0:
                    self.F = F
                    self.Fz = Fz
                    self.FP = FP
                    self.FT = FT
                    self.Fi = Fi
                else:
                    self.F = 0
                self.H = H
                self.P0 = P0
                self.S = S
                self.ng = ng
                self.generatefun()
            except:
                pass

    def generatefun(self):
        ### from analyse
        if self.F > 0:
            F = self.F # no of feed streams
        else:
            F = 0
        t = self.t
        y = self.y
        N = self.N
        Tref = self.Tref
        NL = self.NL
        NC = self.NC
        NE = self.NE
        Lref = self.Lref
        Vref = self.Vref
        I = self.I
        A = self.A
        MW = self.MW
        NV = self.NV
        # Z = self.Z
        R = self.R
        g = self.g
        Ul = self.Ul
        Uv = self.Uv
        sigma = self.sigma
        if F > 0:
            Fz = self.Fz
            FP = self.FP
            FT = self.FT
            Fi = self.Fi
        H = self.H
        P0 = self.P0
        S = self.S
        ng = self.ng

        # Obtain arranged list from Component Table
        arrangedVP = vapourPressureList[1:I + 1]
        arrangedVP_index = [nN_name.index(x) - 1 for x in arrangedVP]  # Because there is white space in first index

        arrangedLD = liquidDensityList[1:I + 1]
        arrangedLD_index = [nN_name.index(x) - 1 for x in arrangedLD]

        arrangedLE = enthalpyLiqList[1:I + 1]
        arrangedLE_index = [nN_name.index(x) - 1 for x in arrangedLE]

        arrangedVE = enthalpyVapList[1:I + 1]
        arrangedVE_index = [nN_name.index(x) - 1 for x in arrangedVE]

        arrangedLH = specificheatLiqList[1:I + 1]
        arrangedLH_index = [nN_name.index(x) - 1 for x in arrangedLH]

        arrangedVH = specificheatVapList[1:I + 1]
        arrangedVH_index = [nN_name.index(x) - 1 for x in arrangedVH]

        def liq_prop(T):
            E = np.zeros((len(T), I))
            C = np.zeros((len(T), I))
            LD = np.zeros((len(T), I))
            for j in range(len(T)):
                for i in range(I):
                    if len(T) > 1:
                        E[j, i] = my_module[arrangedLE_index[i]].ANN(np.atleast_2d(T[j]).T)[0]
                        C[j, i] = my_module[arrangedLH_index[i]].ANN(np.atleast_2d(T[j]).T)[0]
                        LD[j, i] = my_module[arrangedLD_index[i]].ans(T[j][0])
                    elif len(T) == 1:
                        E[j, i] = my_module[arrangedLE_index[i]].ANN([T])[0]
                        C[j, i] = my_module[arrangedLH_index[i]].ANN([T])[0]
                        LD[j, i] = my_module[arrangedLD_index[i]].ans(T[j])
            return E, C, LD

        def Vap_Pressure(T):
            P_s = np.zeros((len(T), I))
            for j in range(len(T)):
                for i in range(I):
                    P_s[j, i] = my_module[arrangedVP_index[i]].ans(T[j])
            return P_s

        ## Results
        t = t / (24 * 3600)
        sy_m = len(y)
        sy_n = len(y[0])
        T = y[:, :N] * Tref
        wsq = np.zeros((sy_m, N - 1))
        wsq[:, :NL - 1] = y[:, NC:NC + NL - 1] * 0.01 * Lref
        wsq[:, NL:N - 1] = y[:, NC + NL - 1:NE] * 0.01 * Vref
        x = np.zeros((sy_m, I, N))
        Wn = np.zeros((sy_m, I, N))
        for i in range(N):
            Wn[:, :, i] = y[:, int(N + I * i):int(N + I * i + I)]
        Wn[:, :, :NL] = Wn[:, :, :NL] * Lref
        Wn[:, :, NL:N] = Wn[:, :, NL:N] * Vref
        sumw = Wn.sum(axis=1)  # equivalent to permute
        for i in range(x.shape[2]):
            mat = Wn[:, :, i]
            summation = mat.sum(axis=1).reshape(-1, 1)
            x[:, :, i] = mat / summation

        # %% ***Density and Pressure Computation****
        rho = np.zeros((sy_m, N))
        P = np.zeros((sy_m, N))
        h = np.zeros((sy_m, 1))
        PV = np.zeros((sy_m, 1))
        Z = np.zeros((sy_m, N))
        for i in range(N):
            if i <= NL - 1:
                T_arr = T[:, i].reshape(-1, 1)
                EL, CL, LD = liq_prop(T_arr)
                rho[:, i] = 1 / np.sum(x[:, :, i] / LD, axis=1)
                if i == 0:
                    h = (NL * sumw[:, i]) / A / rho[:, i]  # liquid level
                    Mw = 1 / np.sum(x[:, :, N - 1] / MW, axis=1)
                    ## Zg and PV calculation
                    Pg = P0
                    for sy in range(sy_m):
                        P1 = np.zeros((N - NL, 1))
                        P1[:] = Pg
                        T1 = np.array(T[sy, NL:N]).reshape(-1, 1)
                        zf = x[sy, :, NL:N]  # 4 by 5 array
                        zf = np.transpose(zf)  # 5 by 4 array
                        AN = np.concatenate((zf, P1, T1), axis=1)
                        BN = my_module[nN_name.index('ModelZg') - 1].ans(AN)
                        BN = np.maximum(BN, 0)
                        BN_list = BN.reshape(1, -1).tolist()[0]
                        Z[sy, NL:N] = BN_list
                        PV[sy] = (NV * Z[sy, N - 1] * R * T[sy, N - 1] * sumw[sy, N - 1]) / (A * (H - h[sy]) * Mw[sy])
                        Pg = PV[sy]

            else:
                Mw = 1 / np.sum(x[:, :, i] / MW, axis=1)
                rho1 = PV * Mw.reshape(-1, 1)
                rho2 = Z[:, i] * R * T[:, i]
                rho2 = rho2.reshape(-1, 1)
                rho3 = rho1 / rho2
                rho3 = rho3.reshape(1, -1)[0]
                rho[:, i] = rho3
                PV_list = PV.reshape(1, -1)[0]
                P[:, i] = PV_list  # vapor phase pressure

        for i in range(NL):
            firstP = PV
            firstP = firstP.reshape(1, -1)[0]
            secondP = (g * h / NL) * (0.5 * rho[:, i] + np.sum(rho[:, i + 1:NL], axis=1)) / 1000
            ansP = firstP + secondP
            P[:, i] = ansP

        # %% *** Evaporation Flux Calculation ***
        T_I = (Ul * T[:, NL - 1] + Uv * T[:, NL]) / (Ul + Uv)  # Interface Temperature
        X = (x[:, :, NL - 1] / MW) / (np.sum(x[:, :, NL - 1] / MW, axis=1).reshape(-1, 1))
        Y = (x[:, :, NL] / MW) / (np.sum(x[:, :, NL] / MW, axis=1).reshape(-1, 1))
        Ps = np.maximum(Vap_Pressure(T_I), 0)
        K = np.maximum(Ps / PV.reshape(-1, 1), 1 * 10 ** (-6))
        J = sigma * A * PV.reshape(-1, 1) * (np.minimum(K * X, 1) - Y) * np.sqrt(MW)
        wsq[:, NL - 1] = np.sum(J, axis=1)

        P2 = np.zeros((F, 1))
        if F > 0:
            Fz = np.transpose(Fz, (1, 2, 0))
            for i in range(sy_m):
                P2[:, 0] = PV[i]
                AN = np.concatenate((Fz[:, :, i], FP[i, :].reshape(-1, 1), P2, FT[i, :].reshape(-1, 1)), axis=1)
                BN = np.maximum(my_module[nN_name.index('ModelF') - 1].ANN(AN), 0)
                T2 = BN[:, 0]
                vf = np.minimum(BN[:, 1:], Fz[:, :, i])
                fv = vf * Fi[i, 0]
                fl = Fi[i, 0] - fv
                if i == 0:
                    VFL = np.array([np.sum(fl, axis=0).sum()]).reshape(-1, 1)
                    VFV = np.array([np.sum(fv, axis=0).sum()]).reshape(-1, 1)
                    TF = T2.reshape(1, -1)
                else:
                    VFL = np.append(VFL, np.array([np.sum(fl, axis=0).sum()]).reshape(-1, 1), axis=0)
                    VFV = np.append(VFV, np.array([np.sum(fv, axis=0).sum()]).reshape(-1, 1), axis=0)
                    TF = np.append(TF, T2.reshape(1, -1), axis=0)

            VF = VFV + wsq[:, NL - 1].reshape(-1,1)

        elif F == 0:
            VF = wsq[:, NL - 1].reshape(-1,1)


        ### Compositions for Plotting
        for i in range(N):
            if i == 0:
                xc = x[:, :, i]
                Wni = Wn[:, :, i]
            else:
                xc = np.append(xc, x[:, :, i], axis=1)
                Wni = np.append(Wni, Wn[:, :, i], axis=1)

        Tl_avg = (T[:, :NL].sum(axis=1))/NL  # Average liquid temperature
        Tv_avg = (T[:, NL:N]).sum(axis=1)/NV  # Average vapour temperature
        xl = np.zeros((sy_m,I))
        xv = np.zeros((sy_m,I))
        for i in range(I):
            xl[:,i] = x[:,i,:NL].sum(axis=1)/NL
            xv[:,i] = x[:,i, NL:N].sum(axis=1)/NV

        ### Product Stream Calculations
        Ts = np.zeros((sy_m, S))
        Ps = np.zeros((sy_m, S))
        xs = np.zeros((sy_m, I, S))
        for j in range(sy_m):
            mg = np.zeros((S, 1))
            for s in range(S):
                if ng[s] > h[j]:
                    mg[s] = np.round((ng[s]-h[j])/((H-h[j])/NV))+NL
                else:
                    mg[s] = np.maximum(np.round(ng[s]/(h[j]/NL)), 1)

            pf = np.zeros((N, S))
            for s in range(S):
                if mg[s, 0] <= NL:
                    xp = [i for i in range(int(mg[s, 0]))]
                    ldf = poisson.pmf(xp, mu=1)
                    xp = [i for i in range(NL - int(mg[s, 0]))]
                    vdf = poisson.pmf(xp, mu=1)
                    for i in range(N):
                        if i <= mg[s, 0] - 1:
                            pf[i, s] = ldf[int(mg[s, 0]) - i - 1]
                        elif i > mg[s, 0] - 1 and i <= NL - 1:
                            pf[i, s] = vdf[i - int(mg[s, 0])]
                        else:
                            pf[i, s] = 0

                else:
                    xp = [i for i in range(int(mg[s, 0]) - NL)]
                    ldf = poisson.pmf(xp, mu=1)
                    xp = [i for i in range(N - int(mg[s, 0] + 1))]
                    vdf = poisson.pmf(xp, mu=1)
                    for i in range(N):
                        if i <= NL - 1:
                            pf[i, s] = 0
                        elif i > NL - 1 and i <= mg[s, 0] - 1:
                            pf[i, s] = ldf[int(mg[s, 0]) - i - 1]
                        else:
                            pf[i, s] = vdf[i - int(mg[s, 0])]

            pf[:NL, :] = pf[:NL, :] / pf[:NL, :].sum(axis=0)
            pf[NL:N, :] = pf[NL:N, :] / pf[NL:N, :].sum(axis=0)
            pf = np.maximum(pf, 0)
            pf = np.nan_to_num(pf)

            Ts[j, :] = np.matmul(T[j, :], pf)
            Ps[j, :] = np.matmul(P[j, :], pf)

            xtemp = np.transpose(x, (1,2,0))
            xtemp = xtemp[:,:,j]
            xs[j,:,:] = np.matmul(xtemp, pf)

        ### For plotting
        self.t = t
        self.T = T
        self.PV = PV
        self.h = h
        self.nlist = [i + 1 for i in range(N)]  ## Because counting starts from 0.
        self.wsqN = wsq[:, NL - 1]
        self.NL = NL
        self.xc = xc
        self.T_I = T_I
        self.VF = VF
        self.Tl_avg = Tl_avg
        self.Tv_avg = Tv_avg
        if F > 0:
            self.VFV = VFV
        self.xl = xl
        self.xv = xv
        self.Ts = Ts
        self.Ps = Ps
        self.P = P
        self.Ts = Ts
        self.Ps = Ps
        self.xs = xs
        self.S = S
        self.resultsButton.setEnabled(True)
        self.resultsfun()
        self.plot1fun()
        print("Ready to Plot.")

    def resultsfun(self):
        self.designButton.setStyleSheet(
            "QPushButton#designButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#designButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.heatLossButton.setStyleSheet(
            "QPushButton#heatLossButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#heatLossButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.initialButton.setStyleSheet(
            "QPushButton#initialButton{background-color: rgb(230, 255, 255); color: rgb(0, 0, 0); border: 2px solid black}"
            "QPushButton#initialButton::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black}")
        self.resultsButton.setStyleSheet(
            "QPushButton#resultsButton{background-color: rgb(255, 0, 0); color: rgb(255, 255, 255); border: 2px solid black}")
        self.designlabel.hide()
        self.heatlosslabel.hide()
        self.initiallabel.hide()
        self.complabel.hide()
        self.resultslabel.show()

    def plot1fun(self):
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot2button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        t = self.t * 24 * 60
        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()
        self.ax.plot(t, self.PV)
        self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax.minorticks_on()
        self.ax.set_xlim(xmin=0)
        self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax.set_ylabel('Pressure (kPa)', fontname="Arial", fontsize=14)
        self.ax.set_title('Pressure Profile', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        self.plotting.draw()

    def plot2fun(self):
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        t = self.t * 24 * 60
        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()
        self.ax.plot(t, self.h)
        self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax.minorticks_on()
        self.ax.set_xlim(xmin=0)
        self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax.set_ylabel('Liquid level (m)', fontname="Arial", fontsize=14)
        self.ax.set_title('Liquid Level Profile', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        self.plotting.draw()

    def plot3fun(self):
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        t = self.t * 24 * 60
        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()
        self.ax.plot(t, self.Tl_avg)
        # self.ax.plot(t, self.Tv_avg)
        self.ax.plot(t, self.T_I)

        legend = ['Avg. Liquid Temperature', 'Interface Temperature']

        for i in range(self.NL,self.N):
            self.ax.plot(t, self.T[:, i])
            legend.append('Disk {} Temperature'.format(str(i+1)))

        self.ax.legend(legend)
        self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        self.ax.minorticks_on()
        self.ax.set_xlim(xmin=0)
        self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax.set_ylabel('Temperature (K)', fontname="Arial", fontsize=14)
        self.ax.set_title('Temperature Profiles', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        self.plotting.draw()

    def plot4fun(self):
        t2 = [self.t[i,0] * 24 * 60 for i in range(len(self.t))]
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()
        self.ax.plot(t2, self.wsqN)
        self.ax.plot(t2, self.VF)
        self.ax.legend(['Net Evaporation Rate', 'Net Vapour Addition Rate'])
        if self.F > 0:
            self.ax.plot(t2, self.VFV)
            self.ax.legend(['Net Evaporation Rate', 'Net Vapour Addition Rate', 'Feed Vapour Addition Rate'])
        self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax.minorticks_on()
        self.ax.set_xlim(xmin=0)
        self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax.set_ylabel('Rate (kg/s)', fontname="Arial", fontsize=14)
        self.ax.set_title('Vapour Flows', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)

        self.plotting.draw()

    def plot5fun(self):
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()

        t = self.t.reshape(1,-1)[0]
        t = t * 24 * 60
        for i in range(self.I):
            self.ax.plot(t, self.xl[:,i])

        legend = ['Component {}'.format(str(i+1)) for i in range(self.I)]
        self.ax.legend(legend)

        self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax.minorticks_on()
        self.ax.set_xlim(xmin=0)
        self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax.set_ylabel('Component Mass Fraction', fontname="Arial", fontsize=14)
        self.ax.set_title('Average Liquid Composition', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)

        self.plotting.draw()

    def plot6fun(self):
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()

        t = self.t.reshape(1, -1)[0]
        t = t * 24 * 60
        for i in range(self.I):
            self.ax.plot(t, self.xv[:, i])

        legend = ['Component {}'.format(str(i + 1)) for i in range(self.I)]
        self.ax.legend(legend)

        self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax.minorticks_on()
        self.ax.set_xlim(xmin=0)
        self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax.set_ylabel('Component Mass Fraction', fontname="Arial", fontsize=14)
        self.ax.set_title('Average Vapour Composition', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        self.plotting.draw()

    def plot7fun(self):
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot8button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()

        if self.S > 0:
            t2 = [self.t[i,0] * 24 * 60 for i in range(len(self.t))]
            self.ax.plot(t2, self.Ts)
            legend = ['Product {} Temperature'.format(str(i+1)) for i in range(len(self.Ts[1]))]
            self.ax.legend(legend)
            self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
            self.ax.minorticks_on()
            self.ax.set_xlim(xmin=0)
            self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
            self.ax.set_ylabel('Temperature (K)', fontname="Arial", fontsize=14)
            self.ax.set_title('Product Temperature', fontname="Arial", fontsize=16)
            for tick in ax.get_xticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)
            for tick in ax.get_yticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)

        self.plotting.draw()

    def plot8fun(self):
        self.plot1button.setStyleSheet(
            "QPushButton#plot1button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot1button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot2button.setStyleSheet(
            "QPushButton#plot2button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot3button.setStyleSheet(
            "QPushButton#plot3button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot3button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot4button.setStyleSheet(
            "QPushButton#plot4button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot4button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot5button.setStyleSheet(
            "QPushButton#plot5button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot5button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot6button.setStyleSheet(
            "QPushButton#plot6button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot6button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot7button.setStyleSheet(
            "QPushButton#plot7button{background-color: rgb(0, 255, 0); color: rgb(0, 0, 0); border: 2px solid black; border-radius: 5px}"
            "QPushButton#plot7button::hover{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")
        self.plot8button.setStyleSheet(
            "QPushButton#plot8button{background-color: rgb(0, 0, 255); color: rgb(255, 255, 255); border: 2px solid black; border-radius: 5px}")

        self.plot_widget.show()
        for ax in self.fig.axes:
            ax.clear()

        if self.S > 0:
            t2 = [self.t[i, 0] * 24 * 60 for i in range(len(self.t))]
            self.ax.plot(t2, self.Ps)
            legend = ['Product {} Pressure'.format(str(i + 1)) for i in range(len(self.Ts[1]))]
            self.ax.legend(legend)
            self.ax.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
            self.ax.minorticks_on()
            self.ax.set_xlim(xmin=0)
            self.ax.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
            self.ax.set_ylabel('Pressure (kPa)', fontname="Arial", fontsize=14)
            self.ax.set_title('Product Pressure', fontname="Arial", fontsize=16)
            for tick in ax.get_xticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)
            for tick in ax.get_yticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)

        self.plotting.draw()


    def resultsExcel(self):
        now = datetime.now()
        current_time = now.strftime("%d-%m-%Y %H%M")
        nameResultsExcel = "Results {}.xlsx".format(current_time)
        current_path = os.getcwd()
        full_path = current_path + '\\Results\\'
        nameResultsExcel = full_path + nameResultsExcel

        wb = openpyxl.Workbook()
        wb.create_sheet("Disk Temperature")
        ws = wb["Disk Temperature"]
        diskName = ["Disk {}".format(str(i+1)) for i in range(self.noDisks)]
        for r in range(len(diskName)):
            ws.cell(1,r+1).value = diskName[r]
        for r in range(len(self.T)):
            for s in range(len(self.T[0])):
                ws.cell(r+2,s+1).value = self.T[r,s]
        wb.save(nameResultsExcel)

        wb.create_sheet("Liquid Level")
        ws = wb["Liquid Level"]
        ws.cell(1,1).value = "h (m)"
        for r in range(len(self.h)):
            ws.cell(r+2, 1).value = self.h[r]
        wb.save(nameResultsExcel)

        wb.create_sheet("Time (min)")
        ws = wb["Time (min)"]
        ws.cell(1,1).value = "Time (min)"
        for r in range(len(self.t)):
            ws.cell(r+2, 1).value = self.t[r,0] * 24 * 60
        wb.save(nameResultsExcel)

        wb.create_sheet("Disk Composition")
        ws = wb["Disk Composition"]
        for r in range(len(self.xc)):
            for s in range(len(self.xc[0])):
                ws.cell(r+1, s+1).value = self.xc[r,s]
        wb.save(nameResultsExcel)

        wb.create_sheet("Interface Temperature")
        ws = wb["Interface Temperature"]
        ws.cell(1,1).value = "Temperature (K)"
        for r in range(len(self.T_I)):
            ws.cell(r+2,1).value = self.T_I[r]
        wb.save(nameResultsExcel)

        wb.create_sheet("Tank Pressure")
        ws = wb["Tank Pressure"]
        ws.cell(1,1).value = "Pressure (kPa)"
        for r in range(len(self.PV)):
            ws.cell(r+2,1).value = self.PV[r,0]
        wb.save(nameResultsExcel)

        wb.create_sheet("wsq")
        ws = wb["wsq"]
        ws.cell(1,1).value = "Evaporation/Condensation Rate (kg/s)"
        for r in range(len(self.wsqN)):
            ws.cell(r+2, 1).value = self.wsqN[r]
        wb.save(nameResultsExcel)

        wb.create_sheet("VF")
        ws = wb["VF"]
        ws.cell(1, 1).value = "Net Vapour Addition Rate (kg/s)"
        for r in range(len(self.VF)):
            ws.cell(r + 2, 1).value = self.VF[r,0]
        wb.save(nameResultsExcel)

        wb.create_sheet("Disk Pressure")
        ws = wb["Disk Pressure"]
        diskName = ["Disk {}".format(str(i + 1)) for i in range(self.noDisks)]
        for r in range(len(diskName)):
            ws.cell(1, r + 1).value = diskName[r]
        for r in range(len(self.P)):
            for s in range(len(self.T[0])):
                ws.cell(r + 2, s + 1).value = self.P[r, s]
        wb.save(nameResultsExcel)

        wb.create_sheet("Product Temperature")
        ws = wb["Product Temperature"]
        pdtName = ["Stream {}".format(str(i + 1)) for i in range(len(self.Ts[1]))]
        for r in range(len(pdtName)):
            ws.cell(1, r + 1).value = pdtName[r]
        for r in range(len(self.Ts)):
            for s in range(len(self.Ts[0])):
                ws.cell(r + 2, s + 1).value = self.Ts[r, s]
        wb.save(nameResultsExcel)

        wb.create_sheet("Product Pressure")
        ws = wb["Product Pressure"]
        pdtName = ["Stream {}".format(str(i + 1)) for i in range(len(self.Ps[1]))]
        for r in range(len(pdtName)):
            ws.cell(1, r + 1).value = pdtName[r]
        for r in range(len(self.Ps)):
            for s in range(len(self.Ps[0])):
                ws.cell(r + 2, s + 1).value = self.Ps[r, s]
        wb.save(nameResultsExcel)

        for i in range(self.S):
            wb.create_sheet("Pdt{} Comp".format(str(i+1)))
            ws = wb["Pdt{} Comp".format(str(i+1))]
            array = self.xs[:,:,i]
            for r in range(len(array)):
                for s in range(len(array[0])):
                    ws.cell(r + 1, s + 1).value = array[r, s]
            wb.save(nameResultsExcel)

    def resultsGraph(self):
        now = datetime.now()
        now_string = now.strftime("%d%m%Y_%H%Mh")
        directory = "Graphs " + now_string
        current_path = os.getcwd()
        current_path = current_path + "/Results/Graphs"
        path = os.path.join(current_path, directory)
        os.mkdir(path)

        ## Plot 1
        t = self.t * 24 * 60
        for ax in self.fig2.axes:
            ax.clear()
        self.ax2.plot(t, self.PV)
        self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax2.minorticks_on()
        self.ax2.set_xlim(xmin=0)
        self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax2.set_ylabel('Pressure (kPa)', fontname="Arial", fontsize=14)
        self.ax2.set_title('Pressure Profile', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)

        fig_dir1 = path + "\\Pressure"
        self.fig2.savefig(fig_dir1)

        ## Plot 2
        t = self.t * 24 * 60
        #self.plot_widget2.show()
        for ax in self.fig2.axes:
            ax.clear()
        self.ax2.plot(t, self.h)
        self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax2.minorticks_on()
        self.ax2.set_xlim(xmin=0)
        self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax2.set_ylabel('Liquid level (m)', fontname="Arial", fontsize=14)
        self.ax2.set_title('Liquid Level Profile', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        fig_dir2 = path + "\\Liquid_Level"
        self.fig2.savefig(fig_dir2)

        ## Plot 3
        t = self.t * 24 * 60
        #self.plot_widget.show()
        for ax in self.fig2.axes:
            ax.clear()
        self.ax2.plot(t, self.Tl_avg)
        self.ax2.plot(t, self.T_I)

        legend = ['Avg. Liquid Temperature', 'Interface Temperature']

        for i in range(self.NL, self.N):
            self.ax2.plot(t, self.T[:, i])
            legend.append('Disk {} Temperature'.format(str(i + 1)))

        self.ax2.legend(legend)
        self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        self.ax2.minorticks_on()
        self.ax2.set_xlim(xmin=0)
        self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax2.set_ylabel('Temperature (K)', fontname="Arial", fontsize=14)
        self.ax2.set_title('Temperature Profiles', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        fig_dir3 = path + "\\Temperature_Profiles"
        self.fig2.savefig(fig_dir3)

        ## Plot 4
        t2 = [self.t[i,0] * 24 * 60 for i in range(len(self.t))]
        #self.plot_widget2.show()
        for ax in self.fig2.axes:
            ax.clear()
        self.ax2.plot(t2, self.wsqN)
        self.ax2.plot(t2, self.VF)
        self.ax2.legend(['Net Evaporation Rate', 'Feed Vapour Addition Rate'])
        if self.F > 0:
            self.ax2.plot(t2, self.VFV)
            self.ax2.legend(['Net Evaporation Rate', 'Feed Vapour Addition Rate', 'Net Vapour Addition Rate'])
        self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
        self.ax2.minorticks_on()
        self.ax2.set_xlim(xmin=0)
        self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax2.set_ylabel('Rate (kg/s)', fontname="Arial", fontsize=14)
        self.ax2.set_title('Vapour Flows', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        fig_dir4 = path + "\\Vapour_Flows"
        self.fig2.savefig(fig_dir4)

        ## Plot 5
        #self.plot_widget.show()
        for ax in self.fig2.axes:
            ax.clear()
        t = self.t.reshape(1, -1)[0]
        t = t * 24 * 60
        for i in range(self.I):
            self.ax2.plot(t, self.xl[:, i])

        legend = ['Component {}'.format(str(i + 1)) for i in range(self.I)]
        self.ax2.legend(legend)

        self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax2.minorticks_on()
        self.ax2.set_xlim(xmin=0)
        self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax2.set_ylabel('Component Mass Fraction', fontname="Arial", fontsize=14)
        self.ax2.set_title('Average Liquid Composition', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        fig_dir5 = path + "\\Liquid Composition"
        self.fig2.savefig(fig_dir5)

        ## Plot 6
        #self.plot_widget2.show()
        for ax in self.fig2.axes:
            ax.clear()
        t = self.t.reshape(1, -1)[0]
        t = t * 24 * 60
        for i in range(self.I):
            self.ax2.plot(t, self.xv[:, i])

        legend = ['Component {}'.format(str(i + 1)) for i in range(self.I)]
        self.ax2.legend(legend)

        self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))

        self.ax2.minorticks_on()
        self.ax2.set_xlim(xmin=0)
        self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
        self.ax2.set_ylabel('Component Mass Fraction', fontname="Arial", fontsize=14)
        self.ax2.set_title('Average Vapour Composition', fontname="Arial", fontsize=16)
        for tick in ax.get_xticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        for tick in ax.get_yticklabels():
            tick.set_fontname("Arial")
            tick.set_fontsize(12)
        fig_dir6 = path + "\\Vapour Composition"
        self.fig2.savefig(fig_dir6)

        ## Plot 7
        if self.S > 0:
            #self.plot_widget2.show()
            for ax in self.fig2.axes:
                ax.clear()

            t2 = [self.t[i, 0] * 24 * 60 for i in range(len(self.t))]
            self.ax2.plot(t2, self.Ts)
            legend = ['Product {} Temperature'.format(str(i + 1)) for i in range(len(self.Ts[1]))]
            self.ax2.legend(legend)
            self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
            self.ax2.minorticks_on()
            self.ax2.set_xlim(xmin=0)
            self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
            self.ax2.set_ylabel('Temperature (K)', fontname="Arial", fontsize=14)
            self.ax2.set_title('Product Temperature', fontname="Arial", fontsize=16)
            for tick in ax.get_xticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)
            for tick in ax.get_yticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)
            fig_dir7 = path + "\\Product Temperature"
            self.fig2.savefig(fig_dir7)

        ## Plot 8
        if self.S > 0:
            #self.plot_widget2.show()
            for ax in self.fig2.axes:
                ax.clear()

            t2 = [self.t[i, 0] * 24 * 60 for i in range(len(self.t))]
            self.ax2.plot(t2, self.Ps)
            legend = ['Product {} Pressure'.format(str(i + 1)) for i in range(len(self.Ts[1]))]
            self.ax2.legend(legend)
            self.ax2.yaxis.set_major_formatter(FormatStrFormatter('%.2f'))
            self.ax2.minorticks_on()
            self.ax2.set_xlim(xmin=0)
            self.ax2.set_xlabel('Time (min)', fontname="Arial", fontsize=14)
            self.ax2.set_ylabel('Pressure (kPa)', fontname="Arial", fontsize=14)
            self.ax2.set_title('Product Pressure', fontname="Arial", fontsize=16)
            for tick in ax.get_xticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)
            for tick in ax.get_yticklabels():
                tick.set_fontname("Arial")
                tick.set_fontsize(12)
            fig_dir8 = path + "\\Product Pressure"
            self.fig2.savefig(fig_dir8)
        print("Plotted")

    def saveinputfun(self):
        ### Base Case.xlsx
        baseCaseExcel = "Base Case.xlsx"
        current_path = os.getcwd()
        base_case_full_path = current_path + '\\' + baseCaseExcel

        ## Feed stream parameters
        F = 0  # Number of Feed Streams
        self.feedLocationList = []

        def isfloat(value):
            try:
                float(value)
                return True
            except ValueError:
                return False

        for key in self.feedLocationValueDict:
            if isfloat(self.feedLocationValueDict[key].text()) and float(
                    self.feedLocationValueDict[key].text()) <= 100 and \
                    float(self.feedLocationValueDict[key].text()) >= 0:
                self.feedLocationList.append(self.feedLocationValueDict[key].text())
                F += 1

        ## Product stream parameters
        S = 0  # total number of product streams
        self.productLocationList = []
        for key in self.productLocationValueDict:
            if isfloat(self.productLocationValueDict[key].text()) and float(
                    self.productLocationValueDict[key].text()) <= 100 and \
                    float(self.productLocationValueDict[key].text()) >= 0:
                self.productLocationList.append(self.productLocationValueDict[key].text())
                S += 1

        ## Save to Design Sheet
        wb = openpyxl.load_workbook(base_case_full_path)
        ws = wb["Design"]
        ws.cell(1, 2).value = self.tankDiametertext.text()
        ws.cell(2, 2).value = self.tankHeighttext.text()
        ws.cell(3, 2).value = self.initialPressuretext.text()
        ws.cell(4, 2).value = self.initialHeighttext.text()
        ws.cell(5, 2).value = self.noComponentstext.currentText()
        ws.cell(6, 2).value = self.jacketStartValue.text()
        ws.cell(7, 2).value = self.jacketEndValue.text()

        for i in range(len(self.feedLocationList)):
            ws.cell(13+i, 2).value = self.feedLocationList[i]

        for j in range(len(self.productLocationList)):
            ws.cell(20+j, 2).value = self.productLocationList[j]

        wb.save(base_case_full_path)

        ## Save to Heat Loss Sheet
        wb = openpyxl.load_workbook(base_case_full_path)
        ws = wb["Heat Loss"]
        ws.cell(1, 2).value = self.sigmaText.text()
        ws.cell(2, 2).value = self.Ultext.text()
        ws.cell(3, 2).value = self.Uvtext.text()
        ws.cell(4, 2).value = self.Uvwtext.text()
        ws.cell(5, 2).value = self.Ulwtext.text()
        ws.cell(6, 2).value = self.Urtext.text()
        ws.cell(7, 2).value = self.Ubtext.text()
        ws.cell(8, 2).value = self.groundTemptext.text()
        ws.cell(9, 2).value = self.ambTemptext.text()
        ws.cell(10, 2).value = self.roofTemptext.text()
        ws.cell(11, 2).value = self.Uvrtext.text()
        ws.cell(12, 2).value = self.Ulrtext.text()
        ws.cell(13, 2).value = self.refridgeTemptext.text()
        wb.save(base_case_full_path)

        ## Save to Initial Sheet
        wb = openpyxl.load_workbook(base_case_full_path)
        ws = wb["Initial"]
        ws.cell(1, 2).value = self.noLDiskstext.text()
        ws.cell(2, 2).value = self.noVDiskstext.text()
        ws.cell(3, 2).value = self.abstoltext.text()
        ws.cell(4, 2).value = self.reltoltext.text()
        ws.cell(5, 2).value = self.numberofIterationstext.text()
        wb.save(base_case_full_path)

        print(self.sigmaText.text())


if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)

    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    MainWindow.show()
    sys.exit(app.exec_())